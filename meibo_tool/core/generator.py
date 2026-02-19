"""テンプレートへのデータ差込処理

SPEC.md 第5章 §5.1〜§5.9 参照。
3 種類のジェネレーター + 共通プレースホルダー置換。
"""

from __future__ import annotations

import os
import re
import shutil
from abc import ABC, abstractmethod
from copy import copy
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image

from templates.template_registry import TEMPLATES
from utils.address import build_address
from utils.date_fmt import DATE_KEYS, format_date
from utils.font_helper import apply_font
from utils.wareki import to_wareki

PLACEHOLDER_RE = re.compile(r'\{\{(.+?)\}\}')

# name_display モードで使用するフィールドキーセット
# kanji モード: かな系フィールドを空白化
# kana   モード: かな値を漢字フィールドに転写し、かなフィールドは空白化
_NAME_KANJI_KEYS: frozenset[str] = frozenset({'氏名', '正式氏名'})
_NAME_KANA_KEYS:  frozenset[str] = frozenset({'氏名かな', '正式氏名かな'})

# kana モード: 漢字キー → 対応するかなキーへのマッピング
_KANJI_TO_KANA: dict[str, str] = {
    '氏名': '氏名かな',
    '正式氏名': '正式氏名かな',
}


# ────────────────────────────────────────────────────────────────────────────
# 共通値解決
# ────────────────────────────────────────────────────────────────────────────

def _resolve_value(key: str, data_row: dict, mode: str) -> str:
    """name_display モードを考慮してデータ行からフィールド値を取得する。

    fill_placeholders / _fill_numbered 共通のロジック。
    """
    original_key = key
    if mode == 'kanji' and key in _NAME_KANA_KEYS:
        return ''
    if mode == 'kana':
        if key in _NAME_KANA_KEYS:
            return ''  # かな行セルを空白化
        if key in _NAME_KANJI_KEYS:
            key = _KANJI_TO_KANA[key]  # 漢字フィールドにかな値を転写
    v = data_row.get(key, '')
    if v is None:
        return ''
    s = str(v).strip()
    if s.lower() == 'nan':
        return ''
    # 日付型フィールドは YY/MM/DD にフォーマット
    if original_key in DATE_KEYS:
        return format_date(s)
    return s


# ────────────────────────────────────────────────────────────────────────────
# プレースホルダー置換（共通）
# ────────────────────────────────────────────────────────────────────────────

def fill_placeholders(ws, data_row: dict | pd.Series, options: dict | None = None) -> None:
    """
    ワークシート内の全プレースホルダー {{...}} をデータで置換する。

    特殊キー:
        {{年度}}       → options['fiscal_year']
        {{年度和暦}}   → 和暦表記の年度
        {{学校名}}     → options['school_name']
        {{担任名}}     → options['teacher_name']
        {{住所}}       → 都道府県+市区町村+町番地+建物名 の結合

    name_display オプション（options['name_display']）:
        'furigana' (デフォルト) → {{氏名}} と {{氏名かな}} の両方を展開
        'kanji'                 → {{氏名かな}} / {{正式氏名かな}} を空白化
        'kana'                  → {{氏名}} に氏名かな値を転写、{{氏名かな}} を空白化
                                   （小さいかな行は空白、大きい名前行にかなを表示）
    """
    options = options or {}
    mode = options.get('name_display', 'furigana')

    def replacer(match: re.Match) -> str:
        key = match.group(1)
        if key == '年度':
            return str(options.get('fiscal_year', ''))
        if key == '年度和暦':
            fy = options.get('fiscal_year', 2025)
            return to_wareki(fy, 4, 1).replace('年', '年度')
        if key == '学校名':
            return options.get('school_name', '')
        if key == '担任名':
            return options.get('teacher_name', '')
        if key == '住所':
            return build_address(data_row)
        return _resolve_value(key, data_row, mode)

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                cell.value = PLACEHOLDER_RE.sub(replacer, cell.value)


def _fill_row_placeholders(
    ws, row_num: int, data_row: dict, options: dict,
) -> None:
    """特定の 1 行のみプレースホルダーを置換する（ListGenerator 用）。"""
    mode = options.get('name_display', 'furigana')

    def replacer(match: re.Match) -> str:
        key = match.group(1)
        if key == '年度':
            return str(options.get('fiscal_year', ''))
        if key == '年度和暦':
            fy = options.get('fiscal_year', 2025)
            return to_wareki(fy, 4, 1).replace('年', '年度')
        if key == '学校名':
            return options.get('school_name', '')
        if key == '担任名':
            return options.get('teacher_name', '')
        if key == '住所':
            return build_address(data_row)
        return _resolve_value(key, data_row, mode)

    for cell in ws[row_num]:
        if isinstance(cell, MergedCell):
            continue
        if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
            cell.value = PLACEHOLDER_RE.sub(replacer, cell.value)


# ────────────────────────────────────────────────────────────────────────────
# 画像付きシート複製
# ────────────────────────────────────────────────────────────────────────────

def copy_sheet_with_images(wb, source_ws, new_title: str):
    """
    画像も含めてワークシートを複製する。
    openpyxl の copy_worksheet は画像をコピーしないため手動で再挿入する。

    落とし穴: Image(img.ref) は動作しない（ref は内部参照であり
    ファイルパスではない）。_data() でバイナリを取得し BytesIO 経由で
    新しい Image を生成する。
    """
    target_ws = wb.copy_worksheet(source_ws)
    target_ws.title = new_title
    for img in source_ws._images:
        img_data = BytesIO(img._data())
        new_img = Image(img_data)
        new_img.anchor = copy(img.anchor)
        new_img.width = img.width
        new_img.height = img.height
        target_ws.add_image(new_img)
    return target_ws


# ────────────────────────────────────────────────────────────────────────────
# 印刷設定
# ────────────────────────────────────────────────────────────────────────────

def setup_print(ws, orientation: str = 'portrait', paper_size: int = 9) -> None:
    """
    印刷設定を適用する。
    paper_size: 9=A4
    orientation: 'portrait' or 'landscape'

    落とし穴: ws.page_setup.fitToPage = True は sheet_properties.pageSetUpPr が
    None のとき AttributeError になる。PageSetupProperties で初期化してから設定する。
    """
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.page_setup.paperSize = paper_size
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # fitToPage は sheet_properties 経由で設定しないと AttributeError になる
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins = PageMargins(
        left=0.39, right=0.39,
        top=0.39, bottom=0.39,
        header=0.2, footer=0.2,
    )
    ws.print_options.horizontalCentered = True


# ────────────────────────────────────────────────────────────────────────────
# ベースクラス
# ────────────────────────────────────────────────────────────────────────────

class BaseGenerator(ABC):
    def __init__(
        self,
        template_path: str,
        output_path: str,
        data: pd.DataFrame,
        options: dict,
    ) -> None:
        self.template_path = template_path
        self.output_path = output_path
        self.data = data
        self.options = options
        self.wb = None

    def generate(self) -> str:
        """テンプレートにデータを差込み、output_path に保存して返す。"""
        # テンプレートをコピーしてから開く（元ファイル保護＋画像消失防止）
        shutil.copy2(self.template_path, self.output_path)
        self.wb = load_workbook(self.output_path)
        try:
            self._populate()
            self._apply_font_all()
            self.wb.save(self.output_path)
        finally:
            self.wb.close()
        return self.output_path

    def _apply_font_all(self) -> None:
        for ws in self.wb.worksheets:
            apply_font(ws)

    def _get_template_meta(self) -> dict:
        """TEMPLATES レジストリからテンプレートファイル名でメタデータを検索する。"""
        for _name, meta in TEMPLATES.items():
            if meta['file'] == os.path.basename(self.template_path):
                return meta
        return {}

    @abstractmethod
    def _populate(self) -> None:
        pass


# ────────────────────────────────────────────────────────────────────────────
# GridGenerator — 名札系（1 ページに N 名をグリッド配置）
# ────────────────────────────────────────────────────────────────────────────

class GridGenerator(BaseGenerator):
    """
    番号付きプレースホルダー（{{氏名_1}} 等）で構成されたグリッドテンプレートを処理する。

    cards_per_page が設定されている場合はページネーションを行う:
      - cards_per_page 以下の人数 → 1 ページで完結（名列表系）
      - cards_per_page を超える人数 → テンプレートシートを複製してページ追加（名札系）

    ページネーション手順:
      1. 必要なページ数分のシートを先にコピー（fill 前に複製しないと fill 済みデータが混入する）
      2. 各シートを対応する児童範囲で fill
    """

    def _populate(self) -> None:
        template_ws = self.wb.active
        meta = self._get_template_meta()
        orientation = meta.get('orientation', 'portrait')
        setup_print(template_ws, orientation=orientation)

        n = len(self.data)
        if n == 0:
            return

        cards_per_page = meta.get('cards_per_page')
        multi_page = cards_per_page is not None and n > cards_per_page

        if not multi_page:
            # ── 単一ページモード（名列表・調べ表など）────────────────────────
            for i, (_, row) in enumerate(self.data.iterrows(), 1):
                _fill_numbered(template_ws, i, row.to_dict(), self.options)
            fill_placeholders(template_ws, self.data.iloc[0].to_dict(), self.options)
        else:
            # ── 複数ページモード（名札など）───────────────────────────────────
            # Step 1: 追加シートを先に複製（fill 前に行う）
            num_pages = (n + cards_per_page - 1) // cards_per_page
            pages = [template_ws]
            for _ in range(1, num_pages):
                new_ws = self.wb.copy_worksheet(template_ws)
                setup_print(new_ws, orientation=orientation)
                pages.append(new_ws)

            # Step 2: 各シートに対応する児童範囲をfill
            for page_num, page_ws in enumerate(pages):
                start = page_num * cards_per_page
                page_data = self.data.iloc[start: start + cards_per_page]
                for slot, (_, row) in enumerate(page_data.iterrows(), 1):
                    _fill_numbered(page_ws, slot, row.to_dict(), self.options)
                fill_placeholders(page_ws, page_data.iloc[0].to_dict(), self.options)


def _fill_numbered(ws, idx: int, data_row: dict, options: dict) -> None:
    """{{氏名_1}}, {{氏名_2}} ... 形式の番号付きプレースホルダーを置換する。

    name_display モード（options['name_display']）に応じて
    氏名系フィールドを選択的に空白化する（fill_placeholders と同一仕様）。
    """
    mode = options.get('name_display', 'furigana')
    suffix = f'_{idx}'

    def replacer(match: re.Match) -> str:
        key = match.group(1)
        if key.endswith(suffix):
            base_key = key[: -len(suffix)]
            if base_key == '住所':
                return build_address(data_row)
            return _resolve_value(base_key, data_row, mode)
        return match.group(0)  # 他のインデックスは変更しない

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                cell.value = PLACEHOLDER_RE.sub(replacer, cell.value)


# ────────────────────────────────────────────────────────────────────────────
# ListGenerator — 名列表・台帳系（データ行を児童数分コピー展開）
# ────────────────────────────────────────────────────────────────────────────

class ListGenerator(BaseGenerator):
    """
    テンプレート内の {{行開始}} 〜 {{行終了}} マーカーで囲まれた行を
    児童数分コピー展開する。
    マーカーがない場合は最初の {{氏名}} が含まれる行をテンプレート行として扱う。
    """

    def _populate(self) -> None:
        ws = self.wb.active
        meta = self._get_template_meta()
        orientation = meta.get('orientation', 'portrait')
        setup_print(ws, orientation=orientation)

        # ヘッダー行・フッター行を特定してデータ行を展開
        template_row = self._find_template_row(ws)
        if template_row is None:
            return

        self._expand_rows(ws, template_row)

        # ヘッダーのプレースホルダー（年度・学校名等）を置換
        first_row = self.data.iloc[0].to_dict() if len(self.data) > 0 else {}
        fill_placeholders(ws, first_row, self.options)

    def _find_template_row(self, ws) -> int | None:
        """氏名系プレースホルダーを含む最初の行番号を返す。"""
        name_pats = ('{{氏名}}', '{{正式氏名}}', '{{氏名かな}}', '{{正式氏名かな}}')
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and any(p in cell.value for p in name_pats):
                    return cell.row
        return None

    def _expand_rows(self, ws, template_row: int) -> None:
        """テンプレート行を児童数分コピーして展開する。"""
        n = len(self.data)
        if n == 0:
            return

        # テンプレート行のセル情報を保存
        tmpl_cells = []
        for cell in ws[template_row]:
            tmpl_cells.append({
                'value': cell.value,
                'font': copy(cell.font),
                'alignment': copy(cell.alignment),
                'border': copy(cell.border),
                'fill': copy(cell.fill),
                'number_format': cell.number_format,
            })

        row_height = ws.row_dimensions[template_row].height

        # 既存行を 1 行だけ利用し、残りを挿入
        for i, (_, data_row) in enumerate(self.data.iterrows()):
            row_num = template_row + i
            if i > 0:
                ws.insert_rows(row_num)

            for col_idx, tmpl in enumerate(tmpl_cells, 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = tmpl['value']
                cell.font = tmpl['font']
                cell.alignment = tmpl['alignment']
                cell.border = tmpl['border']
                cell.fill = tmpl['fill']
                cell.number_format = tmpl['number_format']

            ws.row_dimensions[row_num].height = row_height

            # 当該行のみプレースホルダーを置換（全行スキャンを回避）
            row_dict = data_row.to_dict()
            _fill_row_placeholders(ws, row_num, row_dict, self.options)


# ────────────────────────────────────────────────────────────────────────────
# IndividualGenerator — 個票系（1 名/シートで複製）
# ────────────────────────────────────────────────────────────────────────────

class IndividualGenerator(BaseGenerator):
    """
    テンプレートの 1 枚目シートを児童数分 copy_worksheet で複製する。
    画像は copy_sheet_with_images で再挿入する。
    """

    def _populate(self) -> None:
        template_ws = self.wb.active
        meta = self._get_template_meta()
        orientation = meta.get('orientation', 'portrait')

        if len(self.data) == 0:
            return

        # Step 1: 先に全シートを複製（fill 前に行う — 置換済みデータの混入防止）
        sheets = []
        for i, (_, row) in enumerate(self.data.iterrows()):
            row_dict = row.to_dict()
            shusseki = row_dict.get('出席番号', str(i + 1))
            name = row_dict.get('氏名', row_dict.get('正式氏名', str(i + 1)))
            title = f'{str(shusseki).zfill(2)}_{name}'[:31]  # シート名 31 文字制限

            if i == 0:
                ws = template_ws
                ws.title = title
            else:
                ws = copy_sheet_with_images(self.wb, template_ws, title)
            sheets.append((ws, row_dict))

        # Step 2: 各シートにデータを差し込み
        for ws, row_dict in sheets:
            fill_placeholders(ws, row_dict, self.options)
            setup_print(ws, orientation=orientation)


# ────────────────────────────────────────────────────────────────────────────
# ファクトリー関数
# ────────────────────────────────────────────────────────────────────────────

def create_generator(
    template_name: str,
    output_path: str,
    data: pd.DataFrame,
    options: dict,
) -> BaseGenerator:
    """
    テンプレート名から適切なジェネレーターを返すファクトリー関数。

    options に必要なキー:
        template_dir  : テンプレートフォルダの絶対パス
        fiscal_year   : 年度（整数）
        school_name   : 学校名
        teacher_name  : 担任名（省略可）
    """
    if template_name not in TEMPLATES:
        raise ValueError(f'未定義のテンプレート: {template_name}')

    meta = TEMPLATES[template_name]
    template_path = os.path.join(options['template_dir'], meta['file'])

    if not os.path.exists(template_path):
        raise FileNotFoundError(f'テンプレートが見つかりません: {template_path}')

    gen_type = meta['type']
    if gen_type == 'grid':
        return GridGenerator(template_path, output_path, data, options)
    if gen_type == 'list':
        return ListGenerator(template_path, output_path, data, options)
    if gen_type == 'individual':
        return IndividualGenerator(template_path, output_path, data, options)
    raise ValueError(f'不明なジェネレータータイプ: {gen_type}')
