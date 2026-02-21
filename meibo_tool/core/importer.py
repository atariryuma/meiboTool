"""C4th エクスポート Excel / CSV の読み込み

SPEC.md 第2章 §2.4 参照。
ヘッダー行自動検出 + 全列 dtype=str で読み込む。
"""

import csv
import os

import chardet
import pandas as pd
from openpyxl import load_workbook

from core.mapper import map_columns


def detect_header_row(filepath: str, max_scan: int = 10) -> int:
    """
    Excel ファイルのヘッダー行を自動検出する。
    判定基準: 文字列セルが 5 つ以上連続する最初の行（1-indexed）。
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.active
        result = 1  # フォールバック
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan), 1):
            str_count = sum(
                1 for cell in row
                if cell.value is not None and isinstance(cell.value, str)
            )
            if str_count >= 5:
                result = row_idx
                break
    finally:
        wb.close()
    return result


def detect_encoding(filepath: str) -> str:
    """ファイルのエンコーディングを自動検出する。

    chardet で推定し、信頼度が低い場合は cp932 にフォールバックする。
    """
    with open(filepath, 'rb') as f:
        raw = f.read(65536)  # 先頭 64KB で判定
    result = chardet.detect(raw)
    encoding = (result.get('encoding') or 'cp932').lower()
    # chardet が ascii と判定した場合は utf-8 にフォールバック
    if encoding == 'ascii':
        encoding = 'utf-8'
    # Windows-31J / ISO-2022-JP 系は cp932 に統一
    if encoding in ('windows-1252', 'iso-2022-jp'):
        encoding = 'cp932'
    return encoding


def detect_header_row_csv(
    filepath: str, encoding: str, max_scan: int = 10,
) -> int:
    """
    CSV ファイルのヘッダー行を自動検出する。
    判定基準: 非空文字列セルが 5 つ以上ある最初の行（1-indexed）。
    """
    result = 1  # フォールバック
    with open(filepath, encoding=encoding, newline='') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            if row_idx > max_scan:
                break
            str_count = sum(1 for cell in row if cell.strip())
            if str_count >= 5:
                result = row_idx
                break
    return result


def import_c4th_excel(filepath: str) -> tuple[pd.DataFrame, list[str]]:
    """
    C4th エクスポート Excel を読み込み、マッピング済み DataFrame を返す。

    Returns:
        df_mapped: 内部論理名にリネーム済みの DataFrame
        unmapped:  マッピングできなかったカラム名リスト
    """
    header_row = detect_header_row(filepath)
    df = pd.read_excel(
        filepath,
        header=header_row - 1,  # 0-indexed
        dtype=str,               # 全列文字列（型変換は後工程）
    )
    return _clean_and_map(df)


def import_c4th_csv(filepath: str) -> tuple[pd.DataFrame, list[str]]:
    """
    C4th エクスポート CSV を読み込み、マッピング済み DataFrame を返す。

    Returns:
        df_mapped: 内部論理名にリネーム済みの DataFrame
        unmapped:  マッピングできなかったカラム名リスト
    """
    encoding = detect_encoding(filepath)
    header_row = detect_header_row_csv(filepath, encoding)
    df = pd.read_csv(
        filepath,
        header=header_row - 1,  # 0-indexed
        dtype=str,
        encoding=encoding,
    )
    return _clean_and_map(df)


def import_file(filepath: str) -> tuple[pd.DataFrame, list[str]]:
    """
    拡張子に応じて Excel または CSV を読み込む統合関数。

    Returns:
        df_mapped: 内部論理名にリネーム済みの DataFrame
        unmapped:  マッピングできなかったカラム名リスト
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return import_c4th_csv(filepath)
    return import_c4th_excel(filepath)


def _clean_and_map(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """空白カラム・空白行を除去し、カラムマッピングを適用する。"""
    df = df.loc[:, df.columns.notna()]
    df = df.dropna(how='all')
    df = df.reset_index(drop=True)
    return map_columns(df)
