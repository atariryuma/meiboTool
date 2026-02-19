"""openpyxl ワークシートを PIL Image にレンダリングする。

GUI 右パネルのテンプレートプレビュー用。
印刷品質ではなくレイアウト確認用の簡易レンダリング。
フォントは IPAmj明朝（実印刷用フォント）→ Meiryo → YuGothic → MSGothic の順で検索。
テンプレートに埋め込まれた画像も描画する。
"""

from __future__ import annotations

import functools
import io
import os
from typing import TYPE_CHECKING

from openpyxl.cell.cell import MergedCell
from openpyxl.utils.units import EMU_to_pixels
from PIL import Image, ImageDraw, ImageFont

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# ── 定数 ──────────────────────────────────────────────────────────────────────

_DEFAULT_COL_WIDTH_CHARS = 8.43
_DEFAULT_ROW_HEIGHT_PT = 15.0
_CHAR_WIDTH_PX = 7       # 1 文字幅 ≈ 7px (96 DPI 基準)
_PT_TO_PX = 1.333        # 1pt ≈ 1.333px (96 DPI)
_MARGIN = 4              # 画像外周マージン (px)
_CELL_PAD = 3            # セル内テキストパディング (px)

_BG_COLOR = (255, 255, 255)
_GRID_COLOR = (200, 200, 200)
_BORDER_THIN_COLOR = (120, 120, 120)
_BORDER_MEDIUM_COLOR = (40, 40, 40)
_TEXT_COLOR = (30, 30, 30)

# Windows フォントパス候補（実印刷フォント IPAmj明朝 を最優先）
_FONT_PATHS = [
    'C:/Windows/Fonts/ipamjm.ttf',
    'C:/Windows/Fonts/meiryo.ttc',
    'C:/Windows/Fonts/YuGothR.ttc',
    'C:/Windows/Fonts/msgothic.ttc',
]

# bold 用フォントパス候補
_BOLD_FONT_PATHS = [
    'C:/Windows/Fonts/meiryob.ttc',
    'C:/Windows/Fonts/YuGothB.ttc',
    'C:/Windows/Fonts/msgothic.ttc',
]


# ── フォントキャッシュ ─────────────────────────────────────────────────────────

@functools.lru_cache(maxsize=64)
def _load_font(size_px: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    """システムフォントを読み込む。bold の場合は専用フォントを試す。"""
    paths = _BOLD_FONT_PATHS + _FONT_PATHS if bold else _FONT_PATHS
    for path in paths:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size=size_px, index=0)
            except (OSError, IndexError):
                continue
    return ImageFont.load_default()


# ── ヘルパー ──────────────────────────────────────────────────────────────────

def _color_to_rgb(openpyxl_color) -> tuple[int, int, int] | None:
    """openpyxl Color → (r, g, b) タプル。変換不可なら None。"""
    if openpyxl_color is None:
        return None
    rgb = getattr(openpyxl_color, 'rgb', None)
    if rgb is None or rgb == '00000000':
        return None
    if isinstance(rgb, str) and len(rgb) >= 6:
        # ARGB 形式 'FF00FF00' or RGB 形式 '00FF00'
        hex_str = rgb[-6:]
        try:
            r = int(hex_str[0:2], 16)
            g = int(hex_str[2:4], 16)
            b = int(hex_str[4:6], 16)
            return (r, g, b)
        except ValueError:
            return None
    return None


def _get_col_widths(ws: Worksheet, num_cols: int, scale: float) -> list[float]:
    """各列のピクセル幅リスト。"""
    from openpyxl.utils import get_column_letter
    widths = []
    for c in range(1, num_cols + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width is not None and dim.width > 0:
            w = dim.width
        else:
            w = _DEFAULT_COL_WIDTH_CHARS
        widths.append(w * _CHAR_WIDTH_PX * scale)
    return widths


def _get_row_heights(ws: Worksheet, num_rows: int, scale: float) -> list[float]:
    """各行のピクセル高リスト。"""
    heights = []
    for r in range(1, num_rows + 1):
        dim = ws.row_dimensions.get(r)
        if dim and dim.height is not None and dim.height > 0:
            h = dim.height
        else:
            h = _DEFAULT_ROW_HEIGHT_PT
        heights.append(h * _PT_TO_PX * scale)
    return heights


def _build_merge_info(
    ws: Worksheet,
) -> tuple[dict[tuple[int, int], tuple[int, int, int, int]], set[tuple[int, int]]]:
    """
    結合セル情報を構築する。

    Returns:
        anchors: {(min_row, min_col): (min_row, min_col, max_row, max_col)}
        covered: set of (row, col) — 結合セル内だが左上以外
    """
    anchors: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for merged_range in ws.merged_cells.ranges:
        mr = (merged_range.min_row, merged_range.min_col,
              merged_range.max_row, merged_range.max_col)
        anchors[(mr[0], mr[1])] = mr
        for r in range(mr[0], mr[2] + 1):
            for c in range(mr[1], mr[3] + 1):
                if (r, c) != (mr[0], mr[1]):
                    covered.add((r, c))
    return anchors, covered


def _draw_images(
    ws: Worksheet,
    img: Image.Image,
    col_x: list[float],
    row_y: list[float],
    scale: float,
) -> None:
    """ワークシートに埋め込まれた画像を描画する。"""
    if not hasattr(ws, '_images') or not ws._images:
        return

    for ws_img in ws._images:
        try:
            # 画像データ取得
            img_data = io.BytesIO(ws_img._data())
            pil_img = Image.open(img_data)

            # アンカー位置の取得
            anchor = ws_img.anchor
            if hasattr(anchor, 'min_row') and hasattr(anchor, 'min_col'):
                # TwoCellAnchor / OneCellAnchor — (0-based)
                r = anchor.min_row
                c = anchor.min_col
            elif hasattr(anchor, '_from'):
                r = anchor._from.row
                c = anchor._from.col
            else:
                continue

            if r >= len(row_y) or c >= len(col_x):
                continue

            x = int(col_x[c])
            y = int(row_y[r])

            # 画像サイズの決定
            if hasattr(ws_img, 'width') and hasattr(ws_img, 'height'):
                w = int(EMU_to_pixels(ws_img.width) * scale) if ws_img.width > 1000 else int(ws_img.width * scale)
                h = int(EMU_to_pixels(ws_img.height) * scale) if ws_img.height > 1000 else int(ws_img.height * scale)
            else:
                w = int(pil_img.width * scale)
                h = int(pil_img.height * scale)

            if w > 0 and h > 0:
                pil_img = pil_img.resize((w, h), Image.LANCZOS)
                # RGBA → RGB 合成
                if pil_img.mode == 'RGBA':
                    bg = Image.new('RGB', pil_img.size, _BG_COLOR)
                    bg.paste(pil_img, mask=pil_img.split()[3])
                    pil_img = bg
                elif pil_img.mode != 'RGB':
                    pil_img = pil_img.convert('RGB')

                img.paste(pil_img, (x, y))
        except Exception:
            continue


# ── 描画パス（render_worksheet から分離） ─────────────────────────────────────

def _cell_rect(
    r: int,
    c: int,
    r_idx: int,
    c_idx: int,
    anchors: dict[tuple[int, int], tuple[int, int, int, int]],
    col_x: list[float],
    row_y: list[float],
    num_cols: int,
    num_rows: int,
) -> tuple[float, float, float, float]:
    """セル (r, c) の矩形 (x1, y1, x2, y2) を返す。結合セルの場合は全体矩形。"""
    if (r, c) in anchors:
        mr = anchors[(r, c)]
        return (
            col_x[mr[1] - 1],
            row_y[mr[0] - 1],
            col_x[min(mr[3], num_cols)],
            row_y[min(mr[2], num_rows)],
        )
    return col_x[c_idx], row_y[r_idx], col_x[c_idx + 1], row_y[r_idx + 1]


def _render_fills_and_grid(
    ws: Worksheet,
    draw: ImageDraw.ImageDraw,
    num_rows: int,
    num_cols: int,
    anchors: dict,
    covered: set,
    col_x: list[float],
    row_y: list[float],
) -> None:
    """Pass 1: セル塗りつぶし + グリッド線。"""
    for r_idx in range(num_rows):
        r = r_idx + 1
        for c_idx in range(num_cols):
            c = c_idx + 1
            if (r, c) in covered:
                continue

            x1, y1, x2, y2 = _cell_rect(r, c, r_idx, c_idx, anchors, col_x, row_y, num_cols, num_rows)

            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell) and cell.fill and cell.fill.fgColor:
                fill_rgb = _color_to_rgb(cell.fill.fgColor)
                if fill_rgb and fill_rgb != (0, 0, 0):
                    draw.rectangle([x1, y1, x2, y2], fill=fill_rgb)

            draw.rectangle([x1, y1, x2, y2], outline=_GRID_COLOR)


def _render_borders(
    ws: Worksheet,
    draw: ImageDraw.ImageDraw,
    num_rows: int,
    num_cols: int,
    anchors: dict,
    covered: set,
    col_x: list[float],
    row_y: list[float],
) -> None:
    """Pass 2: セル罫線。"""
    for r_idx in range(num_rows):
        r = r_idx + 1
        for c_idx in range(num_cols):
            c = c_idx + 1
            if (r, c) in covered:
                continue

            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell) or not cell.border:
                continue

            x1, y1, x2, y2 = _cell_rect(r, c, r_idx, c_idx, anchors, col_x, row_y, num_cols, num_rows)

            bdr = cell.border
            for side_name, coords in [
                ('top', [(x1, y1), (x2, y1)]),
                ('bottom', [(x1, y2), (x2, y2)]),
                ('left', [(x1, y1), (x1, y2)]),
                ('right', [(x2, y1), (x2, y2)]),
            ]:
                side = getattr(bdr, side_name, None)
                if side and side.style:
                    if side.style in ('medium', 'thick'):
                        draw.line(coords, fill=_BORDER_MEDIUM_COLOR, width=2)
                    else:
                        draw.line(coords, fill=_BORDER_THIN_COLOR, width=1)


def _render_text(
    ws: Worksheet,
    draw: ImageDraw.ImageDraw,
    num_rows: int,
    num_cols: int,
    anchors: dict,
    covered: set,
    col_x: list[float],
    row_y: list[float],
    scale: float,
) -> None:
    """Pass 3: テキスト描画。"""
    for r_idx in range(num_rows):
        r = r_idx + 1
        for c_idx in range(num_cols):
            c = c_idx + 1
            if (r, c) in covered:
                continue

            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell) or cell.value is None:
                continue

            text = str(cell.value)
            if not text:
                continue

            x1, y1, x2, y2 = _cell_rect(r, c, r_idx, c_idx, anchors, col_x, row_y, num_cols, num_rows)

            # フォント
            font_size_pt = 11
            is_bold = False
            if cell.font:
                if cell.font.size:
                    font_size_pt = cell.font.size
                is_bold = bool(cell.font.bold)
            font = _load_font(max(8, int(font_size_pt * scale)), is_bold)

            # テキストカラー
            text_rgb = _TEXT_COLOR
            if cell.font and cell.font.color:
                c_rgb = _color_to_rgb(cell.font.color)
                if c_rgb:
                    text_rgb = c_rgb

            # アライメント
            h_align = 'left'
            v_align = 'center'
            if cell.alignment:
                if cell.alignment.horizontal:
                    h_align = cell.alignment.horizontal
                if cell.alignment.vertical:
                    v_align = cell.alignment.vertical

            cell_w = x2 - x1
            cell_h = y2 - y1

            # テキスト測定 + 切り詰め
            bbox = draw.textbbox((0, 0), text, font=font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]

            max_text_w = cell_w - _CELL_PAD * 2
            if text_w > max_text_w and max_text_w > 0:
                ratio = max_text_w / text_w
                text = text[:max(1, int(len(text) * ratio))] + '…'
                bbox = draw.textbbox((0, 0), text, font=font)
                text_w = bbox[2] - bbox[0]
                text_h = bbox[3] - bbox[1]

            # 水平位置
            if h_align == 'center':
                tx = x1 + (cell_w - text_w) / 2
            elif h_align == 'right':
                tx = x2 - text_w - _CELL_PAD
            else:
                tx = x1 + _CELL_PAD

            # 垂直位置
            if v_align == 'top':
                ty = y1 + _CELL_PAD
            elif v_align == 'bottom':
                ty = y2 - text_h - _CELL_PAD
            else:
                ty = y1 + (cell_h - text_h) / 2

            draw.text((tx, ty), text, fill=text_rgb, font=font)


# ── メイン API ─────────────────────────────────────────────────────────────────

def render_worksheet(
    ws: Worksheet,
    max_rows: int | None = None,
    max_cols: int | None = None,
    scale: float = 1.5,
) -> Image.Image:
    """
    ワークシートの内容を PIL Image にレンダリングして返す。

    Parameters
    ----------
    ws : openpyxl Worksheet
    max_rows : 描画する最大行数（None=ws.max_row まで）
    max_cols : 描画する最大列数（None=ws.max_column まで）
    scale : レンダリング倍率

    Returns
    -------
    PIL.Image.Image (RGB)
    """
    num_rows = min(ws.max_row or 1, max_rows or (ws.max_row or 1))
    num_cols = min(ws.max_column or 1, max_cols or (ws.max_column or 1))

    col_widths = _get_col_widths(ws, num_cols, scale)
    row_heights = _get_row_heights(ws, num_rows, scale)

    total_w = max(int(sum(col_widths) + _MARGIN * 2), 100)
    total_h = max(int(sum(row_heights) + _MARGIN * 2), 50)

    img = Image.new('RGB', (total_w, total_h), _BG_COLOR)
    draw = ImageDraw.Draw(img)

    anchors, covered = _build_merge_info(ws)

    col_x = [_MARGIN]
    for w in col_widths:
        col_x.append(col_x[-1] + w)

    row_y = [_MARGIN]
    for h in row_heights:
        row_y.append(row_y[-1] + h)

    _render_fills_and_grid(ws, draw, num_rows, num_cols, anchors, covered, col_x, row_y)
    _render_borders(ws, draw, num_rows, num_cols, anchors, covered, col_x, row_y)
    _render_text(ws, draw, num_rows, num_cols, anchors, covered, col_x, row_y, scale)
    _draw_images(ws, img, col_x, row_y, scale)

    return img
