"""gen_from_lay.py の .lay → Excel 変換テスト

テスト対象:
  - 境界座標収集
  - 近接座標マージ
  - 寸法変換
  - セルマッピング
  - ラベル・フィールド配置
  - 罫線変換
  - 実ファイル変換 E2E
"""

from __future__ import annotations

import os
import re

from openpyxl import load_workbook

from core.lay_parser import (
    FontInfo,
    LayFile,
    LayoutObject,
    ObjectType,
    Point,
    Rect,
)
from templates.generators.gen_from_lay import (
    calc_column_widths,
    calc_row_heights,
    collect_boundaries,
    convert_from_layfile,
    merge_close_boundaries,
    object_to_cell_range,
)

# ── ヘルパー ─────────────────────────────────────────────────────────────────


def _make_lay(objects: list[LayoutObject] | None = None) -> LayFile:
    """テスト用の最小 LayFile を構築する。"""
    return LayFile(
        title='テスト',
        version=1600,
        page_width=840,
        page_height=1188,
        objects=objects or [],
    )


def _collect_placeholders(ws) -> set[str]:
    """ワークシート内の全 {{...}} プレースホルダーを収集する。"""
    phs: set[str] = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                phs.update(re.findall(r'\{\{[^}]+\}\}', cell.value))
    return phs


def _collect_texts(ws) -> list[str]:
    """ワークシート内の全テキスト値を収集する。"""
    texts: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                texts.append(cell.value)
    return texts


# ── テストクラス ─────────────────────────────────────────────────────────────


class TestBoundaryCollection:
    """境界座標収集のテスト。"""

    def test_collects_from_rects(self):
        objs = [
            LayoutObject(
                obj_type=ObjectType.LABEL,
                rect=Rect(10, 20, 100, 80),
            ),
        ]
        xs, ys = collect_boundaries(_make_lay(objs))
        assert 10 in xs
        assert 100 in xs
        assert 20 in ys
        assert 80 in ys

    def test_collects_from_lines(self):
        objs = [
            LayoutObject(
                obj_type=ObjectType.LINE,
                line_start=Point(0, 50),
                line_end=Point(200, 50),
            ),
        ]
        xs, ys = collect_boundaries(_make_lay(objs))
        assert 0 in xs
        assert 200 in xs
        assert 50 in ys

    def test_sorted_and_deduplicated(self):
        objs = [
            LayoutObject(obj_type=ObjectType.LABEL, rect=Rect(50, 30, 100, 80)),
            LayoutObject(obj_type=ObjectType.LABEL, rect=Rect(10, 30, 50, 60)),
        ]
        xs, ys = collect_boundaries(_make_lay(objs))
        assert xs == sorted(set(xs))
        assert ys == sorted(set(ys))


class TestMergeCloseBoundaries:
    """近接座標マージのテスト。"""

    def test_no_merge_when_far_apart(self):
        result = merge_close_boundaries([10, 50, 100, 200])
        assert result == [10, 50, 100, 200]

    def test_merge_close_values(self):
        result = merge_close_boundaries([10, 12, 50, 51, 100], tolerance=3)
        assert result == [10, 50, 100]

    def test_empty_input(self):
        assert merge_close_boundaries([]) == []

    def test_single_value(self):
        assert merge_close_boundaries([42]) == [42]


class TestDimensions:
    """寸法変換のテスト。"""

    def test_column_widths(self):
        x_bounds = [0, 100, 200]
        widths = calc_column_widths(x_bounds)
        assert len(widths) == 2
        # 100 units × 0.25mm = 25mm → 25/2.0 = 12.5 Excel chars
        assert abs(widths[0] - 12.5) < 0.01

    def test_row_heights(self):
        y_bounds = [0, 40, 100]
        heights = calc_row_heights(y_bounds)
        assert len(heights) == 2
        # 40 units × 0.25mm = 10mm → 10 × 2.835 = 28.35pt
        assert abs(heights[0] - 28.35) < 0.01

    def test_minimum_width_enforced(self):
        # 1 unit = 0.25mm → 0.125 Excel chars → clamped to 0.5
        widths = calc_column_widths([0, 1])
        assert widths[0] >= 0.5

    def test_minimum_height_enforced(self):
        heights = calc_row_heights([0, 1])
        assert heights[0] >= 3.0


class TestObjectMapping:
    """オブジェクト → セル範囲マッピングのテスト。"""

    def test_single_cell(self):
        obj = LayoutObject(
            obj_type=ObjectType.LABEL,
            rect=Rect(10, 20, 50, 60),
        )
        x_bounds = [10, 50, 100]
        y_bounds = [20, 60, 100]
        result = object_to_cell_range(obj, x_bounds, y_bounds)
        assert result == (1, 1, 1, 1)

    def test_spanning_cells(self):
        obj = LayoutObject(
            obj_type=ObjectType.LABEL,
            rect=Rect(10, 20, 100, 100),
        )
        x_bounds = [10, 50, 100]
        y_bounds = [20, 60, 100]
        result = object_to_cell_range(obj, x_bounds, y_bounds)
        assert result == (1, 1, 2, 2)

    def test_no_rect_returns_none(self):
        obj = LayoutObject(obj_type=ObjectType.LINE)
        result = object_to_cell_range(obj, [0, 100], [0, 100])
        assert result is None


class TestFullConversion:
    """LayFile → Excel の統合テスト。"""

    def test_basic_conversion(self, tmp_path):
        objs = [
            LayoutObject(
                obj_type=ObjectType.LABEL,
                rect=Rect(10, 10, 200, 50),
                text='テスト帳票',
                font=FontInfo('ＭＳ ゴシック', 16.0),
            ),
            LayoutObject(
                obj_type=ObjectType.FIELD,
                rect=Rect(10, 60, 200, 100),
                field_id=108,
                font=FontInfo('ＭＳ 明朝', 10.0),
            ),
            LayoutObject(
                obj_type=ObjectType.LINE,
                line_start=Point(10, 50),
                line_end=Point(200, 50),
            ),
        ]
        lay = _make_lay(objs)
        out = str(tmp_path / 'test.xlsx')
        convert_from_layfile(lay, out)

        assert os.path.isfile(out)
        wb = load_workbook(out)
        ws = wb.active

        texts = _collect_texts(ws)
        assert any('テスト帳票' in t for t in texts)
        assert any('{{氏名}}' in t for t in texts)

    def test_unknown_field_placeholder(self, tmp_path):
        objs = [
            LayoutObject(
                obj_type=ObjectType.FIELD,
                rect=Rect(10, 10, 100, 50),
                field_id=99999,
            ),
        ]
        lay = _make_lay(objs)
        out = str(tmp_path / 'unknown.xlsx')
        convert_from_layfile(lay, out)

        wb = load_workbook(out)
        phs = _collect_placeholders(wb.active)
        assert '{{field_99999}}' in phs


