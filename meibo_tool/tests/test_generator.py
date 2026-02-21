"""core/generator.py のテスト

テスト対象:
  - GridGenerator が番号付きプレースホルダーを実データに置換すること
  - fill_placeholders / setup_print の基本動作
  - name_display モード（furigana/kanji/kana）の差込動作
"""

from __future__ import annotations

import os

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from core.generator import (
    GridGenerator,
    fill_placeholders,
    setup_print,
)
from templates.generators import gen_nafuda

# ────────────────────────────────────────────────────────────────────────────
# フィクスチャ
# ────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def tmpl_nafuda(tmp_path) -> str:
    """名札_1年生用テンプレートを一時ディレクトリに生成して返す。"""
    out = str(tmp_path / '名札_1年生用.xlsx')
    gen_nafuda.generate(out, mode='1年生')
    return out


@pytest.fixture
def tmpl_grid_simple(tmp_path) -> str:
    """GridGenerator テスト用の簡易グリッドテンプレートを生成して返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'テスト'
    ws['A1'] = '{{学年}}年{{組}}組 {{担任名}}'
    for i in range(1, 9):
        ws.cell(row=i + 1, column=1, value=f'{{{{出席番号_{i}}}}}')
        ws.cell(row=i + 1, column=2, value=f'{{{{氏名かな_{i}}}}}')
        ws.cell(row=i + 1, column=3, value=f'{{{{氏名_{i}}}}}')
    out = str(tmp_path / 'テスト_グリッド.xlsx')
    wb.save(out)
    wb.close()
    return out


# ────────────────────────────────────────────────────────────────────────────
# GridGenerator: 基本動作テスト
# ────────────────────────────────────────────────────────────────────────────

class TestGridGeneratorBasic:
    """簡易グリッドテンプレートへの GridGenerator 差込テスト。"""

    def _options(self, tmpl_path: str, name_display: str = 'furigana') -> dict:
        return {
            'fiscal_year': 2025,
            'school_name': '那覇市立天久小学校',
            'teacher_name': '山田先生',
            'template_dir': os.path.dirname(tmpl_path),
            'name_display': name_display,
        }

    def test_output_file_created(self, tmpl_grid_simple, dummy_df, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        gen = GridGenerator(tmpl_grid_simple, out, dummy_df,
                            self._options(tmpl_grid_simple))
        assert gen.generate() == out
        assert os.path.isfile(out)

    def test_header_placeholders_filled(self, tmpl_grid_simple, dummy_df, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        GridGenerator(tmpl_grid_simple, out, dummy_df,
                      self._options(tmpl_grid_simple)).generate()
        title = load_workbook(out).active['A1'].value or ''
        assert '{{学年}}' not in title
        assert '{{組}}' not in title
        assert '山田先生' in title

    def test_data_filled(self, tmpl_grid_simple, dummy_df, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        GridGenerator(tmpl_grid_simple, out, dummy_df,
                      self._options(tmpl_grid_simple)).generate()
        ws = load_workbook(out).active
        assert ws['C2'].value == dummy_df.iloc[0]['氏名']

    def test_unused_slots_blank(self, tmpl_grid_simple, dummy_df, tmp_path):
        """5名データでは No.6 の氏名セルが空になること。"""
        out = str(tmp_path / 'output.xlsx')
        GridGenerator(tmpl_grid_simple, out, dummy_df,
                      self._options(tmpl_grid_simple)).generate()
        ws = load_workbook(out).active
        # Person 6 → row 7, col 3
        assert ws.cell(row=7, column=3).value in (None, '')


# ────────────────────────────────────────────────────────────────────────────
# fill_placeholders: 特殊キーのテスト
# ────────────────────────────────────────────────────────────────────────────

class TestFillPlaceholders:
    def _make_ws(self):
        from openpyxl import Workbook
        return Workbook().active

    def test_fiscal_year(self):
        ws = self._make_ws()
        ws['A1'] = '{{年度}}'
        fill_placeholders(ws, {}, {'fiscal_year': 2025})
        assert ws['A1'].value == '2025'

    def test_school_name(self):
        ws = self._make_ws()
        ws['A1'] = '{{学校名}}'
        fill_placeholders(ws, {}, {'school_name': 'テスト小学校'})
        assert ws['A1'].value == 'テスト小学校'

    def test_teacher_name(self):
        ws = self._make_ws()
        ws['A1'] = '{{担任名}}'
        fill_placeholders(ws, {}, {'teacher_name': '鈴木先生'})
        assert ws['A1'].value == '鈴木先生'

    def test_wareki_nendo(self):
        ws = self._make_ws()
        ws['A1'] = '{{年度和暦}}'
        fill_placeholders(ws, {}, {'fiscal_year': 2025})
        assert ws['A1'].value == '令和7年度'

    def test_address_combined(self):
        ws = self._make_ws()
        ws['A1'] = '{{住所}}'
        data = {
            '都道府県': '沖縄県', '市区町村': '那覇市',
            '町番地': '天久1-2-3', '建物名': '',
        }
        fill_placeholders(ws, data, {})
        assert ws['A1'].value == '沖縄県那覇市天久1-2-3'

    def test_nan_value_becomes_empty(self):
        ws = self._make_ws()
        ws['A1'] = '{{氏名}}'
        fill_placeholders(ws, {'氏名': 'nan'}, {})
        assert ws['A1'].value == ''

    # ── name_display モードテスト ──────────────────────────────────────────

    def test_furigana_mode_fills_both(self):
        """furigana モードでは氏名・氏名かなの両方が展開される。"""
        ws = self._make_ws()
        ws['A1'] = '{{氏名かな}}'
        ws['A2'] = '{{氏名}}'
        data = {'氏名': '山田 太郎', '氏名かな': 'やまだ たろう'}
        fill_placeholders(ws, data, {'name_display': 'furigana'})
        assert ws['A1'].value == 'やまだ たろう'
        assert ws['A2'].value == '山田 太郎'

    def test_kanji_mode_blanks_kana_field(self):
        """kanji モードでは {{氏名かな}} が空白になる。"""
        ws = self._make_ws()
        ws['A1'] = '{{氏名かな}}'
        ws['A2'] = '{{氏名}}'
        data = {'氏名': '山田 太郎', '氏名かな': 'やまだ たろう'}
        fill_placeholders(ws, data, {'name_display': 'kanji'})
        assert ws['A1'].value == ''
        assert ws['A2'].value == '山田 太郎'

    def test_kana_mode_kana_row_blank(self):
        """kana モードでは {{氏名かな}} が空白になる（かな行を細く保つ）。"""
        ws = self._make_ws()
        ws['A1'] = '{{氏名かな}}'
        ws['A2'] = '{{氏名}}'
        data = {'氏名': '山田 太郎', '氏名かな': 'やまだ たろう'}
        fill_placeholders(ws, data, {'name_display': 'kana'})
        assert ws['A1'].value == ''

    def test_kana_mode_puts_kana_in_name_field(self):
        """kana モードでは {{氏名}} にかな値が転写される（氏名行に大きく表示）。"""
        ws = self._make_ws()
        ws['A1'] = '{{氏名かな}}'
        ws['A2'] = '{{氏名}}'
        data = {'氏名': '山田 太郎', '氏名かな': 'やまだ たろう'}
        fill_placeholders(ws, data, {'name_display': 'kana'})
        assert ws['A2'].value == 'やまだ たろう'

    def test_default_mode_same_as_furigana(self):
        """name_display 未指定（デフォルト）は furigana と同じ動作。"""
        ws = self._make_ws()
        ws['A1'] = '{{氏名かな}}'
        ws['A2'] = '{{氏名}}'
        data = {'氏名': '田中 花子', '氏名かな': 'たなか はなこ'}
        fill_placeholders(ws, data, {})
        assert ws['A1'].value == 'たなか はなこ'
        assert ws['A2'].value == '田中 花子'

    def test_kanji_mode_blanks_formal_kana(self):
        """kanji モードでは {{正式氏名かな}} も空白になる。"""
        ws = self._make_ws()
        ws['A1'] = '{{正式氏名かな}}'
        ws['A2'] = '{{正式氏名}}'
        data = {'正式氏名': '山田 太郎', '正式氏名かな': 'やまだ たろう'}
        fill_placeholders(ws, data, {'name_display': 'kanji'})
        assert ws['A1'].value == ''
        assert ws['A2'].value == '山田 太郎'

    def test_kana_mode_puts_formal_kana_in_formal_name_field(self):
        """kana モードでは {{正式氏名}} に正式氏名かなの値が転写される。"""
        ws = self._make_ws()
        ws['A1'] = '{{正式氏名かな}}'
        ws['A2'] = '{{正式氏名}}'
        data = {'正式氏名': '山田 太郎', '正式氏名かな': 'やまだ たろう'}
        fill_placeholders(ws, data, {'name_display': 'kana'})
        assert ws['A1'].value == ''
        assert ws['A2'].value == 'やまだ たろう'


# ────────────────────────────────────────────────────────────────────────────
# setup_print: 印刷設定テスト
# ────────────────────────────────────────────────────────────────────────────

class TestSetupPrint:
    def _make_ws(self):
        from openpyxl import Workbook
        return Workbook().active

    def test_portrait_orientation(self):
        ws = self._make_ws()
        setup_print(ws, orientation='portrait')
        assert ws.page_setup.orientation == 'portrait'

    def test_landscape_orientation(self):
        ws = self._make_ws()
        setup_print(ws, orientation='landscape')
        assert ws.page_setup.orientation == 'landscape'

    def test_paper_size_a4(self):
        ws = self._make_ws()
        setup_print(ws)
        assert ws.page_setup.paperSize == 9

    def test_fit_to_page(self):
        ws = self._make_ws()
        setup_print(ws)
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True
        assert ws.page_setup.fitToWidth == 1


# ────────────────────────────────────────────────────────────────────────────
# GridGenerator: 性別ソートテスト（男女一覧テンプレート用）
# ────────────────────────────────────────────────────────────────────────────

class TestGridGeneratorGenderSort:
    """sort_by='性別' メタデータによるソート機能のテスト。"""

    @pytest.fixture
    def mixed_gender_df(self) -> pd.DataFrame:
        """男女混合の 6 名データ（出席番号順: 女,男,女,男,女,男）。"""
        return pd.DataFrame([
            {'出席番号': '1', '氏名': '佐藤 花子', '氏名かな': 'さとう はなこ',
             '性別': '女', '学年': '3', '組': '1'},
            {'出席番号': '2', '氏名': '田中 太郎', '氏名かな': 'たなか たろう',
             '性別': '男', '学年': '3', '組': '1'},
            {'出席番号': '3', '氏名': '鈴木 美穂', '氏名かな': 'すずき みほ',
             '性別': '女', '学年': '3', '組': '1'},
            {'出席番号': '4', '氏名': '高橋 健太', '氏名かな': 'たかはし けんた',
             '性別': '男', '学年': '3', '組': '1'},
            {'出席番号': '5', '氏名': '山田 凛', '氏名かな': 'やまだ りん',
             '性別': '女', '学年': '3', '組': '1'},
            {'出席番号': '6', '氏名': '伊藤 翔', '氏名かな': 'いとう しょう',
             '性別': '男', '学年': '3', '組': '1'},
        ])

    @pytest.fixture
    def danjo_tmpl(self, tmp_path) -> str:
        """男女一覧テンプレート相当のダミー xlsx を生成。"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '男女一覧'
        # 6 スロットのグリッド
        for i in range(1, 7):
            ws.cell(row=i, column=1, value=f'{{{{出席番号_{i}}}}}')
            ws.cell(row=i, column=2, value=f'{{{{氏名_{i}}}}}')
            ws.cell(row=i, column=3, value=f'{{{{性別_{i}}}}}')
        wb.save(str(tmp_path / '男女一覧.xlsx'))
        wb.close()
        return str(tmp_path / '男女一覧.xlsx')

    def test_gender_sort_men_first(self, danjo_tmpl, mixed_gender_df, tmp_path):
        """sort_by='性別' で男→女の順にソートされる。"""
        output = str(tmp_path / 'output.xlsx')
        import shutil
        shutil.copy2(danjo_tmpl, output)

        options = {
            'template_dir': str(tmp_path),
            'fiscal_year': 2025,
            'school_name': 'テスト小',
            'teacher_name': 'テスト先生',
            'name_display': 'furigana',
        }
        gen = GridGenerator(danjo_tmpl, output, mixed_gender_df, options)
        gen.generate()

        wb = load_workbook(output)
        ws = wb.active
        # 男が先: 出席番号 2, 4, 6 → 女: 1, 3, 5
        assert ws.cell(row=1, column=2).value == '田中 太郎'   # 男・出席番号2
        assert ws.cell(row=2, column=2).value == '高橋 健太'   # 男・出席番号4
        assert ws.cell(row=3, column=2).value == '伊藤 翔'     # 男・出席番号6
        assert ws.cell(row=4, column=2).value == '佐藤 花子'   # 女・出席番号1
        assert ws.cell(row=5, column=2).value == '鈴木 美穂'   # 女・出席番号3
        assert ws.cell(row=6, column=2).value == '山田 凛'     # 女・出席番号5
        wb.close()

    def test_no_sort_without_sort_by(self, tmpl_grid_simple, tmp_path):
        """sort_by がないテンプレートではソートしない（出席番号順のまま）。"""
        df = pd.DataFrame([
            {'出席番号': '1', '氏名': '佐藤 花子', '氏名かな': 'さとう はなこ',
             '性別': '女', '学年': '3', '組': '1'},
            {'出席番号': '2', '氏名': '田中 太郎', '氏名かな': 'たなか たろう',
             '性別': '男', '学年': '3', '組': '1'},
        ])
        output = str(tmp_path / 'output_grid.xlsx')

        options = {
            'template_dir': str(tmp_path),
            'fiscal_year': 2025,
            'school_name': 'テスト小',
            'teacher_name': 'テスト先生',
            'name_display': 'furigana',
        }
        gen = GridGenerator(tmpl_grid_simple, output, df, options)
        gen.generate()

        wb = load_workbook(output)
        ws = wb.active
        # 出席番号順で佐藤 花子が先
        assert ws.cell(row=2, column=3).value == '佐藤 花子'
        wb.close()
