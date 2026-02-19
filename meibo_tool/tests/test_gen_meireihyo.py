"""gen_meireihyo.py テンプレート生成のユニットテスト

掲示用名列表（A4 縦・2列×20名・かな行+氏名行）の構造・プレースホルダー・印刷設定を検証する。
"""

from __future__ import annotations

import re

import pytest
from openpyxl import load_workbook

from templates.generators import gen_meireihyo

# ── フィクスチャ ──────────────────────────────────────────────────────────────

@pytest.fixture
def tmpl(tmp_path) -> str:
    out = str(tmp_path / '掲示用名列表.xlsx')
    gen_meireihyo.generate(out)
    return out


def _collect_placeholders(ws) -> set[str]:
    phs = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                phs.update(re.findall(r'\{\{[^}]+\}\}', cell.value))
    return phs


# ── テスト ──────────────────────────────────────────────────────────────────

class TestMeireihyo:

    def test_file_created(self, tmpl):
        import os
        assert os.path.isfile(tmpl)

    def test_sheet_title(self, tmpl):
        wb = load_workbook(tmpl)
        assert wb.active.title == '名列表'
        wb.close()

    def test_title_row_placeholders(self, tmpl):
        wb = load_workbook(tmpl)
        ws = wb.active
        title = ws.cell(row=1, column=1).value
        assert '{{学年}}' in title
        assert '{{組}}' in title
        assert '{{担任名}}' in title
        wb.close()

    def test_header_row(self, tmpl):
        wb = load_workbook(tmpl)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == '番号'
        assert ws.cell(row=2, column=2).value == '氏名'
        assert ws.cell(row=2, column=4).value == '番号'
        assert ws.cell(row=2, column=5).value == '氏名'
        wb.close()

    def test_placeholders_40_students(self, tmpl):
        """左列20名 + 右列20名 = 40名分のプレースホルダー。"""
        wb = load_workbook(tmpl)
        phs = _collect_placeholders(wb.active)
        for n in range(1, 41):
            assert f'{{{{出席番号_{n}}}}}' in phs
            assert f'{{{{氏名_{n}}}}}' in phs
            assert f'{{{{氏名かな_{n}}}}}' in phs
        # 41 番はない
        assert '{{出席番号_41}}' not in phs
        wb.close()

    def test_orientation_portrait(self, tmpl):
        wb = load_workbook(tmpl)
        assert wb.active.page_setup.orientation == 'portrait'
        wb.close()

    def test_paper_size_a4(self, tmpl):
        wb = load_workbook(tmpl)
        assert wb.active.page_setup.paperSize == 9
        wb.close()

    def test_column_widths(self, tmpl):
        wb = load_workbook(tmpl)
        ws = wb.active
        assert ws.column_dimensions['A'].width == 5.0  # 番号列
        assert ws.column_dimensions['B'].width == 27.0  # 氏名列
        assert ws.column_dimensions['C'].width == 2.0   # 区切り
        wb.close()

    def test_kana_row_height(self, tmpl):
        """かな行（row 3）は 11pt。"""
        wb = load_workbook(tmpl)
        ws = wb.active
        assert ws.row_dimensions[3].height == 11
        wb.close()

    def test_name_row_height(self, tmpl):
        """氏名行（row 4）は 25pt。"""
        wb = load_workbook(tmpl)
        ws = wb.active
        assert ws.row_dimensions[4].height == 25
        wb.close()

    def test_merged_cells_for_numbers(self, tmpl):
        """番号列はかな行+氏名行の2行分マージ。"""
        wb = load_workbook(tmpl)
        ws = wb.active
        merged = [str(r) for r in ws.merged_cells.ranges]
        # A3:A4 は番号列（左）の最初のマージ
        assert any('A3:A4' in m for m in merged)
        wb.close()

    def test_kana_border_has_dotted_bottom(self, tmpl):
        """かな行の下罫線は hair（点線）。"""
        wb = load_workbook(tmpl)
        ws = wb.active
        cell = ws.cell(row=3, column=2)  # 最初のかな行
        assert cell.border.bottom.style == 'hair'
        wb.close()
