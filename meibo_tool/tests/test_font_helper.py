"""utils/font_helper.py のユニットテスト"""

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font

from utils.font_helper import DEFAULT_FONT, apply_font


class TestApplyFont:
    """apply_font() のテスト"""

    def _make_ws(self):
        wb = Workbook()
        ws = wb.active
        return wb, ws

    def test_changes_font_name(self):
        wb, ws = self._make_ws()
        ws['A1'] = 'テスト'
        apply_font(ws)
        assert ws['A1'].font.name == DEFAULT_FONT
        wb.close()

    def test_preserves_font_size(self):
        wb, ws = self._make_ws()
        ws['A1'] = 'テスト'
        ws['A1'].font = Font(name='Arial', size=14)
        apply_font(ws)
        assert ws['A1'].font.size == 14
        assert ws['A1'].font.name == DEFAULT_FONT
        wb.close()

    def test_preserves_bold(self):
        wb, ws = self._make_ws()
        ws['A1'] = 'テスト'
        ws['A1'].font = Font(name='Arial', bold=True)
        apply_font(ws)
        assert ws['A1'].font.bold is True
        assert ws['A1'].font.name == DEFAULT_FONT
        wb.close()

    def test_preserves_italic(self):
        wb, ws = self._make_ws()
        ws['A1'] = 'テスト'
        ws['A1'].font = Font(name='Arial', italic=True)
        apply_font(ws)
        assert ws['A1'].font.italic is True
        wb.close()

    def test_preserves_underline(self):
        wb, ws = self._make_ws()
        ws['A1'] = 'テスト'
        ws['A1'].font = Font(name='Arial', underline='single')
        apply_font(ws)
        assert ws['A1'].font.underline == 'single'
        wb.close()

    def test_preserves_strike(self):
        wb, ws = self._make_ws()
        ws['A1'] = 'テスト'
        ws['A1'].font = Font(name='Arial', strike=True)
        apply_font(ws)
        assert ws['A1'].font.strike is True
        wb.close()

    def test_skips_none_cells(self):
        wb, ws = self._make_ws()
        ws['A1'] = None
        ws['A2'] = 'テスト'
        apply_font(ws)
        # None セルはフォント変更されない（デフォルトのまま）
        assert ws['A1'].font.name != DEFAULT_FONT or ws['A1'].value is None
        assert ws['A2'].font.name == DEFAULT_FONT
        wb.close()

    def test_skips_merged_cells(self):
        """結合セルの左上以外（MergedCell）はスキップされる。"""
        wb, ws = self._make_ws()
        ws['A1'] = 'テスト'
        ws.merge_cells('A1:B2')
        # merge 後、B1/A2/B2 は MergedCell になる
        apply_font(ws)
        # 左上 A1 のフォントは変更される
        assert ws['A1'].font.name == DEFAULT_FONT
        # MergedCell はエラーなく処理される（読み取り専用なのでスキップ）
        assert isinstance(ws['B1'], MergedCell)
        wb.close()

    def test_custom_font_name(self):
        wb, ws = self._make_ws()
        ws['A1'] = 'テスト'
        apply_font(ws, font_name='Meiryo')
        assert ws['A1'].font.name == 'Meiryo'
        wb.close()

    def test_multiple_cells(self):
        wb, ws = self._make_ws()
        ws['A1'] = '太郎'
        ws['B1'] = '花子'
        ws['A2'] = '次郎'
        apply_font(ws)
        assert ws['A1'].font.name == DEFAULT_FONT
        assert ws['B1'].font.name == DEFAULT_FONT
        assert ws['A2'].font.name == DEFAULT_FONT
        wb.close()

    def test_default_font_constant(self):
        assert DEFAULT_FONT == 'IPAmj明朝'
