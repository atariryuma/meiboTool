"""gen_daicho.py テンプレート生成のユニットテスト

修了台帳（shuuryo）と卒業台帳（sotsugyou）の構造・プレースホルダー・印刷設定を検証する。
"""

from __future__ import annotations

import re

import pytest
from openpyxl import load_workbook

from templates.generators import gen_daicho

# ── フィクスチャ ──────────────────────────────────────────────────────────────

@pytest.fixture
def tmpl_shuuryo(tmp_path) -> str:
    out = str(tmp_path / '修了台帳.xlsx')
    gen_daicho.generate(out, mode='shuuryo')
    return out


@pytest.fixture
def tmpl_sotsugyou(tmp_path) -> str:
    out = str(tmp_path / '卒業台帳.xlsx')
    gen_daicho.generate(out, mode='sotsugyou')
    return out


def _collect_placeholders(ws) -> set[str]:
    phs = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                phs.update(re.findall(r'\{\{[^}]+\}\}', cell.value))
    return phs


# ── 修了台帳 ──────────────────────────────────────────────────────────────

class TestShuuryoDaicho:

    def test_file_created(self, tmpl_shuuryo):
        import os
        assert os.path.isfile(tmpl_shuuryo)

    def test_sheet_title(self, tmpl_shuuryo):
        wb = load_workbook(tmpl_shuuryo)
        assert wb.active.title == '修了台帳'
        wb.close()

    def test_title_row_placeholders(self, tmpl_shuuryo):
        wb = load_workbook(tmpl_shuuryo)
        ws = wb.active
        title = ws.cell(row=1, column=1).value
        assert '{{年度和暦}}' in title
        assert '{{担任名}}' in title
        wb.close()

    def test_header_row_labels(self, tmpl_shuuryo):
        wb = load_workbook(tmpl_shuuryo)
        ws = wb.active
        headers = [ws.cell(row=2, column=c).value for c in range(1, 10)]
        assert '正式氏名' in headers
        assert '生年月日' in headers
        assert '性別' in headers
        wb.close()

    def test_data_row_placeholders(self, tmpl_shuuryo):
        wb = load_workbook(tmpl_shuuryo)
        phs = _collect_placeholders(wb.active)
        assert '{{正式氏名}}' in phs
        assert '{{生年月日}}' in phs
        assert '{{出席番号}}' in phs
        wb.close()

    def test_orientation_landscape(self, tmpl_shuuryo):
        wb = load_workbook(tmpl_shuuryo)
        assert wb.active.page_setup.orientation == 'landscape'
        wb.close()

    def test_9_columns(self, tmpl_shuuryo):
        """修了台帳は 9 列。"""
        wb = load_workbook(tmpl_shuuryo)
        ws = wb.active
        # データ行のカラム数
        data_cells = [ws.cell(row=3, column=c).value for c in range(1, 15)]
        non_none = [v for v in data_cells if v is not None]
        assert len(non_none) == 9
        wb.close()

    def test_header_has_fill(self, tmpl_shuuryo):
        wb = load_workbook(tmpl_shuuryo)
        ws = wb.active
        fill = ws.cell(row=2, column=1).fill
        assert fill.fgColor.rgb is not None
        wb.close()


# ── 卒業台帳 ──────────────────────────────────────────────────────────────

class TestSotsugyouDaicho:

    def test_file_created(self, tmpl_sotsugyou):
        import os
        assert os.path.isfile(tmpl_sotsugyou)

    def test_sheet_title(self, tmpl_sotsugyou):
        wb = load_workbook(tmpl_sotsugyou)
        assert wb.active.title == '卒業台帳'
        wb.close()

    def test_title_row_placeholders(self, tmpl_sotsugyou):
        wb = load_workbook(tmpl_sotsugyou)
        ws = wb.active
        title = ws.cell(row=1, column=1).value
        assert '{{学校名}}' in title
        wb.close()

    def test_data_row_has_certificate_number(self, tmpl_sotsugyou):
        wb = load_workbook(tmpl_sotsugyou)
        phs = _collect_placeholders(wb.active)
        assert '{{証書番号}}' in phs
        wb.close()

    def test_8_columns(self, tmpl_sotsugyou):
        """卒業台帳は 8 列。"""
        wb = load_workbook(tmpl_sotsugyou)
        ws = wb.active
        data_cells = [ws.cell(row=3, column=c).value for c in range(1, 15)]
        non_none = [v for v in data_cells if v is not None]
        assert len(non_none) == 8
        wb.close()

    def test_paper_size_a4(self, tmpl_sotsugyou):
        wb = load_workbook(tmpl_sotsugyou)
        assert wb.active.page_setup.paperSize == 9
        wb.close()
