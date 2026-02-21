"""IndividualGenerator のテスト

テスト対象:
  - IndividualGenerator: シート複製・プレースホルダー置換
"""

from __future__ import annotations

import os

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from core.generator import (
    IndividualGenerator,
    copy_sheet_with_images,
)

# ── フィクスチャ ──────────────────────────────────────────────────────────────


@pytest.fixture
def tmpl_individual(tmp_path) -> str:
    """IndividualGenerator テスト用の簡易テンプレートを生成して返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'テンプレート'
    ws['A1'] = '{{学校名}} {{年度和暦}} 家庭調査票'
    ws['A2'] = '氏名: {{正式氏名}}'
    ws['B2'] = 'かな: {{正式氏名かな}}'
    ws['A3'] = '性別: {{性別}}'
    ws['A4'] = '住所: {{住所}}'
    ws['A5'] = '保護者: {{保護者正式名}}'
    out = str(tmp_path / '家庭調査票.xlsx')
    wb.save(out)
    return out


@pytest.fixture
def individual_data() -> pd.DataFrame:
    """IndividualGenerator 用の 3 名データ。"""
    return pd.DataFrame([
        {
            '出席番号': '1', '正式氏名': '山田 太郎', '正式氏名かな': 'やまだ たろう',
            '氏名': '山田 太郎', '性別': '男', '生年月日': '2018-06-15',
            '保護者正式名': '山田 幸子',
            '都道府県': '沖縄県', '市区町村': '那覇市', '町番地': '天久1-2-3', '建物名': '',
            '学年': '1', '組': '1',
        },
        {
            '出席番号': '2', '正式氏名': '田中 花子', '正式氏名かな': 'たなか はなこ',
            '氏名': '田中 花子', '性別': '女', '生年月日': '2018-04-03',
            '保護者正式名': '田中 美穂',
            '都道府県': '沖縄県', '市区町村': '那覇市', '町番地': '古島2-3-4', '建物名': 'コーポA',
            '学年': '1', '組': '1',
        },
        {
            '出席番号': '3', '正式氏名': '鈴木 健太', '正式氏名かな': 'すずき けんた',
            '氏名': '鈴木 健太', '性別': '男', '生年月日': '2018-09-20',
            '保護者正式名': '鈴木 典子',
            '都道府県': '沖縄県', '市区町村': '那覇市', '町番地': '真地3-4-5', '建物名': '',
            '学年': '1', '組': '1',
        },
    ])


def _default_options(tmpl_path: str) -> dict:
    return {
        'fiscal_year': 2025,
        'school_name': '那覇市立天久小学校',
        'teacher_name': '山田先生',
        'template_dir': os.path.dirname(tmpl_path),
        'name_display': 'furigana',
    }


# ── IndividualGenerator ───────────────────────────────────────────────────────


class TestIndividualGenerator:
    def test_output_file_created(self, tmpl_individual, individual_data, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        gen = IndividualGenerator(
            tmpl_individual, out, individual_data, _default_options(tmpl_individual)
        )
        result = gen.generate()
        assert result == out
        assert os.path.isfile(out)

    def test_sheet_count_matches_data(self, tmpl_individual, individual_data, tmp_path):
        """シート数がデータ行数と一致する。"""
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_individual, out, individual_data, _default_options(tmpl_individual)
        ).generate()
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 3

    def test_sheet_names_include_student_info(self, tmpl_individual, individual_data, tmp_path):
        """シート名に出席番号と氏名が含まれる。"""
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_individual, out, individual_data, _default_options(tmpl_individual)
        ).generate()
        wb = load_workbook(out)
        assert '01_山田 太郎' in wb.sheetnames
        assert '02_田中 花子' in wb.sheetnames

    def test_placeholders_filled_per_sheet(self, tmpl_individual, individual_data, tmp_path):
        """各シートのプレースホルダーが該当生徒のデータで置換される。"""
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_individual, out, individual_data, _default_options(tmpl_individual)
        ).generate()
        wb = load_workbook(out)
        # 1枚目のシート
        ws1 = wb.worksheets[0]
        assert '山田 太郎' in (ws1['A2'].value or '')
        # 2枚目
        ws2 = wb.worksheets[1]
        assert '田中 花子' in (ws2['A2'].value or '')

    def test_single_student(self, tmpl_individual, tmp_path):
        """1名データでもエラーにならない。"""
        data = pd.DataFrame([{
            '出席番号': '1', '正式氏名': '山田 太郎', '正式氏名かな': 'やまだ たろう',
            '氏名': '山田 太郎', '性別': '男', '生年月日': '2018-06-15',
            '保護者正式名': '山田 幸子',
            '都道府県': '沖縄県', '市区町村': '那覇市', '町番地': '天久1-2-3', '建物名': '',
            '学年': '1', '組': '1',
        }])
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_individual, out, data, _default_options(tmpl_individual)
        ).generate()
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 1


# ── copy_sheet_with_images ────────────────────────────────────────────────────


class TestCopySheetWithImages:
    def test_copy_without_images(self, tmp_path):
        """画像なしシートのコピーが正常に動作する。"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'テスト'
        new_ws = copy_sheet_with_images(wb, ws, 'コピー')
        assert new_ws.title == 'コピー'
        assert new_ws['A1'].value == 'テスト'
