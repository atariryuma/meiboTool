"""gen_nafuda.py テンプレート生成のユニットテスト

3 モード（通常 / 装飾 / 1年生）のテンプレート構造・プレースホルダー・印刷設定を検証する。
"""

from __future__ import annotations

import re

import pytest
from openpyxl import load_workbook

from templates.generators import gen_nafuda

# ── フィクスチャ ──────────────────────────────────────────────────────────────

@pytest.fixture
def tmpl_normal(tmp_path) -> str:
    out = str(tmp_path / '名札_通常.xlsx')
    gen_nafuda.generate(out, mode='通常')
    return out


@pytest.fixture
def tmpl_deco(tmp_path) -> str:
    out = str(tmp_path / '名札_装飾あり.xlsx')
    gen_nafuda.generate(out, mode='装飾')
    return out


@pytest.fixture
def tmpl_1nen(tmp_path) -> str:
    out = str(tmp_path / '名札_1年生用.xlsx')
    gen_nafuda.generate(out, mode='1年生')
    return out


def _collect_placeholders(ws) -> set[str]:
    """ワークシート内の全 {{...}} プレースホルダーを収集する。"""
    phs = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                phs.update(re.findall(r'\{\{[^}]+\}\}', cell.value))
    return phs


# ── 名札_通常 ──────────────────────────────────────────────────────────────

class TestNafudaNormal:

    def test_file_created(self, tmpl_normal):
        import os
        assert os.path.isfile(tmpl_normal)

    def test_sheet_title(self, tmpl_normal):
        wb = load_workbook(tmpl_normal)
        assert wb.active.title == '名札'
        wb.close()

    def test_placeholders_10_cards(self, tmpl_normal):
        wb = load_workbook(tmpl_normal)
        phs = _collect_placeholders(wb.active)
        for n in range(1, 11):
            assert f'{{{{氏名_{n}}}}}' in phs
            assert f'{{{{出席番号_{n}}}}}' in phs
            assert f'{{{{氏名かな_{n}}}}}' in phs
        wb.close()

    def test_orientation_landscape(self, tmpl_normal):
        wb = load_workbook(tmpl_normal)
        assert wb.active.page_setup.orientation == 'landscape'
        wb.close()

    def test_paper_size_a4(self, tmpl_normal):
        wb = load_workbook(tmpl_normal)
        assert wb.active.page_setup.paperSize == 9
        wb.close()

    def test_column_widths(self, tmpl_normal):
        wb = load_workbook(tmpl_normal)
        ws = wb.active
        assert ws.column_dimensions['A'].width == 4.5
        assert ws.column_dimensions['D'].width == 2.0  # 区切り列
        wb.close()

    def test_merged_cells_exist(self, tmpl_normal):
        wb = load_workbook(tmpl_normal)
        assert len(list(wb.active.merged_cells.ranges)) > 0
        wb.close()


# ── 名札_装飾あり ──────────────────────────────────────────────────────────

class TestNafudaDecorated:

    def test_file_created(self, tmpl_deco):
        import os
        assert os.path.isfile(tmpl_deco)

    def test_placeholders_same_as_normal(self, tmpl_normal, tmpl_deco):
        wb_n = load_workbook(tmpl_normal)
        wb_d = load_workbook(tmpl_deco)
        phs_n = _collect_placeholders(wb_n.active)
        phs_d = _collect_placeholders(wb_d.active)
        assert phs_n == phs_d
        wb_n.close()
        wb_d.close()

    def test_border_style_thick(self, tmpl_deco):
        """装飾版は太い罫線（thick）が使われている。"""
        wb = load_workbook(tmpl_deco)
        ws = wb.active
        cell = ws['A1']
        assert cell.border.top.style == 'thick'
        wb.close()


# ── 名札_1年生用 ──────────────────────────────────────────────────────────

class TestNafuda1nen:

    def test_file_created(self, tmpl_1nen):
        import os
        assert os.path.isfile(tmpl_1nen)

    def test_sheet_title(self, tmpl_1nen):
        wb = load_workbook(tmpl_1nen)
        assert wb.active.title == '名札'
        wb.close()

    def test_placeholders_8_cards(self, tmpl_1nen):
        wb = load_workbook(tmpl_1nen)
        phs = _collect_placeholders(wb.active)
        for n in range(1, 9):
            assert f'{{{{出席番号_{n}}}}}' in phs
            assert f'{{{{氏名かな_{n}}}}}' in phs
        # 9枚目はない
        assert '{{出席番号_9}}' not in phs
        wb.close()

    def test_orientation_portrait(self, tmpl_1nen):
        wb = load_workbook(tmpl_1nen)
        assert wb.active.page_setup.orientation == 'portrait'
        wb.close()

    def test_vertical_text_rotation(self, tmpl_1nen):
        """かな縦書きのセルは textRotation=255。"""
        wb = load_workbook(tmpl_1nen)
        ws = wb.active
        # Row 3 = かな行
        cell = ws.cell(row=3, column=1)
        assert cell.alignment.textRotation == 255
        wb.close()

    def test_8_columns(self, tmpl_1nen):
        wb = load_workbook(tmpl_1nen)
        ws = wb.active
        for col in 'ABCDEFGH':
            assert ws.column_dimensions[col].width == 8.0
        wb.close()
