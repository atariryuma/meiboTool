"""gen_shirabehyo.py テンプレート生成のユニットテスト

調べ表（A4 縦・6列×10行グリッド）の構造・プレースホルダー・印刷設定を検証する。
"""

from __future__ import annotations

import re

import pytest
from openpyxl import load_workbook

from templates.generators import gen_shirabehyo

# ── フィクスチャ ──────────────────────────────────────────────────────────────

@pytest.fixture
def tmpl(tmp_path) -> str:
    out = str(tmp_path / '調べ表.xlsx')
    gen_shirabehyo.generate(out)
    return out


def _collect_placeholders(ws) -> set[str]:
    phs = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                phs.update(re.findall(r'\{\{[^}]+\}\}', cell.value))
    return phs


# ── テスト ──────────────────────────────────────────────────────────────────

class TestShirabehyo:

    def test_file_created(self, tmpl):
        import os
        assert os.path.isfile(tmpl)

    def test_sheet_title(self, tmpl):
        wb = load_workbook(tmpl)
        assert wb.active.title == '調べ表'
        wb.close()

    def test_title_row(self, tmpl):
        wb = load_workbook(tmpl)
        ws = wb.active
        # 左半分: 調査項目名
        assert '（調査項目名）' in (ws.cell(row=1, column=1).value or '')
        # 右半分: 調べ　期限
        assert '調べ' in (ws.cell(row=1, column=4).value or '')
        wb.close()

    def test_sub_header_placeholders(self, tmpl):
        wb = load_workbook(tmpl)
        ws = wb.active
        sub_left = ws.cell(row=2, column=1).value
        sub_right = ws.cell(row=2, column=4).value
        assert '{{学年}}' in sub_left
        assert '{{組}}' in sub_left
        assert '{{担任名}}' in sub_right
        wb.close()

    def test_placeholders_60_students(self, tmpl):
        """6列×10行 = 60名分のプレースホルダー。"""
        wb = load_workbook(tmpl)
        phs = _collect_placeholders(wb.active)
        for n in range(1, 61):
            assert f'{{{{氏名_{n}}}}}' in phs
            assert f'{{{{出席番号_{n}}}}}' in phs
        # 61 番はない
        assert '{{氏名_61}}' not in phs
        wb.close()

    def test_placeholder_numbering_row_major(self, tmpl):
        """番号割り当て: 行優先 (行i, 列j) → N = i*6 + j + 1。"""
        wb = load_workbook(tmpl)
        ws = wb.active
        # Row 0, Col 0 → N=1 (kana_row=3, col=1)
        val = ws.cell(row=3, column=1).value
        assert '{{出席番号_1}}' in val
        # Row 0, Col 5 → N=6 (kana_row=3, col=6)
        val = ws.cell(row=3, column=6).value
        assert '{{出席番号_6}}' in val
        # Row 1, Col 0 → N=7 (kana_row=6, col=1)
        val = ws.cell(row=6, column=1).value
        assert '{{出席番号_7}}' in val
        wb.close()

    def test_orientation_portrait(self, tmpl):
        wb = load_workbook(tmpl)
        assert wb.active.page_setup.orientation == 'portrait'
        wb.close()

    def test_paper_size_a4(self, tmpl):
        wb = load_workbook(tmpl)
        assert wb.active.page_setup.paperSize == 9
        wb.close()

    def test_6_data_columns(self, tmpl):
        wb = load_workbook(tmpl)
        ws = wb.active
        # 列 A〜F はデータ列（等幅 12.5）
        for col in 'ABCDEF':
            assert ws.column_dimensions[col].width == 12.5
        wb.close()

    def test_3_rows_per_student(self, tmpl):
        """各生徒は3行: かな行(10) + 氏名行(22) + 記入行(16)。"""
        wb = load_workbook(tmpl)
        ws = wb.active
        assert ws.row_dimensions[3].height == 10   # かな行
        assert ws.row_dimensions[4].height == 22   # 氏名行
        assert ws.row_dimensions[5].height == 16   # 記入行
        wb.close()

    def test_record_row_is_empty(self, tmpl):
        """記入行（row 5）は空白。"""
        wb = load_workbook(tmpl)
        ws = wb.active
        for col in range(1, 7):
            assert ws.cell(row=5, column=col).value is None
        wb.close()

    def test_sub_header_has_fill(self, tmpl):
        wb = load_workbook(tmpl)
        ws = wb.active
        fill = ws.cell(row=2, column=1).fill
        assert fill.fgColor.rgb is not None
        wb.close()
