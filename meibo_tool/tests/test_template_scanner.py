"""テンプレートスキャナーのユニットテスト"""

from __future__ import annotations

import os

import pytest
from openpyxl import Workbook

from templates.template_registry import TEMPLATES, get_all_templates
from templates.template_scanner import (
    clear_cache,
    scan_template_file,
    scan_template_folder,
)


@pytest.fixture(autouse=True)
def _clear_scanner_cache():
    """各テストの前後でスキャナーキャッシュをクリアする。"""
    clear_cache()
    yield
    clear_cache()


def _save_wb(wb: Workbook, path: str) -> str:
    """Workbook を保存してパスを返す。"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


# ────────────────────────────────────────────────────────────────────────────
# scan_template_file
# ────────────────────────────────────────────────────────────────────────────


class TestScanTemplateFileGrid:
    """番号付きプレースホルダーを持つ grid テンプレートの検出。"""

    def test_detects_grid_type(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{出席番号_1}}'
        ws['B1'] = '{{氏名_1}}'
        ws['A2'] = '{{出席番号_2}}'
        ws['B2'] = '{{氏名_2}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result is not None
        assert result['type'] == 'grid'

    def test_cards_per_page_from_max_number(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=f'{{{{氏名_{i}}}}}')
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['cards_per_page'] == 10

    def test_orientation_from_page_setup(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        ws.page_setup.orientation = 'landscape'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['orientation'] == 'landscape'

    def test_default_orientation_portrait(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['orientation'] == 'portrait'

    def test_filename_in_result(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        path = _save_wb(wb, str(tmp_path / 'テスト名札.xlsx'))

        result = scan_template_file(path)
        assert result['file'] == 'テスト名札.xlsx'

    def test_icon_for_grid(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['icon'] == '📋'


class TestScanTemplateFileList:
    """番号なしプレースホルダー（少数の個人情報フィールド）で list タイプ検出。"""

    def test_detects_list_type(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{年度和暦}}'
        ws['A2'] = '{{出席番号}}'
        ws['B2'] = '{{正式氏名}}'
        ws['C2'] = '{{生年月日}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result is not None
        assert result['type'] == 'list'

    def test_no_cards_per_page(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名}}'
        ws['B1'] = '{{性別}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['cards_per_page'] is None


class TestScanTemplateFileIndividual:
    """個人情報フィールドが多い場合に individual タイプ検出。"""

    def test_detects_individual_type(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{正式氏名}}'
        ws['A2'] = '{{郵便番号}}'
        ws['A3'] = '{{住所}}'
        ws['A4'] = '{{保護者正式名}}'
        ws['A5'] = '{{緊急連絡先}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result is not None
        assert result['type'] == 'individual'

    def test_use_formal_name_detected(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{正式氏名}}'
        ws['A2'] = '{{正式氏名かな}}'
        ws['A3'] = '{{保護者正式名}}'
        ws['A4'] = '{{緊急連絡先}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['use_formal_name'] is True

    def test_icon_for_individual(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{正式氏名}}'
        ws['A2'] = '{{郵便番号}}'
        ws['A3'] = '{{保護者正式名}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['icon'] == '📄'


class TestScanTemplateFileEdgeCases:
    """エッジケース。"""

    def test_no_placeholders_returns_none(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '普通のテキスト'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result is None

    def test_empty_workbook_returns_none(self, tmp_path):
        wb = Workbook()
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result is None

    def test_nonexistent_file_returns_none(self, tmp_path):
        result = scan_template_file(str(tmp_path / 'nonexistent.xlsx'))
        assert result is None

    def test_special_keys_excluded_from_required(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{年度和暦}}'
        ws['B1'] = '{{学校名}}'
        ws['C1'] = '{{担任名}}'
        ws['D1'] = '{{氏名}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert '年度和暦' not in result['required_columns']
        assert '学校名' not in result['required_columns']
        assert '担任名' not in result['required_columns']
        assert '氏名' in result['required_columns']

    def test_mandatory_columns_always_set(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['mandatory_columns'] == ['組', '出席番号']

    def test_use_formal_name_false_when_not_present(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        ws['B1'] = '{{氏名かな_1}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['use_formal_name'] is False

    def test_formal_name_in_numbered(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{正式氏名_1}}'
        ws['B1'] = '{{正式氏名かな_1}}'
        path = _save_wb(wb, str(tmp_path / 'test.xlsx'))

        result = scan_template_file(path)
        assert result['use_formal_name'] is True


# ────────────────────────────────────────────────────────────────────────────
# scan_template_folder
# ────────────────────────────────────────────────────────────────────────────


class TestScanTemplateFolder:
    """フォルダスキャンのテスト。"""

    def test_scans_multiple_files(self, tmp_path):
        for name in ['テンプレA.xlsx', 'テンプレB.xlsx']:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = '{{氏名_1}}'
            wb.save(str(tmp_path / name))

        result = scan_template_folder(str(tmp_path))
        assert 'テンプレA' in result
        assert 'テンプレB' in result

    def test_skips_non_xlsx(self, tmp_path):
        (tmp_path / 'readme.txt').write_text('not a template')
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        wb.save(str(tmp_path / 'real.xlsx'))

        result = scan_template_folder(str(tmp_path))
        assert 'readme' not in result
        assert 'real' in result

    def test_skips_temp_files(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        wb.save(str(tmp_path / '~$temp.xlsx'))
        wb.save(str(tmp_path / 'normal.xlsx'))

        result = scan_template_folder(str(tmp_path))
        assert '~$temp' not in result
        assert 'normal' in result

    def test_nonexistent_dir_returns_empty(self):
        result = scan_template_folder('/nonexistent/path')
        assert result == {}

    def test_skips_non_template_xlsx(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '普通のデータ'
        wb.save(str(tmp_path / 'data.xlsx'))

        result = scan_template_folder(str(tmp_path))
        assert 'data' not in result

    def test_cache_returns_same_result(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        wb.save(str(tmp_path / 'cached.xlsx'))

        result1 = scan_template_folder(str(tmp_path))
        result2 = scan_template_folder(str(tmp_path))
        assert result1 == result2


# ────────────────────────────────────────────────────────────────────────────
# get_all_templates（マージロジック）
# ────────────────────────────────────────────────────────────────────────────


class TestGetAllTemplates:
    """TEMPLATES と自動検出のマージをテストする。"""

    def test_registry_entries_present(self, tmp_path):
        """TEMPLATES にあるテンプレートが結果に含まれる（ファイル存在時）。"""
        # 既存テンプレートをコピー
        src = os.path.join(os.path.dirname(__file__), '..', '..', 'テンプレート')
        src = os.path.abspath(src)
        import shutil
        for _name, meta in TEMPLATES.items():
            f = meta['file']
            src_path = os.path.join(src, f)
            if os.path.exists(src_path):
                shutil.copy2(src_path, str(tmp_path / f))
                break  # 1つあれば十分

        result = get_all_templates(str(tmp_path))
        # 少なくとも1つの TEMPLATES エントリが含まれる
        registry_keys = set(TEMPLATES.keys())
        found = registry_keys & set(result.keys())
        assert len(found) >= 1

    def test_auto_detected_template_included(self, tmp_path):
        """フォルダにのみ存在するテンプレートが結果に含まれる。"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'
        ws['B1'] = '{{出席番号_1}}'
        wb.save(str(tmp_path / 'カスタム名札.xlsx'))

        result = get_all_templates(str(tmp_path))
        assert 'カスタム名札' in result
        assert result['カスタム名札']['type'] == 'grid'

    def test_registry_overrides_scan(self, tmp_path):
        """TEMPLATES にある定義はスキャン結果より優先される。"""
        # TEMPLATES の最初のエントリのファイルを作成
        first_key = next(iter(TEMPLATES))
        meta = TEMPLATES[first_key]
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{氏名_1}}'  # スキャンでは grid と判定される
        wb.save(str(tmp_path / meta['file']))

        result = get_all_templates(str(tmp_path))
        assert first_key in result
        # TEMPLATES のアイコンが保持されている
        assert result[first_key]['icon'] == meta['icon']

    def test_missing_file_excluded(self, tmp_path):
        """ファイルが存在しない enabled テンプレートは結果に含まれない。"""
        result = get_all_templates(str(tmp_path))
        # 空フォルダなので enabled テンプレートのファイルは存在しない
        enabled_keys = {
            k for k, v in TEMPLATES.items() if v.get('enabled', True)
        }
        for key in enabled_keys:
            assert key not in result

    def test_disabled_template_preserved(self, tmp_path):
        """enabled=False のテンプレートは常に結果に含まれる。"""
        result = get_all_templates(str(tmp_path))
        disabled_keys = {
            k for k, v in TEMPLATES.items() if v.get('enabled') is False
        }
        for key in disabled_keys:
            assert key in result
