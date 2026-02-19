"""家庭調査票・学級編成用個票テンプレートのテスト

テスト対象:
  - gen_katei_chousahyo: テンプレート生成・プレースホルダー配置・印刷設定
  - gen_gakkyuu_kojihyo: テンプレート生成・プレースホルダー配置・印刷設定
  - IndividualGenerator による差込生成
"""

from __future__ import annotations

import os

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.generator import IndividualGenerator
from templates.generators import gen_gakkyuu_kojihyo, gen_katei_chousahyo

# ── フィクスチャ ──────────────────────────────────────────────────────────────


@pytest.fixture
def tmpl_katei(tmp_path) -> str:
    """家庭調査票テンプレートを生成して返す。"""
    out = str(tmp_path / '家庭調査票.xlsx')
    gen_katei_chousahyo.generate(out)
    return out


@pytest.fixture
def tmpl_gakkyuu(tmp_path) -> str:
    """学級編成用個票テンプレートを生成して返す。"""
    out = str(tmp_path / '学級編成用個票.xlsx')
    gen_gakkyuu_kojihyo.generate(out)
    return out


@pytest.fixture
def student_data() -> pd.DataFrame:
    """個票テスト用の 2 名データ。"""
    return pd.DataFrame([
        {
            '出席番号': '1', '正式氏名': '山田 太郎', '正式氏名かな': 'やまだ たろう',
            '氏名': '山田 太郎', '性別': '男', '生年月日': '2018-06-15',
            '郵便番号': '900-0005', '電話番号1': '098-123-4567',
            '保護者正式名': '山田 幸子', '保護者正式名かな': 'やまだ さちこ',
            '保護者続柄': '母', '緊急連絡先': '098-999-0000',
            '外国籍': '',
            '都道府県': '沖縄県', '市区町村': '那覇市', '町番地': '天久1-2-3', '建物名': '',
            '学年': '1', '組': '1',
        },
        {
            '出席番号': '2', '正式氏名': '田中 花子', '正式氏名かな': 'たなか はなこ',
            '氏名': '田中 花子', '性別': '女', '生年月日': '2018-04-03',
            '郵便番号': '900-0014', '電話番号1': '098-234-5678',
            '保護者正式名': '田中 美穂', '保護者正式名かな': 'たなか みほ',
            '保護者続柄': '母', '緊急連絡先': '098-888-1111',
            '外国籍': '',
            '都道府県': '沖縄県', '市区町村': '那覇市', '町番地': '古島2-3-4', '建物名': 'コーポA',
            '学年': '1', '組': '1',
        },
    ])


def _default_options(tmpl_path: str) -> dict:
    return {
        'fiscal_year': 2025,
        'school_name': '那覇市立天久小学校',
        'teacher_name': '山田先生',
        'template_dir': os.path.dirname(tmpl_path),
        'name_display': 'furigana',
    }


# ── 家庭調査票テンプレート生成 ────────────────────────────────────────────────


class TestGenKateiChousahyo:
    def test_generates_file(self, tmpl_katei):
        assert os.path.isfile(tmpl_katei)

    def test_sheet_name(self, tmpl_katei):
        wb = load_workbook(tmpl_katei)
        assert wb.active.title == 'テンプレート'

    def test_title_has_placeholders(self, tmpl_katei):
        wb = load_workbook(tmpl_katei)
        ws = wb.active
        title = ws.cell(row=1, column=1).value
        assert '{{学校名}}' in title
        assert '{{年度和暦}}' in title
        assert '家庭調査票' in title

    def test_has_student_name_placeholder(self, tmpl_katei):
        wb = load_workbook(tmpl_katei)
        ws = wb.active
        values = [ws.cell(row=r, column=c).value or ''
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any('{{正式氏名}}' in v for v in values)
        assert any('{{正式氏名かな}}' in v for v in values)

    def test_has_guardian_placeholders(self, tmpl_katei):
        wb = load_workbook(tmpl_katei)
        ws = wb.active
        values = [ws.cell(row=r, column=c).value or ''
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any('{{保護者正式名}}' in v for v in values)
        assert any('{{保護者続柄}}' in v for v in values)
        assert any('{{緊急連絡先}}' in v for v in values)

    def test_portrait_a4(self, tmpl_katei):
        wb = load_workbook(tmpl_katei)
        ws = wb.active
        assert ws.page_setup.orientation == 'portrait'
        assert ws.page_setup.paperSize == 9

    def test_fit_to_page(self, tmpl_katei):
        wb = load_workbook(tmpl_katei)
        ws = wb.active
        assert ws.page_setup.fitToWidth == 1


# ── 学級編成用個票テンプレート生成 ────────────────────────────────────────────


class TestGenGakkyuuKojihyo:
    def test_generates_file(self, tmpl_gakkyuu):
        assert os.path.isfile(tmpl_gakkyuu)

    def test_sheet_name(self, tmpl_gakkyuu):
        wb = load_workbook(tmpl_gakkyuu)
        assert wb.active.title == 'テンプレート'

    def test_title_has_placeholders(self, tmpl_gakkyuu):
        wb = load_workbook(tmpl_gakkyuu)
        ws = wb.active
        title = ws.cell(row=1, column=1).value
        assert '{{学校名}}' in title
        assert '学級編成用個票' in title

    def test_has_student_name_placeholder(self, tmpl_gakkyuu):
        wb = load_workbook(tmpl_gakkyuu)
        ws = wb.active
        values = [ws.cell(row=r, column=c).value or ''
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any('{{正式氏名}}' in v for v in values)
        assert any('{{正式氏名かな}}' in v for v in values)

    def test_has_gaikokuseki_placeholder(self, tmpl_gakkyuu):
        wb = load_workbook(tmpl_gakkyuu)
        ws = wb.active
        values = [ws.cell(row=r, column=c).value or ''
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
        assert any('{{外国籍}}' in v for v in values)

    def test_has_blank_sections(self, tmpl_gakkyuu):
        """手書き用空白セクション（学力・行動特性等）のヘッダーがある。"""
        wb = load_workbook(tmpl_gakkyuu)
        ws = wb.active
        values = [ws.cell(row=r, column=1).value or ''
                  for r in range(1, ws.max_row + 1)]
        section_labels = ' '.join(values)
        assert '学力' in section_labels
        assert '行動特性' in section_labels
        assert '引継ぎ' in section_labels

    def test_portrait_a4(self, tmpl_gakkyuu):
        wb = load_workbook(tmpl_gakkyuu)
        ws = wb.active
        assert ws.page_setup.orientation == 'portrait'
        assert ws.page_setup.paperSize == 9


# ── IndividualGenerator 統合テスト ─────────────────────────────────────────────


class TestIndividualKatei:
    def test_generates_two_sheets(self, tmpl_katei, student_data, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_katei, out, student_data, _default_options(tmpl_katei)
        ).generate()
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 2

    def test_student_name_filled(self, tmpl_katei, student_data, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_katei, out, student_data, _default_options(tmpl_katei)
        ).generate()
        wb = load_workbook(out)
        # 全セルを走査して山田太郎の名前が含まれることを確認
        ws1 = wb.worksheets[0]
        all_values = [ws1.cell(row=r, column=c).value or ''
                      for r in range(1, ws1.max_row + 1)
                      for c in range(1, ws1.max_column + 1)]
        assert any('山田 太郎' in v for v in all_values)

    def test_school_name_filled(self, tmpl_katei, student_data, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_katei, out, student_data, _default_options(tmpl_katei)
        ).generate()
        wb = load_workbook(out)
        ws1 = wb.worksheets[0]
        title = ws1.cell(row=1, column=1).value or ''
        assert '那覇市立天久小学校' in title
        assert '{{学校名}}' not in title


class TestIndividualGakkyuu:
    def test_generates_two_sheets(self, tmpl_gakkyuu, student_data, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_gakkyuu, out, student_data, _default_options(tmpl_gakkyuu)
        ).generate()
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 2

    def test_second_sheet_has_different_student(self, tmpl_gakkyuu, student_data, tmp_path):
        out = str(tmp_path / 'output.xlsx')
        IndividualGenerator(
            tmpl_gakkyuu, out, student_data, _default_options(tmpl_gakkyuu)
        ).generate()
        wb = load_workbook(out)
        ws2 = wb.worksheets[1]
        all_values = [ws2.cell(row=r, column=c).value or ''
                      for r in range(1, ws2.max_row + 1)
                      for c in range(1, ws2.max_column + 1)]
        assert any('田中 花子' in v for v in all_values)
