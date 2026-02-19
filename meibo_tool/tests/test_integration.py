"""統合テスト: データ読込 → マッピング → 全テンプレート生成の一気通貫

ダミー C4th Excel を生成 → importer.py で読み込み（マッピング含む）
→ 全 enabled テンプレートで create_generator → generate() の成功を確認する。
"""

from __future__ import annotations

import os
import re
import shutil

import pytest
from openpyxl import load_workbook

from core.generator import create_generator
from core.importer import import_c4th_excel
from templates.template_registry import TEMPLATES, get_all_templates
from tests.generate_dummy import generate_multi_class

# ── enabled テンプレート一覧 ──────────────────────────────────────────────────

_ENABLED = [name for name, meta in TEMPLATES.items() if meta.get('enabled', True)]
_LIST_INDIVIDUAL = [
    name for name, meta in TEMPLATES.items()
    if meta.get('enabled', True) and meta['type'] in ('list', 'individual')
]


# ── フィクスチャ ──────────────────────────────────────────────────────────────


@pytest.fixture(scope='module')
def c4th_excel(tmp_path_factory) -> str:
    """ダミー C4th Excel を生成して返す。"""
    out_dir = tmp_path_factory.mktemp('fixtures')
    out_path = str(out_dir / 'dummy_c4th.xlsx')
    df = generate_multi_class(classes=[(3, 1), (3, 2)], n_per_class=15)
    df.to_excel(out_path, index=False)
    return out_path


@pytest.fixture(scope='module')
def imported_data(c4th_excel):
    """importer で読み込んだマッピング済み DataFrame。"""
    df, _unmapped = import_c4th_excel(c4th_excel)
    return df


@pytest.fixture(scope='module')
def template_dir(tmp_path_factory) -> str:
    """テンプレート群をテスト用 tmp にコピーして返す。"""
    src = os.path.join(os.path.dirname(__file__), '..', '..', 'テンプレート')
    src = os.path.abspath(src)
    dst = str(tmp_path_factory.mktemp('templates'))
    for f in os.listdir(src):
        if f.endswith('.xlsx'):
            shutil.copy2(os.path.join(src, f), os.path.join(dst, f))
    return dst


def _make_options(tmpl_dir: str) -> dict:
    return {
        'fiscal_year': 2025,
        'school_name': '那覇市立天久小学校',
        'teacher_name': '山田先生',
        'template_dir': tmpl_dir,
        'name_display': 'furigana',
    }


# ── テスト ────────────────────────────────────────────────────────────────────


class TestImportPipeline:
    """データ読込パイプラインの基本検証。"""

    def test_import_returns_dataframe(self, imported_data):
        assert len(imported_data) == 30  # 15名 × 2クラス

    def test_mapped_columns_exist(self, imported_data):
        required = ['氏名', '氏名かな', '正式氏名', '性別', '生年月日', '組']
        for col in required:
            assert col in imported_data.columns, f'{col} がマッピングされていない'

    def test_filter_by_class(self, imported_data):
        kumi1 = imported_data[imported_data['組'] == '1']
        kumi2 = imported_data[imported_data['組'] == '2']
        assert len(kumi1) == 15
        assert len(kumi2) == 15


class TestAllTemplateGeneration:
    """全 enabled テンプレートで create_generator → generate() が成功する。"""

    @pytest.mark.parametrize('template_name', _ENABLED)
    def test_generate_succeeds(self, template_name, imported_data, template_dir, tmp_path):
        """テンプレートごとに生成が成功し、出力ファイルが存在する。"""
        data = imported_data[imported_data['組'] == '1'].copy()
        all_templates = get_all_templates(template_dir)
        meta = all_templates[template_name]
        out_path = str(tmp_path / f'{template_name}_output.xlsx')
        options = _make_options(template_dir)

        template_file = os.path.join(template_dir, meta['file'])
        if not os.path.exists(template_file):
            pytest.skip(f'テンプレートファイルが見つかりません: {meta["file"]}')

        gen = create_generator(template_name, out_path, data, options)
        result = gen.generate()
        assert os.path.isfile(result), f'{template_name}: 出力ファイルが作成されなかった'

    @pytest.mark.parametrize('template_name', _ENABLED)
    def test_output_has_data(self, template_name, imported_data, template_dir, tmp_path):
        """出力ファイルにデータが含まれている（空ファイルではない）。"""
        data = imported_data[imported_data['組'] == '1'].copy()
        all_templates = get_all_templates(template_dir)
        meta = all_templates[template_name]
        out_path = str(tmp_path / f'{template_name}_data.xlsx')
        options = _make_options(template_dir)

        template_file = os.path.join(template_dir, meta['file'])
        if not os.path.exists(template_file):
            pytest.skip(f'テンプレートファイルが見つかりません: {meta["file"]}')

        gen = create_generator(template_name, out_path, data, options)
        gen.generate()

        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.max_row >= 2, f'{template_name}: データが空'


class TestPlaceholderResolution:
    """生成後のファイルにプレースホルダーが残っていないことを確認する。"""

    @pytest.mark.parametrize('template_name', _LIST_INDIVIDUAL)
    def test_no_remaining_placeholders(
        self, template_name, imported_data, template_dir, tmp_path
    ):
        """list / individual 型テンプレートでプレースホルダーが残っていないこと。"""
        data = imported_data[imported_data['組'] == '1'].copy()
        all_templates = get_all_templates(template_dir)
        meta = all_templates[template_name]
        out_path = str(tmp_path / f'{template_name}_ph.xlsx')
        options = _make_options(template_dir)

        template_file = os.path.join(template_dir, meta['file'])
        if not os.path.exists(template_file):
            pytest.skip(f'テンプレートファイルが見つかりません: {meta["file"]}')

        gen = create_generator(template_name, out_path, data, options)
        gen.generate()

        wb = load_workbook(out_path)
        placeholder_re = re.compile(r'\{\{.+?\}\}')
        remaining = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        found = placeholder_re.findall(cell.value)
                        if found:
                            remaining.extend(found)

        assert not remaining, (
            f'{template_name}: 未解決のプレースホルダー: {remaining[:5]}'
        )
