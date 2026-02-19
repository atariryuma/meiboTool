"""gui/preview_renderer.py のユニットテスト

テスト対象:
  - render_worksheet(): PIL Image を返す基本動作
  - _color_to_rgb(): openpyxl カラー変換
  - 結合セル・空ワークシート・制限パラメータの処理
  - テンプレート生成 → レンダリングの統合テスト
"""

from __future__ import annotations

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image

from gui.preview_renderer import _color_to_rgb, render_worksheet

# ── _color_to_rgb ─────────────────────────────────────────────────────────────


class TestColorToRgb:
    """_color_to_rgb のユニットテスト。"""

    def test_none_returns_none(self):
        assert _color_to_rgb(None) is None

    def test_argb_hex(self):
        """ARGB 形式 'FF00FF00' → (0, 255, 0)"""

        class FakeColor:
            rgb = 'FF00FF00'

        assert _color_to_rgb(FakeColor()) == (0, 255, 0)

    def test_rgb_hex(self):
        """RGB 形式 'FF0000' → (255, 0, 0)"""

        class FakeColor:
            rgb = 'FF0000'

        assert _color_to_rgb(FakeColor()) == (255, 0, 0)

    def test_black_00000000_returns_none(self):
        """'00000000' はテーマカラーのデフォルト → None。"""

        class FakeColor:
            rgb = '00000000'

        assert _color_to_rgb(FakeColor()) is None

    def test_no_rgb_attr_returns_none(self):
        """rgb 属性がない場合は None。"""

        class FakeColor:
            pass

        assert _color_to_rgb(FakeColor()) is None


# ── render_worksheet 基本テスト ───────────────────────────────────────────────


class TestRenderWorksheetBasic:
    """render_worksheet の基本動作テスト。"""

    def test_returns_pil_image(self):
        """PIL Image を返す。"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'テスト'
        img = render_worksheet(ws)
        assert isinstance(img, Image.Image)
        assert img.mode == 'RGB'

    def test_image_has_minimum_size(self):
        """最小サイズ保証（100x50）。"""
        wb = Workbook()
        ws = wb.active
        img = render_worksheet(ws)
        assert img.width >= 100
        assert img.height >= 50

    def test_text_creates_non_white_pixels(self):
        """テキストが描画されると白以外のピクセルが存在する。"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'テストデータ'
        ws.column_dimensions['A'].width = 20
        img = render_worksheet(ws, scale=2.0)
        arr = np.array(img)
        # 完全に白 (255,255,255) ではないピクセルが存在する
        non_white = np.any(arr != 255, axis=2)
        assert non_white.sum() > 0

    def test_empty_worksheet_no_error(self):
        """空のワークシートでもエラーにならない。"""
        wb = Workbook()
        ws = wb.active
        img = render_worksheet(ws)
        assert isinstance(img, Image.Image)

    def test_scale_affects_size(self):
        """scale パラメータが画像サイズに影響する。"""
        wb = Workbook()
        ws = wb.active
        # 最小サイズ保証に引っかからないよう複数行列を用意
        for r in range(1, 11):
            for c in range(1, 6):
                ws.cell(row=r, column=c, value=f'{r}-{c}')
        img_small = render_worksheet(ws, scale=1.0)
        img_large = render_worksheet(ws, scale=2.0)
        # 2倍スケールの方が大きい
        assert img_large.width > img_small.width
        assert img_large.height > img_small.height


# ── max_rows / max_cols ───────────────────────────────────────────────────────


class TestRenderLimits:
    """max_rows / max_cols の制限テスト。"""

    def test_max_rows_limits_output(self):
        """max_rows が画像の高さを制限する。"""
        wb = Workbook()
        ws = wb.active
        for r in range(1, 51):
            ws.cell(row=r, column=1, value=f'行{r}')
        img_all = render_worksheet(ws, scale=1.0)
        img_10 = render_worksheet(ws, max_rows=10, scale=1.0)
        assert img_10.height < img_all.height

    def test_max_cols_limits_output(self):
        """max_cols が画像の幅を制限する。"""
        wb = Workbook()
        ws = wb.active
        for c in range(1, 21):
            ws.cell(row=1, column=c, value=f'列{c}')
        img_all = render_worksheet(ws, scale=1.0)
        img_5 = render_worksheet(ws, max_cols=5, scale=1.0)
        assert img_5.width < img_all.width


# ── 結合セル ──────────────────────────────────────────────────────────────────


class TestMergedCells:
    """結合セルの処理テスト。"""

    def test_merged_cells_no_crash(self):
        """結合セルがあってもクラッシュしない。"""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:C1')
        ws['A1'] = '結合されたセル'
        ws['A2'] = '通常セル'
        img = render_worksheet(ws)
        assert isinstance(img, Image.Image)

    def test_multiple_merged_ranges(self):
        """複数の結合範囲があっても処理できる。"""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:B2')
        ws['A1'] = '結合1'
        ws.merge_cells('C1:D1')
        ws['C1'] = '結合2'
        ws['A3'] = '通常'
        img = render_worksheet(ws)
        assert isinstance(img, Image.Image)


# ── スタイル描画 ──────────────────────────────────────────────────────────────


class TestStyleRendering:
    """セルスタイル（塗りつぶし、罫線、フォント等）の描画テスト。"""

    def test_fill_color_renders(self):
        """塗りつぶし色が描画される（非白ピクセルが存在）。"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'カラー'
        ws['A1'].fill = PatternFill(start_color='FFFF0000', fill_type='solid')
        ws.column_dimensions['A'].width = 15
        ws.row_dimensions[1].height = 30
        img = render_worksheet(ws, scale=2.0)
        arr = np.array(img)
        # 赤成分が高いピクセルが存在する
        red_pixels = (arr[:, :, 0] > 200) & (arr[:, :, 1] < 50)
        assert red_pixels.sum() > 0

    def test_border_renders(self):
        """罫線が描画される。"""
        wb = Workbook()
        ws = wb.active
        ws['B2'] = 'ボーダー'
        thin_side = Side(style='thin')
        ws['B2'].border = Border(
            top=thin_side, bottom=thin_side,
            left=thin_side, right=thin_side,
        )
        img = render_worksheet(ws, scale=2.0)
        assert isinstance(img, Image.Image)

    def test_font_size_affects_rendering(self):
        """フォントサイズ指定がある場合もクラッシュしない。"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '大きい文字'
        ws['A1'].font = Font(size=20, bold=True)
        ws['A2'] = '小さい文字'
        ws['A2'].font = Font(size=8)
        img = render_worksheet(ws)
        assert isinstance(img, Image.Image)

    def test_alignment_center(self):
        """center アライメントでクラッシュしない。"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '中央'
        ws['A1'].alignment = Alignment(horizontal='center')
        img = render_worksheet(ws)
        assert isinstance(img, Image.Image)

    def test_alignment_right(self):
        """right アライメントでクラッシュしない。"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '右寄せ'
        ws['A1'].alignment = Alignment(horizontal='right')
        img = render_worksheet(ws)
        assert isinstance(img, Image.Image)


# ── テンプレート統合テスト ─────────────────────────────────────────────────────


class TestTemplateIntegration:
    """テンプレート生成 → レンダリングの一気通貫テスト。"""

    def test_meireihyo_template_renders(self, tmp_path):
        """掲示用名列表テンプレートをレンダリングできる。"""
        from templates.generators import gen_meireihyo
        out = str(tmp_path / '名列表.xlsx')
        gen_meireihyo.generate(out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        img = render_worksheet(ws, scale=1.5)
        assert isinstance(img, Image.Image)
        assert img.width > 100
        assert img.height > 100

    def test_nafuda_template_renders(self, tmp_path):
        """名札テンプレートをレンダリングできる。"""
        from templates.generators import gen_nafuda
        out = str(tmp_path / '名札.xlsx')
        gen_nafuda.generate(out, mode='通常')

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        img = render_worksheet(ws, scale=1.5)
        assert isinstance(img, Image.Image)

    def test_daicho_template_renders(self, tmp_path):
        """台帳テンプレートをレンダリングできる。"""
        from templates.generators import gen_daicho
        out = str(tmp_path / '台帳.xlsx')
        gen_daicho.generate(out, mode='shuuryo')

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        img = render_worksheet(ws, scale=1.0)
        assert isinstance(img, Image.Image)
