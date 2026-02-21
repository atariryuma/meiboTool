"""templates/template_registry.py のユニットテスト"""

import os

from openpyxl import Workbook

from templates.template_registry import (
    _FILE_TO_KEY,
    CATEGORY_ORDER,
    TEMPLATES,
    get_all_templates,
    get_display_groups,
)

# ────────────────────────────────────────────────────────────────────────────
# TEMPLATES dict の構造テスト
# ────────────────────────────────────────────────────────────────────────────

class TestTemplatesDict:
    """TEMPLATES 定数の構造・整合性テスト"""

    def test_all_have_file(self):
        for name, meta in TEMPLATES.items():
            assert 'file' in meta, f'{name}: file がない'

    def test_all_have_type(self):
        for name, meta in TEMPLATES.items():
            assert meta['type'] in ('grid', 'list', 'individual'), (
                f'{name}: type が不正: {meta.get("type")}'
            )

    def test_all_have_category(self):
        for name, meta in TEMPLATES.items():
            assert 'category' in meta, f'{name}: category がない'

    def test_all_categories_in_order(self):
        """全カテゴリが CATEGORY_ORDER に含まれている。"""
        for name, meta in TEMPLATES.items():
            cat = meta.get('category')
            assert cat in CATEGORY_ORDER, (
                f'{name}: category "{cat}" が CATEGORY_ORDER にない'
            )

    def test_all_have_required_columns(self):
        for name, meta in TEMPLATES.items():
            assert 'required_columns' in meta, f'{name}: required_columns がない'
            assert isinstance(meta['required_columns'], list)

    def test_all_have_mandatory_columns(self):
        for name, meta in TEMPLATES.items():
            assert 'mandatory_columns' in meta, f'{name}: mandatory_columns がない'

    def test_file_names_are_xlsx(self):
        for name, meta in TEMPLATES.items():
            assert meta['file'].endswith('.xlsx'), f'{name}: file が .xlsx でない'

    def test_file_names_unique(self):
        files = [meta['file'] for meta in TEMPLATES.values()]
        assert len(files) == len(set(files)), 'ファイル名に重複がある'

    def test_no_disabled_templates(self):
        """現在 enabled=False のテンプレートはない。"""
        disabled = [n for n, m in TEMPLATES.items() if m.get('enabled') is False]
        assert disabled == []

    def test_grid_templates_count(self):
        grids = [n for n, m in TEMPLATES.items() if m['type'] == 'grid']
        assert len(grids) >= 8


# ────────────────────────────────────────────────────────────────────────────
# _FILE_TO_KEY 逆引きインデックス
# ────────────────────────────────────────────────────────────────────────────

class TestFileToKey:
    """_FILE_TO_KEY 逆引き dict のテスト"""

    def test_has_all_entries(self):
        assert len(_FILE_TO_KEY) == len(TEMPLATES)

    def test_lookup(self):
        assert _FILE_TO_KEY['ラベル_色付き.xlsx'] == 'ラベル_色付き'

    def test_reverse_consistency(self):
        for name, meta in TEMPLATES.items():
            assert _FILE_TO_KEY[meta['file']] == name


# ────────────────────────────────────────────────────────────────────────────
# get_all_templates() マージロジック
# ────────────────────────────────────────────────────────────────────────────

def _create_template_xlsx(folder: str, name: str, placeholders: list[str] | None = None):
    """テスト用にダミー .xlsx を作成する。"""
    wb = Workbook()
    ws = wb.active
    if placeholders:
        for i, ph in enumerate(placeholders):
            ws.cell(row=1, column=i + 1, value=ph)
    else:
        ws['A1'] = '{{氏名_1}}'
    wb.save(os.path.join(folder, name))
    wb.close()


class TestGetAllTemplates:
    """get_all_templates() のマージロジックテスト"""

    def test_templates_with_existing_files(self, tmp_path):
        """TEMPLATES に定義されていてファイルも存在 → 結果に含まれる。"""
        # 1つだけ作成
        _create_template_xlsx(str(tmp_path), 'ラベル_色付き.xlsx')
        result = get_all_templates(str(tmp_path))
        assert 'ラベル_色付き' in result
        assert result['ラベル_色付き']['icon'] == '🌸'

    def test_templates_without_files_excluded(self, tmp_path):
        """TEMPLATES にあるがファイルがない → 結果から除外。"""
        # 空フォルダ（ファイルなし）
        result = get_all_templates(str(tmp_path))
        # enabled テンプレートでファイルがないものは除外される
        for name, meta in TEMPLATES.items():
            if meta.get('enabled') is not False:
                assert name not in result, f'{name} はファイルがないのに含まれている'

    def test_sort_by_metadata_present(self, tmp_path):
        """sort_by メタデータがあるテンプレートが正しく返される。"""
        _create_template_xlsx(str(tmp_path), '男女一覧.xlsx')
        result = get_all_templates(str(tmp_path))
        assert '男女一覧' in result
        assert result['男女一覧']['sort_by'] == '性別'

    def test_scan_only_templates_added(self, tmp_path):
        """TEMPLATES にないがフォルダにある → スキャン結果で追加。"""
        _create_template_xlsx(
            str(tmp_path), 'カスタム帳票.xlsx',
            ['{{氏名_1}}', '{{氏名_2}}', '{{氏名_3}}'],
        )
        result = get_all_templates(str(tmp_path))
        assert 'カスタム帳票' in result
        assert result['カスタム帳票']['type'] == 'grid'

    def test_scan_does_not_duplicate_templates(self, tmp_path):
        """TEMPLATES に登録済みのファイルはスキャンで重複追加されない。"""
        _create_template_xlsx(str(tmp_path), 'ラベル_色付き.xlsx')
        _create_template_xlsx(str(tmp_path), 'カスタム帳票.xlsx', ['{{氏名_1}}'])
        result = get_all_templates(str(tmp_path))
        # ラベル_色付き は TEMPLATES 版が使われる
        assert result['ラベル_色付き']['icon'] == '🌸'
        # カスタム帳票はスキャン版
        assert 'カスタム帳票' in result

    def test_returns_dict(self, tmp_path):
        result = get_all_templates(str(tmp_path))
        assert isinstance(result, dict)


# ────────────────────────────────────────────────────────────────────────────
# get_display_groups()
# ────────────────────────────────────────────────────────────────────────────

class TestGetDisplayGroups:
    """get_display_groups() のテスト（template_dir='' で TEMPLATES のみ使用）"""

    def test_returns_list_of_tuples(self):
        groups = get_display_groups()
        assert isinstance(groups, list)
        for cat_name, items in groups:
            assert isinstance(cat_name, str)
            assert isinstance(items, list)

    def test_categories_in_correct_order(self):
        groups = get_display_groups()
        cat_names = [cat for cat, _ in groups]
        # CATEGORY_ORDER の順序で並んでいることを確認
        order = {cat: i for i, cat in enumerate(CATEGORY_ORDER)}
        indices = [order.get(c, 999) for c in cat_names]
        assert indices == sorted(indices)

    def test_danjo_ichiran_included(self):
        """男女一覧が表示グループに含まれる（enabled=True）。"""
        groups = get_display_groups()
        all_keys = []
        for _cat, items in groups:
            for _display, key, _icon, _desc in items:
                all_keys.append(key)
        assert '男女一覧' in all_keys

    def test_all_enabled_templates_present(self):
        """enabled なテンプレートは全て表示グループに含まれる。"""
        groups = get_display_groups()
        all_keys = set()
        for _cat, items in groups:
            for _display, key, _icon, _desc in items:
                all_keys.add(key)
        for name, meta in TEMPLATES.items():
            if meta.get('enabled', True):
                assert name in all_keys, f'{name} が表示グループにない'

    def test_item_tuple_structure(self):
        """各アイテムが (display_name, key, icon, description) の4要素。"""
        groups = get_display_groups()
        for _cat, items in groups:
            for item in items:
                assert len(item) == 4
                display_name, key, icon, desc = item
                assert isinstance(display_name, str)
                assert isinstance(key, str)
                assert isinstance(icon, str)
                assert isinstance(desc, str)

    def test_has_expected_categories(self):
        groups = get_display_groups()
        cat_names = {cat for cat, _ in groups}
        assert '名札・ラベル' in cat_names
        assert '名簿・出欠表' in cat_names


# ────────────────────────────────────────────────────────────────────────────
# CATEGORY_ORDER
# ────────────────────────────────────────────────────────────────────────────

class TestCategoryOrder:
    """CATEGORY_ORDER 定数のテスト"""

    def test_is_list(self):
        assert isinstance(CATEGORY_ORDER, list)

    def test_contains_expected(self):
        assert '名札・ラベル' in CATEGORY_ORDER
        assert 'その他' in CATEGORY_ORDER

    def test_other_is_last(self):
        assert CATEGORY_ORDER[-1] == 'その他'
