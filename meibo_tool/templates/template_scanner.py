"""テンプレート自動検出スキャナー

テンプレートフォルダ内の .xlsx ファイルを解析し、
プレースホルダーからタイプ・cards_per_page・orientation 等のメタデータを自動検出する。

Excelで手作りしたテンプレートをフォルダに置くだけで自動認識される。
template_registry.py の TEMPLATES にエントリがある場合はそちらが優先（オーバーライド）。
"""

from __future__ import annotations

import os
import re
from typing import Any

from openpyxl import load_workbook

# プレースホルダー正規表現
_PH_RE = re.compile(r'\{\{(.+?)\}\}')
# 番号付きプレースホルダー: {{field_N}}
_NUMBERED_RE = re.compile(r'^(.+)_(\d+)$')

# 共通ヘッダーフィールド（データフィールドではない）
_SPECIAL_KEYS: frozenset[str] = frozenset({
    '年度', '年度和暦', '学校名', '担任名',
})

# individual タイプの判定に使う個人情報フィールド
_INDIVIDUAL_MARKERS: frozenset[str] = frozenset({
    '郵便番号', '保護者正式名', '保護者正式名かな',
    '保護者続柄', '緊急連絡先', '保護者名', '保護者名かな',
})

# タイプ別デフォルトアイコン
_DEFAULT_ICONS: dict[str, str] = {
    'grid': '📋',
    'list': '📋',
    'individual': '📄',
}


def scan_template_file(path: str) -> dict[str, Any] | None:
    """単一の .xlsx ファイルを解析してテンプレートメタデータを返す。

    プレースホルダーが1つも見つからない場合は None を返す（テンプレートではない）。

    Returns:
        dict with keys: file, type, cards_per_page, orientation,
                        use_formal_name, required_columns, mandatory_columns,
                        icon, description
        or None if not a template.
    """
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return None

    try:
        ws = wb.active
        if ws is None:
            return None

        placeholders: set[str] = set()
        numbered_max: dict[str, int] = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                    for match in _PH_RE.finditer(cell.value):
                        key = match.group(1)
                        placeholders.add(key)
                        m = _NUMBERED_RE.match(key)
                        if m:
                            base = m.group(1)
                            num = int(m.group(2))
                            numbered_max[base] = max(
                                numbered_max.get(base, 0), num
                            )

        if not placeholders:
            return None

        # ── タイプ判定 ────────────────────────────────────────────────────
        if numbered_max:
            tmpl_type = 'grid'
            cards_per_page = max(numbered_max.values())
        else:
            base_keys = {
                _NUMBERED_RE.match(k).group(1) if _NUMBERED_RE.match(k) else k
                for k in placeholders
            }
            individual_count = len(base_keys & _INDIVIDUAL_MARKERS)
            tmpl_type = 'individual' if individual_count >= 2 else 'list'
            cards_per_page = None

        # ── orientation ───────────────────────────────────────────────────
        orientation = getattr(ws.page_setup, 'orientation', None) or 'portrait'

        # ── use_formal_name ───────────────────────────────────────────────
        formal_in_plain = bool(
            {'正式氏名', '正式氏名かな'} & placeholders
        )
        formal_in_numbered = any(
            k.startswith('正式氏名') for k in numbered_max
        )
        use_formal = formal_in_plain or formal_in_numbered

        # ── required_columns ──────────────────────────────────────────────
        required: set[str] = set()
        for key in placeholders:
            m = _NUMBERED_RE.match(key)
            base = m.group(1) if m else key
            if base not in _SPECIAL_KEYS and base not in ('組', '出席番号'):
                required.add(base)

        # ── メタデータ構築 ────────────────────────────────────────────────
        filename = os.path.basename(path)
        display_name = os.path.splitext(filename)[0]

        return {
            'file': filename,
            'type': tmpl_type,
            'cards_per_page': cards_per_page,
            'orientation': orientation,
            'use_formal_name': use_formal,
            'required_columns': sorted(required),
            'mandatory_columns': ['組', '出席番号'],
            'icon': _DEFAULT_ICONS.get(tmpl_type, '📋'),
            'description': f'{display_name}（自動検出）',
        }
    finally:
        wb.close()


# ── mtime キャッシュ ──────────────────────────────────────────────────────

_cache: dict[str, tuple[float, dict[str, Any]]] = {}  # path → (mtime, meta)


def _get_cached(path: str) -> dict[str, Any] | None:
    """mtime が変わっていなければキャッシュを返す。変わっていれば None。"""
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return None
    cached = _cache.get(path)
    if cached and cached[0] == mtime:
        return cached[1]
    return None


def _set_cache(path: str, meta: dict[str, Any]) -> None:
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return
    _cache[path] = (mtime, meta)


def scan_template_folder(template_dir: str) -> dict[str, dict[str, Any]]:
    """テンプレートフォルダ内の全 .xlsx を解析してメタデータ辞書を返す。

    キーはファイル名（拡張子なし）、値はメタデータ dict。
    プレースホルダーのないファイルは除外される。
    """
    result: dict[str, dict[str, Any]] = {}
    if not os.path.isdir(template_dir):
        return result

    for fname in os.listdir(template_dir):
        if not fname.endswith('.xlsx') or fname.startswith('~$'):
            continue
        path = os.path.join(template_dir, fname)
        if not os.path.isfile(path):
            continue

        # キャッシュチェック
        cached = _get_cached(path)
        if cached is not None:
            key = os.path.splitext(fname)[0]
            result[key] = cached
            continue

        meta = scan_template_file(path)
        if meta is not None:
            key = os.path.splitext(fname)[0]
            result[key] = meta
            _set_cache(path, meta)

    return result


def clear_cache() -> None:
    """テスト用: キャッシュをクリアする。"""
    _cache.clear()
