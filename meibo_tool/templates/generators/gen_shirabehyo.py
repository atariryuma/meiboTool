"""調べ表テンプレート生成スクリプト

SPEC.md §4.5 参照。
A4 縦・6列×10行グリッド（最大 60 名）。
タイトル部:
  左上 空白大セル（手書きで調査項目名を記入）
  右上 「調べ / 期限：」（斜線セル風）
  サブ: 「{{学年}}年{{組}}組」 「担任：{{担任名}}」

各セル（1名分）:
  上段（小）: {{出席番号_N}}  {{氏名かな_N}}
  中段（大）: {{氏名_N}}
  下段（空白）: 記入欄

プレースホルダー番号の割り当て:
  列1: _1, _7, _13, _19, _25, _31, _37, _43, _49, _55
  列2: _2, _8, ...
  列j, 行i: N = (i-1)*6 + j  （行優先・1-indexed）
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ── スタイル定数 ──────────────────────────────────────────────────────────

FONT_FAMILY = 'IPAmj明朝'
_HEADER_BG  = 'F0F0F0'   # サブヘッダー行のみ薄グレー（データ行は白/塗りなし）

FILL_HEADER = PatternFill(fill_type='solid', fgColor=_HEADER_BG)

_THIN   = Side(style='thin',   color='000000')
_MEDIUM = Side(style='medium', color='808080')

BORDER_THIN   = Border(top=_THIN,   bottom=_THIN,   left=_THIN,   right=_THIN)
BORDER_MEDIUM = Border(top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM, right=_MEDIUM)

ALIGN_C  = Alignment(horizontal='center', vertical='center',  wrap_text=True)
ALIGN_CL = Alignment(horizontal='left',   vertical='center',  wrap_text=True)
ALIGN_CB = Alignment(horizontal='center', vertical='bottom',  wrap_text=True)
ALIGN_CT = Alignment(horizontal='center', vertical='top',     wrap_text=True)

FONT_TITLE = Font(name=FONT_FAMILY, size=14, bold=True)
FONT_SUB   = Font(name=FONT_FAMILY, size=9)
FONT_NO    = Font(name=FONT_FAMILY, size=8)
FONT_KANA  = Font(name=FONT_FAMILY, size=8)
FONT_NAME  = Font(name=FONT_FAMILY, size=12, bold=True)

COLS = 6
ROWS = 10


def _ph(base: str, n: int) -> str:
    return '{{' + base + '_' + str(n) + '}}'


def _cell(ws, row: int, col: int, *, value=None, font=None, fill=None,
          border=None, alignment=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border
    if alignment is not None:
        c.alignment = alignment
    return c


def _apply_print(ws) -> None:
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.20, right=0.20, top=0.20, bottom=0.20, header=0.10, footer=0.10
    )
    ws.print_options.horizontalCentered = True


def generate(output_path: str) -> None:
    """調べ表テンプレートを生成して output_path に保存する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = '調べ表'

    # ── 列幅・行高 ──────────────────────────────────────────────────────────
    # タイトル行(1): 高 30、サブ行(2): 高 16
    # データ行セット（3行/1名）: かな行(高10) + 氏名行(高22) + 記入行(高16)
    # 列幅: 各セルの Data 列は均等
    col_letters = list('ABCDEFGH')  # A〜F=データ, G=余白（未使用）
    data_width = 12.5
    for _i, c in enumerate(col_letters[:COLS]):
        ws.column_dimensions[c].width = data_width

    ws.row_dimensions[1].height = 30   # タイトル
    ws.row_dimensions[2].height = 16   # サブヘッダー

    # ── タイトル行 ──────────────────────────────────────────────────────────
    # 左半分: 空白（手書き調査項目名用）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    _cell(ws, 1, 1, value='（調査項目名）', font=FONT_TITLE,
          alignment=Alignment(horizontal='center', vertical='center'))

    # 右半分: 「調べ　期限：　　　」
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
    _cell(ws, 1, 4, value='調べ　期限：　　　　', font=FONT_TITLE,
          border=BORDER_THIN, alignment=ALIGN_C)

    # ── サブヘッダー行（薄グレー）────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    _cell(ws, 2, 1,
          value='{{学年}}年{{組}}組',
          font=FONT_SUB, fill=FILL_HEADER, alignment=ALIGN_CL)

    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=6)
    _cell(ws, 2, 4,
          value='担任：{{担任名}}',
          font=FONT_SUB, fill=FILL_HEADER,
          alignment=Alignment(horizontal='right', vertical='center'))

    # ── データグリッド（3行セット × 10行 = 30行、列は6列） ────────────────
    # 番号の割り当て: 行優先（行 i=0..9, 列 j=0..5） → N = i*6 + j + 1
    for i in range(ROWS):
        kana_row = 3 + i * 3
        name_row = 4 + i * 3
        rec_row  = 5 + i * 3

        ws.row_dimensions[kana_row].height = 10
        ws.row_dimensions[name_row].height = 22
        ws.row_dimensions[rec_row].height  = 16

        for j in range(COLS):
            n = i * COLS + j + 1
            col = j + 1

            # かな行: 番号 + かな（横並び）
            _cell(ws, kana_row, col,
                  value=f'{_ph("出席番号", n)} {_ph("氏名かな", n)}',
                  font=FONT_KANA,
                  border=BORDER_THIN,
                  alignment=ALIGN_CB)

            # 氏名行: 氏名（大きめ中央）
            _cell(ws, name_row, col,
                  value=_ph('氏名', n), font=FONT_NAME,
                  border=BORDER_THIN, alignment=ALIGN_C)

            # 記入欄: 空白（手書き記入用）
            _cell(ws, rec_row, col,
                  border=BORDER_THIN)

    _apply_print(ws)

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f'Generated: {output_path}')


if __name__ == '__main__':
    _root = Path(__file__).resolve().parents[3]
    generate(str(_root / 'テンプレート' / '調べ表.xlsx'))
