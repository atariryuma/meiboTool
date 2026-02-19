"""様式R2年度名簿 型.xls → xlsx テンプレート変換スクリプト

元の .xls ファイルから書式（フォント・色・列幅・行高・結合セル・罫線・背景色）を
そのまま引き継いで openpyxl 形式の .xlsx テンプレートを生成する。

使用方法（meibo_tool/ ディレクトリから実行）:
    python -m templates.generators.gen_from_legacy

生成されるテンプレート:
    名札_通常.xlsx       ← ラベル 大 (sheet 7)  cards_per_page=20（両面同一生徒）
    名札_装飾あり.xlsx   ← ラベル大色付 (sheet 9) cards_per_page=40（青/赤・2-up）
    ラベル_大2.xlsx      ← ラベル 大 (2) (sheet 1) cards_per_page=40
    ラベル_小.xlsx       ← ラベル 小 (sheet 6)   cards_per_page=20（両面 2-up）
    ラベル_特大.xlsx     ← ラベル 特大 (sheet 8) cards_per_page=40
    横名簿.xlsx          ← 横名簿 (sheet 2)       cards_per_page=40
    縦一週間.xlsx        ← 縦一週間 (sheet 3)     cards_per_page なし（40スロット固定）
"""

from __future__ import annotations

import os
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell as _MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ── xls シートインデックス ────────────────────────────────────────────────────
_SI_LABEL_LARGE2    = 1   # ラベル 大 (2)   → ラベル_大2.xlsx
_SI_YOKONABOKU      = 2   # 横名簿           → 横名簿.xlsx
_SI_TATE_ISSHUUKAN  = 3   # 縦一週間         → 縦一週間.xlsx
_SI_LABEL_SMALL     = 6   # ラベル 小        → ラベル_小.xlsx
_SI_LABEL_LARGE     = 7   # ラベル 大        → 名札_通常.xlsx（上書き）
_SI_LABEL_XL        = 8   # ラベル 特大      → ラベル_特大.xlsx
_SI_LABEL_COLOR     = 9   # ラベル大色付     → 名札_装飾あり.xlsx（上書き）

# xlrd 罫線スタイル番号 → openpyxl スタイル文字列
_LINE_STYLE: dict[int, str | None] = {
    0: None,          1: 'thin',          2: 'medium',
    3: 'dashed',      4: 'dotted',        5: 'thick',
    6: 'double',      7: 'hair',          8: 'mediumDashed',
    9: 'dashDot',     10: 'mediumDashDot', 11: 'dashDotDot',
    12: 'mediumDashDotDot',               13: 'slantDashDot',
}

# クラス名ヘッダープレースホルダー（fill_placeholders で置換される）
_HEADER_PH = '{{学年}}年{{組}}組'


# ────────────────────────────────────────────────────────────────────────────
# xlrd → openpyxl 書式変換ユーティリティ
# ────────────────────────────────────────────────────────────────────────────

def _color_to_hex(wb_xls, colour_index: int) -> str | None:
    """xlrd colour_index → 'RRGGBB' 文字列（# なし）。automatic/system 色は None。"""
    if colour_index in (32767, 64, 65, -1):
        return None
    rgb = wb_xls.colour_map.get(colour_index)
    if rgb:
        return '{:02X}{:02X}{:02X}'.format(*rgb)
    return None


def _make_side(wb_xls, line_style: int, colour_index: int) -> Side:
    style = _LINE_STYLE.get(line_style)
    if not style:
        return Side()
    color = _color_to_hex(wb_xls, colour_index) or '000000'
    return Side(border_style=style, color=color)


def _xf_to_font(wb_xls, xf_idx: int) -> Font:
    xf = wb_xls.xf_list[xf_idx]
    f = wb_xls.font_list[xf.font_index]
    color = _color_to_hex(wb_xls, f.colour_index) or '000000'
    return Font(
        name=f.name,
        size=f.height / 20,   # twips → pt
        bold=bool(f.bold),
        color=color,
    )


def _xf_to_border(wb_xls, xf_idx: int) -> Border:
    xf = wb_xls.xf_list[xf_idx]
    b = xf.border
    return Border(
        left=_make_side(wb_xls, b.left_line_style,   b.left_colour_index),
        right=_make_side(wb_xls, b.right_line_style,  b.right_colour_index),
        top=_make_side(wb_xls, b.top_line_style,    b.top_colour_index),
        bottom=_make_side(wb_xls, b.bottom_line_style, b.bottom_colour_index),
    )


def _xf_to_fill(wb_xls, xf_idx: int) -> PatternFill | None:
    xf = wb_xls.xf_list[xf_idx]
    bg = xf.background
    if bg.fill_pattern == 1:
        fg = _color_to_hex(wb_xls, bg.pattern_colour_index)
        if fg:
            return PatternFill(fill_type='solid', fgColor=fg)
    return None


def _xf_to_alignment(wb_xls, xf_idx: int) -> Alignment:
    _HA = {
        0: 'general', 1: 'left', 2: 'center', 3: 'right',
        4: 'fill', 5: 'justify', 6: 'centerContinuous',
    }
    _VA = {0: 'top', 1: 'center', 2: 'bottom', 3: 'justify', 4: 'distributed'}
    a = wb_xls.xf_list[xf_idx].alignment
    return Alignment(
        horizontal=_HA.get(a.hor_align, 'general'),
        vertical=_VA.get(a.vert_align, 'bottom'),
        wrap_text=bool(a.text_wrapped),
        text_rotation=a.rotation if a.rotation else 0,
    )


def _copy_dims(ws, sh, wb_xls) -> None:
    """列幅・行高を xlrd sheet → openpyxl ws に転写する。"""
    for ci in range(sh.ncols):
        cinfo = sh.colinfo_map.get(ci)
        if cinfo is not None:
            ws.column_dimensions[get_column_letter(ci + 1)].width = cinfo.width / 256
    for ri in range(sh.nrows):
        rinfo = sh.rowinfo_map.get(ri)
        if rinfo is not None and rinfo.height > 0:
            ws.row_dimensions[ri + 1].height = rinfo.height / 20


def _apply_merges(ws, sh) -> None:
    """結合セルを xlrd → openpyxl に転写する。
    xlrd 形式: (rlo, rhi_excl, clo, chi_excl)
    openpyxl  : start_row=rlo+1, end_row=rhi_excl, start_col=clo+1, end_col=chi_excl
    """
    for (r1, r2, c1, c2) in sh.merged_cells:
        ws.merge_cells(
            start_row=r1 + 1, end_row=r2,
            start_column=c1 + 1, end_column=c2,
        )


def _clone_all_cells(ws, sh, wb_xls) -> None:
    """全セルの値・書式を xlrd sheet → openpyxl ws に複製する。
    結合セルの非左上セル（MergedCell）は書き込みをスキップする。
    """
    for r in range(sh.nrows):
        for c in range(sh.ncols):
            cell = ws.cell(row=r + 1, column=c + 1)
            if isinstance(cell, _MergedCell):
                continue
            xf_idx = sh.cell_xf_index(r, c)
            raw = sh.cell_value(r, c)
            cell.value = raw if raw != '' else None
            cell.font = _xf_to_font(wb_xls, xf_idx)
            cell.border = _xf_to_border(wb_xls, xf_idx)
            cell.alignment = _xf_to_alignment(wb_xls, xf_idx)
            fill = _xf_to_fill(wb_xls, xf_idx)
            if fill is not None:
                cell.fill = fill


def _set_val(ws, row_0: int, col_0: int, value) -> None:
    """0-indexed 座標でセルに値を設定する。MergedCell は無視。"""
    cell = ws.cell(row=row_0 + 1, column=col_0 + 1)
    if not isinstance(cell, _MergedCell):
        cell.value = value


def _setup_print(ws, orientation: str = 'portrait') -> None:
    """A4 用紙・用紙内に収まるよう印刷設定を適用する。"""
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.20, right=0.20, top=0.20, bottom=0.20,
        header=0.10, footer=0.10,
    )


def _ph(base: str, n: int) -> str:
    """GridGenerator 用番号付きプレースホルダーを返す: {{base_n}}"""
    return '{{' + base + '_' + str(n) + '}}'


def _save(wb, output_path: str, label: str) -> None:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f'Generated ({label}): {output_path}')


# ────────────────────────────────────────────────────────────────────────────
# テンプレート生成関数
# ────────────────────────────────────────────────────────────────────────────

def gen_nafuda_large(wb_xls, output_path: str) -> None:
    """ラベル 大 (sheet 7) → 名札_通常.xlsx

    構造: 7列 × 23行
      - row 0     : クラス名ヘッダー（col 1, merged col 2）
      - rows 2-21 : 1行=1生徒。col 1(青)・col 4(赤) に同一生徒名（両面印刷用）
      - row 22    : フッター（col 1, merged col 2）
    cards_per_page=20（1ページ20人、1-20 → ページ1、21-40 → ページ2）
    """
    sh = wb_xls.sheets()[_SI_LABEL_LARGE]
    wb = Workbook()
    ws = wb.active
    ws.title = 'ラベル'

    _copy_dims(ws, sh, wb_xls)
    _apply_merges(ws, sh)
    _clone_all_cells(ws, sh, wb_xls)

    # ヘッダー（row 0, col 1 が結合の左上）
    _set_val(ws, 0, 1, _HEADER_PH)

    # データ行 2-21: 左（青）と右（赤）に同一生徒プレースホルダー
    for i in range(20):
        r = i + 2    # rows 2-21 (0-indexed)
        n = i + 1    # N = 1-20
        _set_val(ws, r, 1, _ph('氏名', n))   # 青（表面）
        _set_val(ws, r, 4, _ph('氏名', n))   # 赤（裏面）

    _setup_print(ws, orientation='portrait')
    _save(wb, output_path, '名札_通常（ラベル大）')


def gen_nafuda_color(wb_xls, output_path: str) -> None:
    """ラベル大色付 (sheet 9) → 名札_装飾あり.xlsx

    構造: 24列 × 103行
      - row 1           : クラス名ヘッダー（col 2 merged / col 14 merged、左右 2-up）
      - rows 3,6,...,60 : データ行（3行1組の中央行）
                          col 2 =青(1-20), col 8 =赤(21-40)
                          col 14=青(1-20), col 20=赤(21-40) ← 2-up 複製
    cards_per_page=40（1ページ40人 = 左20人×右20人×両面）
    """
    sh = wb_xls.sheets()[_SI_LABEL_COLOR]
    wb = Workbook()
    ws = wb.active
    ws.title = 'ラベル'

    _copy_dims(ws, sh, wb_xls)
    _apply_merges(ws, sh)
    _clone_all_cells(ws, sh, wb_xls)

    # ヘッダー（左半・右半）
    _set_val(ws, 1, 2,  _HEADER_PH)   # 左半（col 2, merged col 3）
    _set_val(ws, 1, 14, _HEADER_PH)   # 右半（col 14, merged col 15）

    # データ: 20グループ × 3行、データ行 = 3*k (k=1..20)
    for k in range(1, 21):
        r = 3 * k    # rows 3, 6, 9, ..., 60 (0-indexed)
        _set_val(ws, r, 2,  _ph('氏名', k))        # 左前（青）
        _set_val(ws, r, 8,  _ph('氏名', k + 20))   # 左後（赤）
        _set_val(ws, r, 14, _ph('氏名', k))        # 右前（青、2-up複製）
        _set_val(ws, r, 20, _ph('氏名', k + 20))   # 右後（赤、2-up複製）

    _setup_print(ws, orientation='portrait')
    _save(wb, output_path, '名札_装飾あり（ラベル大色付）')


def gen_nafuda_large2(wb_xls, output_path: str) -> None:
    """ラベル 大 (2) (sheet 1) → ラベル_大2.xlsx

    構造: 7列 × 180行
      - rows 0-1  : クラス名ヘッダー（col 1 merged col 2-5、2行結合）
      - rows 2-21 : col 1=1-20（黒）、col 4=21-40（黒）
    cards_per_page=40（左=1-20、右=21-40）
    """
    sh = wb_xls.sheets()[_SI_LABEL_LARGE2]
    wb = Workbook()
    ws = wb.active
    ws.title = 'ラベル'

    _copy_dims(ws, sh, wb_xls)
    _apply_merges(ws, sh)
    _clone_all_cells(ws, sh, wb_xls)

    # ヘッダー（rows 0-1 merged、col 1 が左上）
    _set_val(ws, 0, 1, _HEADER_PH)

    # データ行 2-21
    for i in range(20):
        r = i + 2
        n = i + 1
        _set_val(ws, r, 1, _ph('氏名', n))        # 左（1-20）
        _set_val(ws, r, 4, _ph('氏名', n + 20))   # 右（21-40）

    _setup_print(ws, orientation='portrait')
    _save(wb, output_path, 'ラベル_大2')


def gen_nafuda_small(wb_xls, output_path: str) -> None:
    """ラベル 小 (sheet 6) → ラベル_小.xlsx

    構造: 14列 × 24行
      - row 0     : クラス名ヘッダー（col 1 merged col 2）
      - rows 2-21 : 1行=1生徒 × 4ポジション（両面 2-up）
                    col 1=青前, col 4=赤後, col 8=青前(複製), col 11=赤後(複製)
    cards_per_page=20
    """
    sh = wb_xls.sheets()[_SI_LABEL_SMALL]
    wb = Workbook()
    ws = wb.active
    ws.title = 'ラベル'

    _copy_dims(ws, sh, wb_xls)
    _apply_merges(ws, sh)
    _clone_all_cells(ws, sh, wb_xls)

    # ヘッダー
    _set_val(ws, 0, 1, _HEADER_PH)

    # データ行 2-21: 同一生徒を4か所に配置
    for i in range(20):
        r = i + 2
        n = i + 1
        _set_val(ws, r, 1,  _ph('氏名', n))   # col B 青前
        _set_val(ws, r, 4,  _ph('氏名', n))   # col E 赤後（同一生徒）
        _set_val(ws, r, 8,  _ph('氏名', n))   # col I 青前（2-up複製）
        _set_val(ws, r, 11, _ph('氏名', n))   # col L 赤後（2-up複製）

    _setup_print(ws, orientation='portrait')
    _save(wb, output_path, 'ラベル_小')


def gen_nafuda_xl(wb_xls, output_path: str) -> None:
    """ラベル 特大 (sheet 8) → ラベル_特大.xlsx

    構造: 4列 × 20行
      - rows 0-19 : col 0=1-20（黒 72pt）、col 2=21-40（黒 72pt）
      ヘッダー行なし（全行がデータ）
    cards_per_page=40（左=1-20、右=21-40）
    """
    sh = wb_xls.sheets()[_SI_LABEL_XL]
    wb = Workbook()
    ws = wb.active
    ws.title = 'ラベル'

    _copy_dims(ws, sh, wb_xls)
    _apply_merges(ws, sh)
    _clone_all_cells(ws, sh, wb_xls)

    # ヘッダーなし。全20行がデータ行。
    for i in range(20):
        r = i   # rows 0-19 (0-indexed)
        n = i + 1
        _set_val(ws, r, 0, _ph('氏名', n))        # 左（1-20）
        _set_val(ws, r, 2, _ph('氏名', n + 20))   # 右（21-40）

    _setup_print(ws, orientation='portrait')
    _save(wb, output_path, 'ラベル_特大')


def gen_yokonaboku(wb_xls, output_path: str) -> None:
    """横名簿 (sheet 2) → 横名簿.xlsx

    構造: 25列 × 23行
      - row 0      : クラス名ヘッダー（col 0 merged col 1）
      - row 1      : 出席日マーク行（'/' が10列×2 = 20か所）
      - rows 3-22  : データ行（20行 × 2列 = 40人）
                     col 0=番号L, col 1=氏名L, col 13=番号R, col 14=氏名R
    cards_per_page=40
    """
    sh = wb_xls.sheets()[_SI_YOKONABOKU]
    wb = Workbook()
    ws = wb.active
    ws.title = '横名簿'

    _copy_dims(ws, sh, wb_xls)
    _apply_merges(ws, sh)
    _clone_all_cells(ws, sh, wb_xls)

    # ヘッダー（row 0, col 0 merged col 1）
    _set_val(ws, 0, 0, _HEADER_PH)

    # データ行 3-22 (0-indexed)
    for i in range(20):
        r = i + 3    # rows 3-22
        n = i + 1
        _set_val(ws, r, 0,  _ph('出席番号', n))        # 左 番号
        _set_val(ws, r, 1,  _ph('氏名', n))            # 左 氏名
        _set_val(ws, r, 13, _ph('出席番号', n + 20))   # 右 番号
        _set_val(ws, r, 14, _ph('氏名', n + 20))       # 右 氏名

    _setup_print(ws, orientation='landscape')
    _save(wb, output_path, '横名簿')


def gen_tate_isshuukan(wb_xls, output_path: str) -> None:
    """縦一週間 (sheet 3) → 縦一週間.xlsx

    構造: 30列 × 46行（GridGenerator、カードなし = 40スロット固定）
      - row 0      : クラス名ヘッダー（col 1）、col 5-29 に '/' 出欠日ヘッダー
      - rows 1-4   : サブヘッダー（白文字の隠し参照値）
      - rows 5-44  : データ行（1行=1生徒）
                     col 0=出席番号、col 4=氏名、col 5-29=出欠欄（空欄）
      - row 45     : フッター行
    cards_per_page なし（単一ページ、40スロット）
    """
    sh = wb_xls.sheets()[_SI_TATE_ISSHUUKAN]
    wb = Workbook()
    ws = wb.active
    ws.title = '縦一週間'

    _copy_dims(ws, sh, wb_xls)
    _apply_merges(ws, sh)
    _clone_all_cells(ws, sh, wb_xls)

    # ヘッダー（row 0, col 1）
    _set_val(ws, 0, 1, _HEADER_PH)

    # データ行 5-44 (0-indexed): 40スロット
    for i in range(40):
        r = i + 5    # rows 5-44
        n = i + 1
        _set_val(ws, r, 0, _ph('出席番号', n))   # col A 出席番号
        _set_val(ws, r, 4, _ph('氏名', n))       # col E 氏名（広い列）

    _setup_print(ws, orientation='portrait')
    _save(wb, output_path, '縦一週間')


# ────────────────────────────────────────────────────────────────────────────
# 公開 API
# ────────────────────────────────────────────────────────────────────────────

def generate(template_dir: str, xls_path: str) -> None:
    """
    様式R2年度名簿 型.xls から xlsx テンプレートを一括生成する。

    Args:
        template_dir: 出力先テンプレートフォルダの絶対パス
        xls_path:     元の .xls ファイルの絶対パス
    """
    wb_xls = xlrd.open_workbook(
        xls_path, formatting_info=True, encoding_override='cp932'
    )

    gen_nafuda_large(wb_xls,    os.path.join(template_dir, '名札_通常.xlsx'))
    gen_nafuda_color(wb_xls,    os.path.join(template_dir, '名札_装飾あり.xlsx'))
    gen_nafuda_large2(wb_xls,   os.path.join(template_dir, 'ラベル_大2.xlsx'))
    gen_nafuda_small(wb_xls,    os.path.join(template_dir, 'ラベル_小.xlsx'))
    gen_nafuda_xl(wb_xls,       os.path.join(template_dir, 'ラベル_特大.xlsx'))
    gen_yokonaboku(wb_xls,      os.path.join(template_dir, '横名簿.xlsx'))
    gen_tate_isshuukan(wb_xls,  os.path.join(template_dir, '縦一週間.xlsx'))


# ────────────────────────────────────────────────────────────────────────────
# スタンドアロン実行
# ────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    _root = Path(__file__).resolve().parents[3]
    _xls = _root / '様式R2年度名簿 型.xls'
    _tmpl = _root / 'テンプレート'
    generate(str(_tmpl), str(_xls))
