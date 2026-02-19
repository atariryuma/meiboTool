"""掲示用名列表テンプレート生成スクリプト

SPEC.md §4.4 参照。
A4縦・2列（左:No.1〜20、右:No.21〜40）・タイトル・白背景。

1ファイル（掲示用名列表.xlsx）ですべての name_display モードに対応する。
  furigana : 上段かな（小）+ 下段漢字（大）を両方表示
  kanji    : 上段かな行を空白のまま（11pt の細い空間）、下段漢字を表示
  kana     : 上段かな行を空白のまま（11pt の細い空間）、下段にかな値を表示

ふりがな行 (11pt) は常に存在するが、furigana 以外のモードでは空白になるため
視覚的な邪魔にならない細いラインとして機能する。

テンプレート内のプレースホルダー形式（GridGenerator が _fill_numbered() で展開）:
  ヘッダー : {{学年}} {{組}} {{担任名}}
  データ   : {{出席番号_1}}〜{{出席番号_40}}
             {{氏名かな_1}}〜{{氏名かな_40}}  ← かな行（11pt）
             {{氏名_1}}〜{{氏名_40}}          ← 名前行（25pt）
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ────────────────────────────────────────────────────────────────────────────
# スタイル定数（SPEC §4.0.2）
# ────────────────────────────────────────────────────────────────────────────

FONT_FAMILY  = 'IPAmj明朝'
_HEADER_BG   = 'F0F0F0'   # 列ヘッダーのみ薄グレー（データ行は白/塗りなし）

FILL_HEADER  = PatternFill(fill_type='solid', fgColor=_HEADER_BG)

_THIN   = Side(style='thin',   color='000000')
_DOTTED = Side(style='hair',   color='AAAAAA')

BORDER_FULL     = Border(top=_THIN,   bottom=_THIN,   left=_THIN, right=_THIN)
BORDER_KANA_TOP = Border(top=_THIN,   bottom=_DOTTED, left=_THIN, right=_THIN)
BORDER_NAME_BOT = Border(top=_DOTTED, bottom=_THIN,   left=_THIN, right=_THIN)
BORDER_NO_FULL  = Border(top=_THIN,   bottom=_THIN,   left=_THIN, right=_THIN)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_KANA   = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

FONT_TITLE  = Font(name=FONT_FAMILY, size=18, bold=True)
FONT_HEADER = Font(name=FONT_FAMILY, size=9,  bold=True)
FONT_NO     = Font(name=FONT_FAMILY, size=11, bold=False)
FONT_NAME   = Font(name=FONT_FAMILY, size=14, bold=True)
FONT_KANA   = Font(name=FONT_FAMILY, size=9,  bold=False)

MAX_PER_COL = 20   # 1列あたりの最大人数

# 行高（pt）
# かな行を 11pt にすることで 9pt フォントが収まる（フォントサイズ × 1.3 程度）。
# furigana 以外のモードでは空白ラインとして自然な細い区切りになる。
# 合計: タイトル(42) + ヘッダー(18) + 20組×(11+25) = 780pt → A4(842pt) に収まる
KANA_HEIGHT = 11   # pt  ← 全モード共通
NAME_HEIGHT = 25   # pt


# ────────────────────────────────────────────────────────────────────────────
# ヘルパー
# ────────────────────────────────────────────────────────────────────────────

def _ph(base: str, n: int) -> str:
    return '{{' + base + '_' + str(n) + '}}'


def _cell(ws, row: int, col: int, *,
          value=None, font=None, fill=None, border=None, alignment=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border
    if alignment is not None:
        c.alignment = alignment
    return c


def _apply_print_settings(ws) -> None:
    ws.page_setup.paperSize   = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.20, right=0.20,
        top=0.20,  bottom=0.20,
        header=0.10, footer=0.10,
    )
    ws.print_options.horizontalCentered = True


# ────────────────────────────────────────────────────────────────────────────
# テンプレートレイアウト生成（1種のみ）
# ────────────────────────────────────────────────────────────────────────────

def _build(ws) -> None:
    """
    2行/人レイアウト（上段: かな 11pt / 下段: 氏名 25pt）。

    列構成（5列）:
      A: 番号（左）2行分マージ 幅 5
      B: 氏名テキスト（左）上=かな / 下=氏名 幅 27
      C: 区切り 幅 2
      D: 番号（右）2行分マージ 幅 5
      E: 氏名テキスト（右）上=かな / 下=氏名 幅 27

    furigana モード: かな行=かな値、氏名行=漢字値
    kanji    モード: かな行=空白（細い線）、氏名行=漢字値
    kana     モード: かな行=空白（細い線）、氏名行=かな値
                   （generator._fill_numbered が kana モードで氏名→かな値に転写）
    """
    for col, width in zip('ABCDE', [5.0, 27.0, 2.0, 5.0, 27.0], strict=True):
        ws.column_dimensions[col].width = width

    # タイトル行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    _cell(ws, 1, 1,
          value='{{学年}}年{{組}}組　名列表　　担任：{{担任名}}',
          font=FONT_TITLE,
          alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 42

    # ヘッダー行（列ラベルのみ薄グレー）
    for col, label in [(1, '番号'), (2, '氏名'), (4, '番号'), (5, '氏名')]:
        _cell(ws, 2, col,
              value=label, font=FONT_HEADER,
              fill=FILL_HEADER, border=BORDER_FULL, alignment=ALIGN_CENTER)
    ws.row_dimensions[2].height = 18

    for i in range(MAX_PER_COL):
        kana_row = 3 + i * 2
        name_row = 3 + i * 2 + 1
        ln = i + 1    # 左 No. (1–20)
        rn = i + 21   # 右 No. (21–40)

        # ── 番号列（A / D）: 2行マージ ──────────────────────────────────────
        ws.merge_cells(start_row=kana_row, start_column=1,
                       end_row=name_row,  end_column=1)
        _cell(ws, kana_row, 1,
              value=_ph('出席番号', ln), font=FONT_NO,
              border=BORDER_NO_FULL, alignment=ALIGN_CENTER)

        ws.merge_cells(start_row=kana_row, start_column=4,
                       end_row=name_row,  end_column=4)
        _cell(ws, kana_row, 4,
              value=_ph('出席番号', rn), font=FONT_NO,
              border=BORDER_NO_FULL, alignment=ALIGN_CENTER)

        # ── かな行（上段・11pt）──────────────────────────────────────────────
        _cell(ws, kana_row, 2,
              value=_ph('氏名かな', ln), font=FONT_KANA,
              border=BORDER_KANA_TOP, alignment=ALIGN_KANA)
        _cell(ws, kana_row, 5,
              value=_ph('氏名かな', rn), font=FONT_KANA,
              border=BORDER_KANA_TOP, alignment=ALIGN_KANA)
        _cell(ws, kana_row, 3)  # 区切り

        # ── 氏名行（下段・25pt）──────────────────────────────────────────────
        _cell(ws, name_row, 2,
              value=_ph('氏名', ln), font=FONT_NAME,
              border=BORDER_NAME_BOT, alignment=ALIGN_CENTER)
        _cell(ws, name_row, 5,
              value=_ph('氏名', rn), font=FONT_NAME,
              border=BORDER_NAME_BOT, alignment=ALIGN_CENTER)
        _cell(ws, name_row, 3)  # 区切り

        ws.row_dimensions[kana_row].height = KANA_HEIGHT
        ws.row_dimensions[name_row].height = NAME_HEIGHT


# ────────────────────────────────────────────────────────────────────────────
# 公開 API
# ────────────────────────────────────────────────────────────────────────────

def generate(output_path: str) -> None:
    """掲示用名列表テンプレートを生成して output_path に保存する。

    name_display モード（furigana / kanji / kana）は GridGenerator が
    データ差込時に _fill_numbered を通じて適用する。テンプレートは1種のみ。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '名列表'

    _build(ws)
    _apply_print_settings(ws)

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f'Generated: {output_path}')


# ────────────────────────────────────────────────────────────────────────────
# スタンドアロン実行
# ────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    _project_root = Path(__file__).resolve().parents[3]
    generate(str(_project_root / 'テンプレート' / '掲示用名列表.xlsx'))
