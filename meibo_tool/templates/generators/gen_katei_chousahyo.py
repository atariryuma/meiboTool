"""家庭調査票テンプレート生成スクリプト

SPEC.md §4.8 参照。
A4 縦置き・Individual 型（IndividualGenerator が 1 名/シートで複製する）。

差込フィールド:
  正式氏名、正式氏名かな、生年月日、性別、郵便番号、住所（結合）、
  電話番号1、保護者正式名、保護者正式名かな、保護者続柄、緊急連絡先

空白セクション（手書き用）:
  家族構成、保健調査、児童の生活、自由記入、公開レベル、緊急時引渡し欄
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ── スタイル定数 ──────────────────────────────────────────────────────────

FONT_FAMILY = 'IPAmj明朝'
_THIN = Side(style='thin', color='000000')
_MEDIUM = Side(style='medium', color='000000')

BORDER_THIN = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
BORDER_OUTER = Border(top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM, right=_MEDIUM)
BORDER_TOP_MEDIUM = Border(top=_MEDIUM, bottom=_THIN, left=_THIN, right=_THIN)

FILL_HEADER = PatternFill(fill_type='solid', fgColor='F0F0F0')

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_TL = Alignment(horizontal='left', vertical='top', wrap_text=True)

FONT_TITLE = Font(name=FONT_FAMILY, size=14, bold=True)
FONT_SECTION = Font(name=FONT_FAMILY, size=10, bold=True)
FONT_LABEL = Font(name=FONT_FAMILY, size=9, bold=True)
FONT_DATA = Font(name=FONT_FAMILY, size=10)
FONT_SMALL = Font(name=FONT_FAMILY, size=8)
FONT_NOTE = Font(name=FONT_FAMILY, size=7, color='666666')


def _cell(ws, row: int, col: int, *, value=None, font=None, fill=None,
          border=None, alignment=None, width=None, height=None):
    """セルにスタイルを適用するヘルパー。"""
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border
    if alignment is not None:
        c.alignment = alignment
    if width is not None:
        ws.column_dimensions[get_column_letter(col)].width = width
    if height is not None:
        ws.row_dimensions[row].height = height
    return c


def _apply_borders(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    """指定範囲に thin 罫線を適用する。"""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = BORDER_THIN


def _apply_print(ws) -> None:
    """A4 縦置き印刷設定。"""
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.39, right=0.39, top=0.39, bottom=0.39, header=0.20, footer=0.20
    )
    ws.print_options.horizontalCentered = True


def _build_template(ws) -> None:
    """家庭調査票テンプレートの全レイアウトを構築する。"""
    # 全体で 8 列を使用（A〜H）
    col_widths = [4, 10, 10, 12, 12, 10, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── タイトル行 ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _cell(ws, row, 1, value='{{学校名}}　{{年度和暦}}年度　家庭調査票',
          font=FONT_TITLE, alignment=ALIGN_C, height=32)
    row += 1

    # ── 基本情報セクション ──────────────────────────────────────────────
    # 学年・組・番
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    _cell(ws, row, 1, value='{{学年}}年 {{組}}組', font=FONT_DATA,
          alignment=ALIGN_C, border=BORDER_THIN, height=22)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    _cell(ws, row, 3, value='番号: {{出席番号}}', font=FONT_DATA,
          alignment=ALIGN_C, border=BORDER_THIN)
    # 性別
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=5)
    _cell(ws, row, 5, value='{{性別}}', font=FONT_DATA,
          alignment=ALIGN_C, border=BORDER_THIN)
    # 生年月日
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
    _cell(ws, row, 6, value='生年月日: {{生年月日}}', font=FONT_DATA,
          alignment=ALIGN_C, border=BORDER_THIN)
    _apply_borders(ws, row, row, 1, 8)
    row += 1

    # ── 児童氏名 ──────────────────────────────────────────────────────
    _cell(ws, row, 1, value='ふりがな', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN, height=20)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    _cell(ws, row, 2, value='{{正式氏名かな}}', font=FONT_DATA,
          alignment=ALIGN_L, border=BORDER_THIN)
    row += 1

    _cell(ws, row, 1, value='氏名', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN, height=28)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    _cell(ws, row, 2, value='{{正式氏名}}', font=Font(name=FONT_FAMILY, size=14),
          alignment=ALIGN_L, border=BORDER_THIN)
    row += 1

    # ── 住所・電話 ────────────────────────────────────────────────────
    _cell(ws, row, 1, value='〒', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN, height=20)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    _cell(ws, row, 2, value='{{郵便番号}}', font=FONT_DATA,
          alignment=ALIGN_L, border=BORDER_THIN)
    _cell(ws, row, 5, value='電話', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN)
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
    _cell(ws, row, 6, value='{{電話番号1}}', font=FONT_DATA,
          alignment=ALIGN_L, border=BORDER_THIN)
    row += 1

    _cell(ws, row, 1, value='住所', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN, height=28)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    _cell(ws, row, 2, value='{{住所}}', font=FONT_DATA,
          alignment=ALIGN_L, border=BORDER_THIN)
    row += 1

    # ── 保護者情報 ────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _cell(ws, row, 1, value='保護者', font=FONT_SECTION,
          fill=FILL_HEADER, alignment=ALIGN_L, border=BORDER_TOP_MEDIUM, height=20)
    row += 1

    _cell(ws, row, 1, value='続柄', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN, height=20)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
    _cell(ws, row, 2, value='{{保護者続柄}}', font=FONT_DATA,
          alignment=ALIGN_C, border=BORDER_THIN)
    _cell(ws, row, 3, value='ふりがな', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    _cell(ws, row, 4, value='{{保護者正式名かな}}', font=FONT_DATA,
          alignment=ALIGN_L, border=BORDER_THIN)
    row += 1

    _cell(ws, row, 1, value='氏名', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN, height=24)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    _cell(ws, row, 2, value='{{保護者正式名}}',
          font=Font(name=FONT_FAMILY, size=12),
          alignment=ALIGN_L, border=BORDER_THIN)
    row += 1

    _cell(ws, row, 1, value='緊急', font=FONT_LABEL,
          fill=FILL_HEADER, alignment=ALIGN_C, border=BORDER_THIN, height=20)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    _cell(ws, row, 2, value='{{緊急連絡先}}', font=FONT_DATA,
          alignment=ALIGN_L, border=BORDER_THIN)
    row += 1

    # ── 家族構成（空白・手書き用） ─────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _cell(ws, row, 1, value='家族構成', font=FONT_SECTION,
          fill=FILL_HEADER, alignment=ALIGN_L, border=BORDER_TOP_MEDIUM, height=18)
    row += 1

    # 空白行 4 行（手書きスペース）
    for _i in range(4):
        _apply_borders(ws, row, row, 1, 8)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── 保健調査（空白・手書き用） ─────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _cell(ws, row, 1, value='保健調査（アレルギー・既往症・配慮事項）', font=FONT_SECTION,
          fill=FILL_HEADER, alignment=ALIGN_L, border=BORDER_TOP_MEDIUM, height=18)
    row += 1

    for _i in range(3):
        _apply_borders(ws, row, row, 1, 8)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── 児童の生活（空白・手書き用） ───────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _cell(ws, row, 1, value='児童の生活（習い事・通学方法・特記事項）', font=FONT_SECTION,
          fill=FILL_HEADER, alignment=ALIGN_L, border=BORDER_TOP_MEDIUM, height=18)
    row += 1

    for _i in range(3):
        _apply_borders(ws, row, row, 1, 8)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── 緊急時引渡し欄（空白・手書き用） ───────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _cell(ws, row, 1,
          value='緊急時引渡し（引渡し先の氏名・続柄・連絡先）', font=FONT_SECTION,
          fill=FILL_HEADER, alignment=ALIGN_L, border=BORDER_TOP_MEDIUM, height=18)
    row += 1

    for _i in range(3):
        _apply_borders(ws, row, row, 1, 8)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── 公開レベル・自由記入（空白・手書き用） ─────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _cell(ws, row, 1, value='自由記入欄', font=FONT_SECTION,
          fill=FILL_HEADER, alignment=ALIGN_L, border=BORDER_TOP_MEDIUM, height=18)
    row += 1

    for _i in range(3):
        _apply_borders(ws, row, row, 1, 8)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── フッター注釈 ──────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _cell(ws, row, 1,
          value='※ この調査票は学校内での指導目的にのみ使用し、外部に提供することはありません。',
          font=FONT_NOTE, alignment=ALIGN_L, height=14)


# ── 公開 API ──────────────────────────────────────────────────────────────

def generate(output_path: str) -> None:
    """家庭調査票テンプレートを生成して output_path に保存する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'テンプレート'

    _build_template(ws)
    _apply_print(ws)

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f'Generated: {output_path}')


if __name__ == '__main__':
    _root = Path(__file__).resolve().parents[3]
    _tmpl = _root / 'テンプレート'
    generate(str(_tmpl / '家庭調査票.xlsx'))
