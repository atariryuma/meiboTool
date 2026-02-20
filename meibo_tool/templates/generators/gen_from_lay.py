"""スズキ校務 .lay → Excel テンプレート変換器

.lay バイナリレイアウトファイルを読み込み、openpyxl の .xlsx テンプレートに変換する。

使用方法:
    from templates.generators.gen_from_lay import convert
    convert('input.lay', 'output.xlsx')

スタンドアロン実行:
    cd meibo_tool && python -m templates.generators.gen_from_lay INPUT.lay OUTPUT.xlsx
"""

from __future__ import annotations

import bisect
import os
import sys

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell as _MergedCell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from core.lay_parser import (
    LayFile,
    LayoutObject,
    ObjectType,
    parse_lay,
    resolve_field_name,
)

# ── 定数 ─────────────────────────────────────────────────────────────────────

# Excel 列幅の変換係数: 1 Excel 文字幅 ≈ 2.0mm（デフォルト 11pt Calibri 基準）
_MM_TO_EXCEL_WIDTH = 2.0

# Excel 行高さ変換: 1mm = 2.835pt
_MM_TO_PT = 2.835

# 0.25mm/unit（.lay 座標単位）
_UNIT_TO_MM = 0.25

# 近接座標マージの許容範囲（単位数）
_BOUNDARY_TOLERANCE = 3

# 罫線スタイル
_THIN_SIDE = Side(border_style='thin', color='000000')

# デフォルトフォント
_DEFAULT_FONT = 'ＭＳ 明朝'


# ── 境界座標 ─────────────────────────────────────────────────────────────────


def collect_boundaries(
    lay: LayFile,
) -> tuple[list[int], list[int]]:
    """全オブジェクトから X/Y 境界座標を収集する。

    Returns:
        (sorted_x_boundaries, sorted_y_boundaries)
    """
    xs: set[int] = set()
    ys: set[int] = set()

    for obj in lay.objects:
        if obj.obj_type == ObjectType.LINE:
            if obj.line_start:
                xs.add(obj.line_start.x)
                ys.add(obj.line_start.y)
            if obj.line_end:
                xs.add(obj.line_end.x)
                ys.add(obj.line_end.y)
        elif obj.rect:
            xs.add(obj.rect.left)
            xs.add(obj.rect.right)
            ys.add(obj.rect.top)
            ys.add(obj.rect.bottom)

    return sorted(xs), sorted(ys)


def merge_close_boundaries(
    bounds: list[int], tolerance: int = _BOUNDARY_TOLERANCE,
) -> list[int]:
    """許容範囲内の近接座標をマージする。"""
    if not bounds:
        return []
    merged = [bounds[0]]
    for b in bounds[1:]:
        if b - merged[-1] > tolerance:
            merged.append(b)
    return merged


# ── 寸法変換 ─────────────────────────────────────────────────────────────────


def calc_column_widths(x_bounds: list[int]) -> list[float]:
    """X 境界座標から Excel 列幅のリストを計算する。"""
    widths: list[float] = []
    for i in range(len(x_bounds) - 1):
        mm = (x_bounds[i + 1] - x_bounds[i]) * _UNIT_TO_MM
        excel_w = mm / _MM_TO_EXCEL_WIDTH
        widths.append(max(excel_w, 0.5))
    return widths


def calc_row_heights(y_bounds: list[int]) -> list[float]:
    """Y 境界座標から Excel 行高さ (pt) のリストを計算する。"""
    heights: list[float] = []
    for i in range(len(y_bounds) - 1):
        mm = (y_bounds[i + 1] - y_bounds[i]) * _UNIT_TO_MM
        pt = mm * _MM_TO_PT
        heights.append(max(pt, 3.0))
    return heights


# ── セルマッピング ───────────────────────────────────────────────────────────


def _find_cell_index(coord: int, bounds: list[int]) -> int:
    """座標値から対応するセルインデックス (0-based) を返す。

    bounds[i] <= coord < bounds[i+1] ならインデックス i を返す。
    """
    idx = bisect.bisect_right(bounds, coord) - 1
    return max(0, min(idx, len(bounds) - 2))


def object_to_cell_range(
    obj: LayoutObject, x_bounds: list[int], y_bounds: list[int],
) -> tuple[int, int, int, int] | None:
    """オブジェクトの矩形を Excel セル範囲 (row1, col1, row2, col2) に変換する。

    Returns:
        (start_row, start_col, end_row, end_col) 1-indexed,
        または矩形がない場合は None。
    """
    if obj.rect is None:
        return None

    col_s = _find_cell_index(obj.rect.left, x_bounds)
    col_e = _find_cell_index(obj.rect.right - 1, x_bounds)
    row_s = _find_cell_index(obj.rect.top, y_bounds)
    row_e = _find_cell_index(obj.rect.bottom - 1, y_bounds)

    # 1-indexed に変換
    return row_s + 1, col_s + 1, row_e + 1, col_e + 1


# ── アライメント ─────────────────────────────────────────────────────────────

_H_ALIGN_MAP: dict[int, str] = {
    0: 'left',
    1: 'center',
    2: 'right',
}

_V_ALIGN_MAP: dict[int, str] = {
    0: 'top',
    1: 'center',
    2: 'bottom',
}


def _lay_alignment(h_align: int, v_align: int) -> Alignment:
    """lay アライメントコードを openpyxl Alignment に変換する。"""
    return Alignment(
        horizontal=_H_ALIGN_MAP.get(h_align, 'center'),
        vertical=_V_ALIGN_MAP.get(v_align, 'center'),
        wrap_text=True,
    )


# ── ラベル・フィールド配置 ───────────────────────────────────────────────────


def _place_label(
    ws: Worksheet, obj: LayoutObject,
    x_bounds: list[int], y_bounds: list[int],
) -> None:
    """ラベル（固定テキスト）をセルに配置する。"""
    cell_range = object_to_cell_range(obj, x_bounds, y_bounds)
    if not cell_range:
        return

    r1, c1, r2, c2 = cell_range

    # 既にマージ済みセルの場合はスキップ
    cell = ws.cell(row=r1, column=c1)
    if isinstance(cell, _MergedCell):
        return

    if r2 > r1 or c2 > c1:
        ws.merge_cells(
            start_row=r1, start_column=c1,
            end_row=r2, end_column=c2,
        )

    text = obj.prefix + obj.text + obj.suffix
    cell.value = text
    cell.font = Font(
        name=obj.font.name or _DEFAULT_FONT,
        size=obj.font.size_pt,
    )
    cell.alignment = _lay_alignment(obj.h_align, obj.v_align)


def _place_field(
    ws: Worksheet, obj: LayoutObject,
    x_bounds: list[int], y_bounds: list[int],
) -> None:
    """データフィールドを {{プレースホルダー}} として配置する。"""
    cell_range = object_to_cell_range(obj, x_bounds, y_bounds)
    if not cell_range:
        return

    r1, c1, r2, c2 = cell_range

    # 既にマージ済みセルの場合はスキップ
    cell = ws.cell(row=r1, column=c1)
    if isinstance(cell, _MergedCell):
        return

    if r2 > r1 or c2 > c1:
        ws.merge_cells(
            start_row=r1, start_column=c1,
            end_row=r2, end_column=c2,
        )

    logical_name = resolve_field_name(obj.field_id)
    placeholder = f'{obj.prefix}{{{{{logical_name}}}}}{obj.suffix}'
    cell.value = placeholder
    cell.font = Font(
        name=obj.font.name or _DEFAULT_FONT,
        size=obj.font.size_pt,
    )
    cell.alignment = _lay_alignment(obj.h_align, obj.v_align)


# ── 罫線変換 ─────────────────────────────────────────────────────────────────


def _apply_lines(
    ws: Worksheet, lines: list[LayoutObject],
    x_bounds: list[int], y_bounds: list[int],
    num_cols: int, num_rows: int,
) -> None:
    """LINE オブジェクトをセル罫線に変換して適用する。

    水平線 → 対応行の top/bottom border
    垂直線 → 対応列の left/right border
    """
    # border_map[row][col] = {top, bottom, left, right}
    border_map: dict[tuple[int, int], dict[str, bool]] = {}

    def _mark(row: int, col: int, side: str) -> None:
        if 1 <= row <= num_rows and 1 <= col <= num_cols:
            key = (row, col)
            if key not in border_map:
                border_map[key] = {}
            border_map[key][side] = True

    for line in lines:
        if not line.line_start or not line.line_end:
            continue

        x1, y1 = line.line_start.x, line.line_start.y
        x2, y2 = line.line_end.x, line.line_end.y

        # 水平線 (y が同じ)
        if abs(y1 - y2) <= _BOUNDARY_TOLERANCE:
            y = y1
            row_idx = bisect.bisect_right(y_bounds, y - 1)
            min_x = min(x1, x2)
            max_x = max(x1, x2)
            col_start = _find_cell_index(min_x, x_bounds)
            col_end = _find_cell_index(max_x - 1, x_bounds)

            # y が境界線と一致する場合: 上のセルの bottom / 下のセルの top
            closest_bound_idx = bisect.bisect_left(y_bounds, y)

            if closest_bound_idx < len(y_bounds) and \
                    abs(y_bounds[closest_bound_idx] - y) <= _BOUNDARY_TOLERANCE:
                # 境界と一致 → 上側セルの bottom + 下側セルの top
                for c in range(col_start, col_end + 1):
                    if closest_bound_idx > 0:
                        _mark(closest_bound_idx, c + 1, 'bottom')
                    if closest_bound_idx < num_rows:
                        _mark(closest_bound_idx + 1, c + 1, 'top')
            else:
                # 境界と不一致 → 最寄り行の bottom
                for c in range(col_start, col_end + 1):
                    _mark(row_idx, c + 1, 'bottom')

        # 垂直線 (x が同じ)
        elif abs(x1 - x2) <= _BOUNDARY_TOLERANCE:
            x = x1
            min_y = min(y1, y2)
            max_y = max(y1, y2)
            row_start = _find_cell_index(min_y, y_bounds)
            row_end = _find_cell_index(max_y - 1, y_bounds)

            closest_bound_idx = bisect.bisect_left(x_bounds, x)

            if closest_bound_idx < len(x_bounds) and \
                    abs(x_bounds[closest_bound_idx] - x) <= _BOUNDARY_TOLERANCE:
                for r in range(row_start, row_end + 1):
                    if closest_bound_idx > 0:
                        _mark(r + 1, closest_bound_idx, 'right')
                    if closest_bound_idx < num_cols:
                        _mark(r + 1, closest_bound_idx + 1, 'left')
            else:
                col_idx = _find_cell_index(x, x_bounds)
                for r in range(row_start, row_end + 1):
                    _mark(r + 1, col_idx + 1, 'right')

    # border_map を適用
    for (row, col), sides in border_map.items():
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, _MergedCell):
            continue
        existing = cell.border
        cell.border = Border(
            top=_THIN_SIDE if sides.get('top') else existing.top,
            bottom=_THIN_SIDE if sides.get('bottom') else existing.bottom,
            left=_THIN_SIDE if sides.get('left') else existing.left,
            right=_THIN_SIDE if sides.get('right') else existing.right,
        )


# ── 印刷設定 ─────────────────────────────────────────────────────────────────


def _apply_print_settings(ws: Worksheet, lay: LayFile) -> None:
    """ページサイズに基づいて印刷設定を適用する。"""
    ws.page_setup.paperSize = 9  # A4
    if lay.page_width > lay.page_height:
        ws.page_setup.orientation = 'landscape'
    else:
        ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.39, right=0.39, top=0.39, bottom=0.39,
        header=0.20, footer=0.20,
    )
    ws.print_options.horizontalCentered = True


# ── 公開 API ─────────────────────────────────────────────────────────────────


def convert(lay_path: str, output_path: str) -> str:
    """`.lay` ファイルを Excel テンプレートに変換する。

    Args:
        lay_path: 入力 .lay ファイルのパス
        output_path: 出力 .xlsx ファイルのパス

    Returns:
        output_path
    """
    lay = parse_lay(lay_path)
    return convert_from_layfile(lay, output_path)


def convert_from_layfile(lay: LayFile, output_path: str) -> str:
    """パース済み LayFile を Excel テンプレートに変換する。

    Args:
        lay: パース済み LayFile
        output_path: 出力 .xlsx ファイルのパス

    Returns:
        output_path
    """
    # 1. 境界座標の収集・マージ
    x_bounds, y_bounds = collect_boundaries(lay)
    x_bounds = merge_close_boundaries(x_bounds)
    y_bounds = merge_close_boundaries(y_bounds)

    if len(x_bounds) < 2 or len(y_bounds) < 2:
        raise ValueError('座標データが不足しています')

    # 2. 寸法計算
    col_widths = calc_column_widths(x_bounds)
    row_heights = calc_row_heights(y_bounds)
    num_cols = len(col_widths)
    num_rows = len(row_heights)

    # 3. ワークブック作成
    wb = Workbook()
    ws = wb.active
    # タイトルから不要文字を除去
    title = lay.title.strip('\x00').strip()
    ws.title = title[:31] if title else 'テンプレート'

    # 4. 列幅・行高さ設定
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    for i, h in enumerate(row_heights):
        ws.row_dimensions[i + 1].height = h

    # 5. ラベル配置
    for obj in lay.labels:
        _place_label(ws, obj, x_bounds, y_bounds)

    # 6. フィールド配置
    for obj in lay.fields:
        _place_field(ws, obj, x_bounds, y_bounds)

    # 7. 罫線適用
    _apply_lines(ws, lay.lines, x_bounds, y_bounds, num_cols, num_rows)

    # 8. 印刷設定
    _apply_print_settings(ws, lay)

    # 9. 保存
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    return output_path


# ── スタンドアロン実行 ───────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python -m templates.generators.gen_from_lay INPUT.lay OUTPUT.xlsx')
        sys.exit(1)
    result = convert(sys.argv[1], sys.argv[2])
    print(f'Generated: {result}')
