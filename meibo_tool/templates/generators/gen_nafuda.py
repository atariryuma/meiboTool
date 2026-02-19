"""名札テンプレート生成スクリプト

SPEC.md §4.1〜§4.3 参照。
3 種類の名札テンプレートを生成する:
  '通常'   : 名札_通常.xlsx    — A4 横・2列×5行・10枚/ページ
  '装飾'   : 名札_装飾あり.xlsx — 名札_通常 + ピンク装飾枠
  '1年生'  : 名札_1年生用.xlsx  — A4 縦・8枚/ページ・かな縦書き

各カードのプレースホルダー（GridGenerator が _fill_numbered で置換）:
  {{出席番号_N}}, {{氏名かな_N}}, {{氏名_N}}  （N=1〜最大枚数）
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ── スタイル定数 ──────────────────────────────────────────────────────────

FONT_FAMILY = 'IPAmj明朝'

_THIN   = Side(style='thin',   color='000000')
_MEDIUM = Side(style='medium', color='000000')
_THICK  = Side(style='thick',  color='C06080')   # 装飾枠用

BORDER_THIN = Border(top=_THIN,   bottom=_THIN,   left=_THIN,   right=_THIN)
BORDER_CARD = Border(top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM, right=_MEDIUM)
BORDER_DECO = Border(top=_THICK,  bottom=_THICK,  left=_THICK,  right=_THICK)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_VCENTER = Alignment(horizontal='center', vertical='center')

FONT_NO       = Font(name=FONT_FAMILY, size=10)
FONT_KANA     = Font(name=FONT_FAMILY, size=12)
FONT_NAME     = Font(name=FONT_FAMILY, size=28, bold=True)
FONT_KANA_1NEN = Font(name=FONT_FAMILY, size=40, bold=True)
FONT_NO_1NEN  = Font(name=FONT_FAMILY, size=14)


def _ph(base: str, n: int) -> str:
    return '{{' + base + '_' + str(n) + '}}'


def _cell(ws, row: int, col: int, *, value=None, font=None, fill=None,
          border=None, alignment=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border
    if alignment is not None:
        c.alignment = alignment
    return c


def _apply_print(ws, orientation: str = 'landscape') -> None:
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.20, right=0.20, top=0.20, bottom=0.20, header=0.10, footer=0.10
    )
    ws.print_options.horizontalCentered = True


# ────────────────────────────────────────────────────────────────────────────
# 名札_通常 / 名札_装飾あり（A4 横・2列×5行・10枚/ページ）
# ────────────────────────────────────────────────────────────────────────────

def _build_normal(ws, decorated: bool = False) -> None:
    """
    カード構成（A4 横・10枚）:
      列: A(No左,4) B(かな左,20) C(氏名左,20) D(区切,2) E(No右,4) F(かな右,20) G(氏名右,20)
      行: 2行1組(かな行+氏名行) × 5組 = 10行

    左列: カード 1〜5（_1〜_5）
    右列: カード 6〜10（_6〜_10）
    """
    border = BORDER_DECO if decorated else BORDER_CARD

    ws.column_dimensions['A'].width = 4.5
    ws.column_dimensions['B'].width = 20.0
    ws.column_dimensions['C'].width = 22.0
    ws.column_dimensions['D'].width = 2.0
    ws.column_dimensions['E'].width = 4.5
    ws.column_dimensions['F'].width = 20.0
    ws.column_dimensions['G'].width = 22.0

    KANA_H = 16
    NAME_H = 46

    for i in range(5):
        ln = i + 1    # 左 No.1〜5
        rn = i + 6    # 右 No.6〜10
        kana_row = 1 + i * 2
        name_row = 2 + i * 2

        # ── 左カード ──────────────────────────────────────────────────────
        # 番号（出席番号）: かな行+氏名行をマージ
        ws.merge_cells(
            start_row=kana_row, start_column=1,
            end_row=name_row, end_column=1
        )
        _cell(ws, kana_row, 1,
              value=_ph('出席番号', ln), font=FONT_NO,
              border=border, alignment=ALIGN_CENTER)

        # かな行
        ws.merge_cells(
            start_row=kana_row, start_column=2,
            end_row=kana_row, end_column=3
        )
        _cell(ws, kana_row, 2,
              value=_ph('氏名かな', ln), font=FONT_KANA,
              border=border,
              alignment=Alignment(horizontal='center', vertical='bottom', wrap_text=True))

        # 氏名行
        ws.merge_cells(
            start_row=name_row, start_column=2,
            end_row=name_row, end_column=3
        )
        _cell(ws, name_row, 2,
              value=_ph('氏名', ln), font=FONT_NAME,
              border=border, alignment=ALIGN_CENTER)

        # 区切り
        _cell(ws, kana_row, 4)
        _cell(ws, name_row, 4)

        # ── 右カード ──────────────────────────────────────────────────────
        ws.merge_cells(
            start_row=kana_row, start_column=5,
            end_row=name_row, end_column=5
        )
        _cell(ws, kana_row, 5,
              value=_ph('出席番号', rn), font=FONT_NO,
              border=border, alignment=ALIGN_CENTER)

        ws.merge_cells(
            start_row=kana_row, start_column=6,
            end_row=kana_row, end_column=7
        )
        _cell(ws, kana_row, 6,
              value=_ph('氏名かな', rn), font=FONT_KANA,
              border=border,
              alignment=Alignment(horizontal='center', vertical='bottom', wrap_text=True))

        ws.merge_cells(
            start_row=name_row, start_column=6,
            end_row=name_row, end_column=7
        )
        _cell(ws, name_row, 6,
              value=_ph('氏名', rn), font=FONT_NAME,
              border=border, alignment=ALIGN_CENTER)

        ws.row_dimensions[kana_row].height = KANA_H
        ws.row_dimensions[name_row].height = NAME_H


# ────────────────────────────────────────────────────────────────────────────
# 名札_1年生用（A4 縦・8列・かな縦書き・8枚/ページ）
# ────────────────────────────────────────────────────────────────────────────

def _build_1nen(ws) -> None:
    """
    縦長短冊型: A4 縦に 8 枚を横並び。
    各短冊 = 1 列（幅 8）× 複数行:
      Row1: 番号（横書き）
      Row2: （空白）
      Row3〜5: かな（縦書き、textRotation=255、3行分マージ）

    列: A〜H（8 列） + I（区切り） の 9 列
    """
    VERT = Alignment(
        textRotation=255,  # 縦書き
        horizontal='center',
        vertical='center',
        wrap_text=True,
    )

    col_names = list('ABCDEFGH')
    col_widths = {c: 8.0 for c in col_names}
    col_widths['I'] = 1.0  # 余白
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 20   # 番号行
    ws.row_dimensions[2].height = 8    # 空白
    ws.row_dimensions[3].height = 80   # かな（縦書き）

    for i in range(8):
        n = i + 1
        col = i + 1  # 1-indexed

        # 番号行
        _cell(ws, 1, col,
              value=_ph('出席番号', n), font=FONT_NO_1NEN,
              border=BORDER_THIN,
              alignment=ALIGN_CENTER)

        # 空白行
        _cell(ws, 2, col, border=BORDER_THIN)

        # かな縦書き
        _cell(ws, 3, col,
              value=_ph('氏名かな', n), font=FONT_KANA_1NEN,
              border=BORDER_THIN, alignment=VERT)


# ────────────────────────────────────────────────────────────────────────────
# 公開 API
# ────────────────────────────────────────────────────────────────────────────

def generate(output_path: str, mode: str = '通常') -> None:
    """
    名札テンプレートを生成して output_path に保存する。

    mode:
      '通常'   — 名札_通常（2列×5行、装飾なし）
      '装飾'   — 名札_装飾あり（2列×5行、ピンク枠装飾）
      '1年生'  — 名札_1年生用（8列縦書き）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '名札'

    if mode == '1年生':
        _build_1nen(ws)
        _apply_print(ws, orientation='portrait')
    else:
        decorated = (mode == '装飾')
        _build_normal(ws, decorated=decorated)
        _apply_print(ws, orientation='landscape')

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f'Generated ({mode}): {output_path}')


# ────────────────────────────────────────────────────────────────────────────
# スタンドアロン実行
# ────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    _root = Path(__file__).resolve().parents[3]
    _tmpl = _root / 'テンプレート'
    generate(str(_tmpl / '名札_通常.xlsx'), mode='通常')
    generate(str(_tmpl / '名札_装飾あり.xlsx'), mode='装飾')
    generate(str(_tmpl / '名札_1年生用.xlsx'), mode='1年生')
