"""名札テンプレート集.xlsx → 個別プレースホルダーテンプレートに変換する

レガシーの全シート一体型テンプレート（Excel 数式で「基本入力」シートを参照する方式）を、
アプリのテンプレートエンジンが認識する {{placeholder}} 形式の個別ファイルに変換する。

変換ロジック:
  1. 基本入力!A{row} → {{出席番号_{row-1}}}
  2. 基本入力!B{row} → {{氏名_{row-1}}}
  3. CONCATENATE(A,B) → {{出席番号_N}}{{氏名_N}}
  4. シート内参照 (=C4) → 参照先のプレースホルダーに解決
  5. 名前プレースホルダーと同じ行の静的整数 → {{出席番号_N}}
  6. 残りの壊れた数式はクリア

使い方:
    cd meibo_tool && python -m templates.generators.convert_legacy_xlsx

入力:
    テンプレート/名札テンプレート集.xlsx

出力:
    テンプレート/ 以下に個別 .xlsx ファイル（9種）
"""

from __future__ import annotations

import os
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell as _MergedCell
from openpyxl.worksheet.dimensions import ColumnDimension

# 日本語 Excel の phonetic 属性に対する openpyxl 互換パッチ
_orig_cd_init = ColumnDimension.__init__


def _patched_cd_init(self, *args, **kwargs):
    kwargs.pop('phonetic', None)
    _orig_cd_init(self, *args, **kwargs)


ColumnDimension.__init__ = _patched_cd_init

# 基本入力 シートへの参照パターン（$付き/なし両対応）
_KIHON_REF = re.compile(r'基本入力!\$?([A-Z])\$?(\d+)')

# シート名 → 出力ファイル名
_SHEET_MAP: dict[str, str] = {
    '掲示用名列表': '掲示用名列表.xlsx',
    '横名簿': '横名簿.xlsx',
    '縦一週間': '縦一週間.xlsx',
    '男女一覧': '男女一覧.xlsx',
    '調べ物表': '調べ表.xlsx',
    'ラベル 小': 'ラベル_小.xlsx',
    'ラベル 大': 'ラベル_大2.xlsx',
    'ラベル 特大': 'ラベル_特大.xlsx',
    'ラベル大色付': '名札_装飾あり.xlsx',
}


def _formula_to_placeholder(formula: str) -> str | None:
    """基本入力 参照の数式を {{placeholder}} に変換する。

    例: =IF(基本入力!B2="","",基本入力!B2) → {{氏名_1}}
        =CONCATENATE(基本入力!$A2,基本入力!$B2) → {{出席番号_1}}{{氏名_1}}
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return None

    refs = _KIHON_REF.findall(formula)
    if not refs:
        return None

    # 生徒番号ごとにカラム（A=番号, B=名前）を集約
    students: dict[int, dict[str, bool]] = {}
    for col, row_str in refs:
        n = int(row_str) - 1  # 基本入力 row 2 = 生徒1
        if n < 1 or n > 40:
            continue
        if col == 'A':
            students.setdefault(n, {})['num'] = True
        elif col == 'B':
            students.setdefault(n, {})['name'] = True

    if not students:
        return None

    parts: list[str] = []
    for n in sorted(students.keys()):
        info = students[n]
        if info.get('num'):
            parts.append(f'{{{{出席番号_{n}}}}}')
        if info.get('name'):
            parts.append(f'{{{{氏名_{n}}}}}')

    return ''.join(parts) if parts else None


def _resolve_cell_ref(formula: str, ws) -> str | None:
    """シート内セル参照（=C4 など）を参照先の値に解決する。

    参照先がプレースホルダー文字列の場合のみ返す。
    """
    if not formula or not formula.startswith('='):
        return None

    m = re.match(r'^=([A-Z]+)(\d+)$', formula.strip())
    if not m:
        return None

    coord = f'{m.group(1)}{m.group(2)}'
    try:
        cell = ws[coord]
        if isinstance(cell, _MergedCell):
            return None
        val = cell.value
        if val and isinstance(val, str) and '{{' in val:
            return val
    except Exception:
        pass

    return None


def convert_sheet(legacy_path: str, sheet_name: str, output_path: str) -> int:
    """レガシーファイルの1シートをプレースホルダーテンプレートに変換する。

    Returns:
        変換されたプレースホルダー数
    """
    wb = load_workbook(legacy_path)

    # 対象シート以外を削除
    for name in wb.sheetnames:
        if name != sheet_name:
            del wb[name]

    # 残ったシートを確実にアクティブにする
    ws = wb.worksheets[0]
    wb.active = 0
    converted = 0

    # Pass 1: 基本入力 参照をプレースホルダーに変換
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, _MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                placeholder = _formula_to_placeholder(cell.value)
                if placeholder:
                    cell.value = placeholder
                    converted += 1

    # Pass 2: シート内セル参照を解決（=C4 → プレースホルダー）
    # 参照チェーンに対応するため変化がなくなるまで繰り返す
    max_iterations = 5
    for _ in range(max_iterations):
        changed = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, _MergedCell):
                    continue
                if (cell.value and isinstance(cell.value, str)
                        and cell.value.startswith('=')):
                    resolved = _resolve_cell_ref(cell.value, ws)
                    if resolved:
                        cell.value = resolved
                        converted += 1
                        changed = True
        if not changed:
            break

    # Pass 3: 名前プレースホルダーと同じ行の静的整数を出席番号に変換
    _name_ph_re = re.compile(r'\{\{氏名_(\d+)\}\}')
    for row in ws.iter_rows():
        name_students: set[int] = set()
        for cell in row:
            if isinstance(cell, _MergedCell):
                continue
            if cell.value and isinstance(cell.value, str):
                for match in _name_ph_re.finditer(cell.value):
                    name_students.add(int(match.group(1)))

        if name_students:
            for cell in row:
                if isinstance(cell, _MergedCell):
                    continue
                if isinstance(cell.value, (int, float)):
                    n = int(cell.value)
                    if n in name_students:
                        cell.value = f'{{{{出席番号_{n}}}}}'
                        converted += 1

    # Pass 4: 残りの壊れた数式をクリア（削除されたシートへの参照等）
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, _MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = ''

    # 保存
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()
    return converted


def generate(template_dir: str, legacy_path: str) -> list[str]:
    """全シートを変換する。

    Returns:
        生成されたファイルパスのリスト
    """
    wb = load_workbook(legacy_path, data_only=False)
    sheet_names = wb.sheetnames
    wb.close()

    # シート名のスペース差異に対応（例: "縦一週間 " → "縦一週間"）
    stripped_to_actual = {name.strip(): name for name in sheet_names}

    generated: list[str] = []

    for sheet_name, output_file in _SHEET_MAP.items():
        actual_name = stripped_to_actual.get(sheet_name)
        if actual_name is None:
            print(f'  SKIP (not found): {sheet_name}')
            continue

        output_path = os.path.join(template_dir, output_file)
        try:
            count = convert_sheet(legacy_path, actual_name, output_path)
            print(f'  OK: {sheet_name} → {output_file} ({count} placeholders)')
            generated.append(output_path)
        except Exception as e:
            print(f'  ERROR [{sheet_name}]: {e}')

    return generated


if __name__ == '__main__':
    _root = Path(__file__).resolve().parents[3]
    _legacy = _root / 'テンプレート' / '名札テンプレート集.xlsx'
    _tmpl = _root / 'テンプレート'

    if not _legacy.exists():
        print(f'ERROR: {_legacy} が見つかりません')
        raise SystemExit(1)

    print(f'入力: {_legacy}')
    print(f'出力先: {_tmpl}')
    result = generate(str(_tmpl), str(_legacy))
    print(f'\n完了: {len(result)} ファイル生成')
