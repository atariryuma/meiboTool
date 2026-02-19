"""修了台帳・卒業台帳テンプレート生成スクリプト

SPEC.md §4.6〜§4.7 参照。
どちらも A4 横置き・List 型（ListGenerator がデータ行を展開する）。

修了台帳 (mode='shuuryo'):
  タイトル: 「{{年度和暦}}年度 {{学年}}年{{組}}組 修了台帳　担任：{{担任名}}」
  ヘッダー列: 学校をまった日|番|正式氏名|正式氏名かな|性別|生年月日|保護者正式名|住所|電話番号1

卒業台帳 (mode='sotsugyou'):
  タイトル: 「{{年度和暦}}年度 卒業生名簿　{{学校名}}」
  ヘッダー列: 証書番号|正式氏名|正式氏名かな|生年月日|住所|保護者正式名|進学先|担任名
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ── スタイル定数 ──────────────────────────────────────────────────────────

FONT_FAMILY = 'IPAmj明朝'
_HEADER_BG  = 'F0F0F0'   # カラムヘッダーのみ薄グレー（データ行は白/塗りなし）

FILL_HEADER = PatternFill(fill_type='solid', fgColor=_HEADER_BG)

_THIN = Side(style='thin', color='000000')

BORDER_THIN = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left',   vertical='center', wrap_text=True)

FONT_TITLE  = Font(name=FONT_FAMILY, size=14, bold=True)
FONT_HEADER = Font(name=FONT_FAMILY, size=8,  bold=True)
FONT_DATA   = Font(name=FONT_FAMILY, size=8)


def _cell(ws, row: int, col: int, *, value=None, font=None, fill=None,
          border=None, alignment=None, width=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border
    if alignment is not None:
        c.alignment = alignment
    if width is not None:
        import openpyxl.utils as utils
        ws.column_dimensions[utils.get_column_letter(col)].width = width
    return c


def _apply_print(ws) -> None:
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.20, right=0.20, top=0.20, bottom=0.20, header=0.10, footer=0.10
    )


# ────────────────────────────────────────────────────────────────────────────
# 修了台帳
# ────────────────────────────────────────────────────────────────────────────

# (列ラベル, プレースホルダー, 幅)
_SHUURYO_COLS = [
    ('学校をまった日', '{{転入日}}',       12),
    ('番',             '{{出席番号}}',      5),
    ('正式氏名',       '{{正式氏名}}',      18),
    ('正式氏名かな',   '{{正式氏名かな}}',  18),
    ('性別',           '{{性別}}',          5),
    ('生年月日',       '{{生年月日}}',      13),
    ('保護者正式名',   '{{保護者正式名}}',  18),
    ('住所',           '{{住所}}',          30),
    ('電話番号',       '{{電話番号1}}',     14),
]

_SOTSUGYOU_COLS = [
    ('証書番号',       '{{証書番号}}',      10),
    ('正式氏名',       '{{正式氏名}}',      18),
    ('正式氏名かな',   '{{正式氏名かな}}',  18),
    ('生年月日',       '{{生年月日}}',      13),
    ('住所',           '{{住所}}',          30),
    ('保護者正式名',   '{{保護者正式名}}',  18),
    ('進学先',         '{{進学先}}',        18),
    ('担任名',         '{{担任名}}',        12),
]


def _build_daicho(ws, title_value: str, col_specs: list) -> None:
    """台帳テンプレート共通レイアウト。"""
    total_cols = len(col_specs)

    # タイトル行（塗りなし・大字のみで視認性を確保）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    _cell(ws, 1, 1,
          value=title_value, font=FONT_TITLE,
          alignment=ALIGN_C)
    ws.row_dimensions[1].height = 30

    # ヘッダー行（カラムラベルのみ薄グレー）
    for col_idx, (label, _ph, width) in enumerate(col_specs, 1):
        _cell(ws, 2, col_idx,
              value=label, font=FONT_HEADER,
              fill=FILL_HEADER, border=BORDER_THIN, alignment=ALIGN_C,
              width=width)
    ws.row_dimensions[2].height = 20

    # データ行（ListGenerator がここを複製する・塗りなし）
    for col_idx, (_, placeholder, _) in enumerate(col_specs, 1):
        align = ALIGN_C if col_idx <= 2 else ALIGN_L
        _cell(ws, 3, col_idx,
              value=placeholder, font=FONT_DATA,
              border=BORDER_THIN, alignment=align)
    ws.row_dimensions[3].height = 16


# ────────────────────────────────────────────────────────────────────────────
# 公開 API
# ────────────────────────────────────────────────────────────────────────────

def generate(output_path: str, mode: str = 'shuuryo') -> None:
    """
    台帳テンプレートを生成して output_path に保存する。

    mode:
      'shuuryo'   — 修了台帳
      'sotsugyou' — 卒業台帳
    """
    wb = Workbook()
    ws = wb.active

    if mode == 'sotsugyou':
        ws.title = '卒業台帳'
        title = '{{年度和暦}}年度  卒業生名簿　　{{学校名}}'
        _build_daicho(ws, title, _SOTSUGYOU_COLS)
    else:  # shuuryo
        ws.title = '修了台帳'
        title = '{{年度和暦}}年度  {{学年}}年{{組}}組  修了台帳　　担任：{{担任名}}'
        _build_daicho(ws, title, _SHUURYO_COLS)

    _apply_print(ws)

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f'Generated ({mode}): {output_path}')


if __name__ == '__main__':
    _root = Path(__file__).resolve().parents[3]
    _tmpl = _root / 'テンプレート'
    generate(str(_tmpl / '修了台帳.xlsx'), mode='shuuryo')
    generate(str(_tmpl / '卒業台帳.xlsx'), mode='sotsugyou')
