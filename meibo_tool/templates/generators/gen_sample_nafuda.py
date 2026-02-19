"""手作りテンプレートのサンプル生成スクリプト

Excelで直接テンプレートを作る際の参考用に、最小限の名札テンプレートを生成する。
生成後はExcelで開いて自由に調整可能。

使い方:
    cd meibo_tool && python -m templates.generators.gen_sample_nafuda

生成ファイル:
    テンプレート/サンプル_名札.xlsx

このスクリプトは「Excelで手作りするとこうなる」という参考実装。
実際の運用では、Excelで直接作成・編集してテンプレートフォルダに保存するだけでOK。
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ── スタイル ──────────────────────────────────────────────────────────────

FONT_FAMILY = 'IPAmj明朝'
_MEDIUM = Side(style='medium', color='000000')
BORDER = Border(top=_MEDIUM, bottom=_MEDIUM, left=_MEDIUM, right=_MEDIUM)
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _ph(base: str, n: int) -> str:
    """プレースホルダー文字列を生成する。"""
    return '{{' + base + '_' + str(n) + '}}'


def generate(output_path: str) -> None:
    """シンプルな名札テンプレート（A4横・2列×5行=10枚/ページ）を生成する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = '名札'

    # ── 列幅 ────────────────────────────────────────────────────────────
    # A: 出席番号(左)  B: 氏名(左)  C: 区切り  D: 出席番号(右)  E: 氏名(右)
    ws.column_dimensions['A'].width = 6.0
    ws.column_dimensions['B'].width = 30.0
    ws.column_dimensions['C'].width = 2.0
    ws.column_dimensions['D'].width = 6.0
    ws.column_dimensions['E'].width = 30.0

    # ── カード配置（2列×5行=10枚） ────────────────────────────────────
    for i in range(5):
        left_n = i + 1      # 左列: 1〜5
        right_n = i + 6     # 右列: 6〜10
        row = i + 1

        ws.row_dimensions[row].height = 55   # 各カード行高

        # 左カード: 出席番号
        c = ws.cell(row=row, column=1)
        c.value = _ph('出席番号', left_n)
        c.font = Font(name=FONT_FAMILY, size=12)
        c.border = BORDER
        c.alignment = ALIGN_C

        # 左カード: 氏名
        c = ws.cell(row=row, column=2)
        c.value = _ph('氏名', left_n)
        c.font = Font(name=FONT_FAMILY, size=28, bold=True)
        c.border = BORDER
        c.alignment = ALIGN_C

        # 右カード: 出席番号
        c = ws.cell(row=row, column=4)
        c.value = _ph('出席番号', right_n)
        c.font = Font(name=FONT_FAMILY, size=12)
        c.border = BORDER
        c.alignment = ALIGN_C

        # 右カード: 氏名
        c = ws.cell(row=row, column=5)
        c.value = _ph('氏名', right_n)
        c.font = Font(name=FONT_FAMILY, size=28, bold=True)
        c.border = BORDER
        c.alignment = ALIGN_C

    # ── 印刷設定（Excelで「ページレイアウト」から設定する内容と同等） ──
    ws.page_setup.paperSize = 9        # A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.20, right=0.20,
        top=0.20, bottom=0.20,
        header=0.10, footer=0.10,
    )
    ws.print_options.horizontalCentered = True

    # ── 保存 ────────────────────────────────────────────────────────────
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f'Generated sample: {output_path}')


if __name__ == '__main__':
    _root = Path(__file__).resolve().parents[3]
    generate(str(_root / 'テンプレート' / 'サンプル_名札.xlsx'))
